import tkinter as tk
from tkinter import Tk, Canvas, Text, Button, PhotoImage, messagebox, filedialog, ttk
from tkcalendar import DateEntry
import webbrowser
import os
import csv
import chardet
import pandas as pd
import openpyxl
import pyodbc
import camelot
from pdf2image import convert_from_path
import base64
from io import BytesIO
import tempfile
from PIL import ImageTk, Image
import datetime
import uuid





# CRIAR O EXE
# pip install --upgrade pyinstaller
# pip install --upgrade sqlalchemy
# pyinstaller --onefile --windowed --icon=C:\Users\Administrador\Desktop\python\BOLA-LUZITIC.ico --hidden-import babel.numbers gui.py



def load_base64_image(base64_data):
    return PhotoImage(data=base64_data)


def abrir_email():
    webbrowser.open('mailto:suporte@luzitic.pt')


def fechar_aplicacao():
    window.destroy()


def processar_arquivo_PorFazer():
    messagebox.showinfo('Configuração com Modelo de Importação',
                        'A importação que escolheu tem um modelo de importção configurado')


def processar_arquivo_ALTICE():
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.csv')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        def obter_valor():
            valorMES = entry3.get_date()
            nome_arquivo, extensao = os.path.splitext(
                os.path.basename(filename))
            with open(filename, newline='') as f_in, open('C:\\importacao\\' + nome_arquivo + '.csv', 'w', newline='') as f_out:

                # Criar um leitor CSV e um gravador CSV
                leitor = csv.reader(f_in, delimiter=';')
                gravador = csv.writer(f_out, delimiter=';')

                rows = list(leitor)[6:-3]
                gravador.writerows(rows)
            with open('C:\\importacao\\' + nome_arquivo + '.csv', 'rb') as f:
                result = chardet.detect(f.read())
                encoding = result['encoding']

            df = pd.read_csv('C:\\importacao\\' + nome_arquivo +
                             '.csv', encoding=encoding, delimiter=";")

            df.to_excel('C:\\importacao\\' + nome_arquivo +
                        '.xlsx', index=False)

            os.remove('C:\\importacao\\' + nome_arquivo + '.csv')
            # Exibir uma mensagem de conclusão
            df = pd.read_excel('C:\\importacao\\' + nome_arquivo +
                               '.xlsx')

            preco = df.loc[0, "Plano de Preços"]

            num_rows = df.shape[0]
            num_rows2 = float(num_rows + 1)
            numero = float(preco.replace(',', '.'))
            media = numero/num_rows2

            df["VALOR MENSALIDADE"] = media
            valor = df["Valor (s/IVA)"].str.replace(',', '.').astype(float)
            df['VALOR DO CARTAO'] = media + valor
            df['valorIva'] = (df['VALOR DO CARTAO']* 0.23)

            df['DATA FATURA'] = valorMES

            df = df.iloc[1:]
            df.to_excel('C:\\importacao\\' + nome_arquivo +
                        '.xlsx', index=False)
            server = 'SBs2019-ISAAC\ABMN'
            database = 'aTrans'
            username = 'Bds'
            password = 'olivettiBDS1'

            # Criação da conexão com o SQL Server
            conn = pyodbc.connect(
                'DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+password)
            # Create a DataFrame to store the query results
            for index, row in df.iterrows():
                ntele = str(row[2])

                # Create the SQL query
                sql = f"SELECT id FROM Motoristas WHERE IdCartaoTlm = (SELECT Id FROM CartoesTlm WHERE NumCartao = '{ntele}')"
                # sql = f"SELECT top 1 matricula FROM viaturas WHERE idMotorista = (SELECT nome FROM Motoristas WHERE IdCartaoTlm = (SELECT Id FROM CartoesTlm WHERE NumCartao = '{ntele}'))"

                # Execute the query and retrieve the results
                df_query = pd.read_sql_query(sql, conn)
                # Check if df_query has any results
                if not df_query.empty:
                    # Extract the value from the 'nome' column of df_query
                    id = df_query['id'].iloc[0]
                else:
                    id = 'B46D277F-9A32-43BF-95DE-AB0D2016E75E'  # Define um valor vazio caso não haja resultados

                # Assign the extracted value to the 'nome' column of df
                df.at[index, 'idMotorista'] = id
                # Fechar a conexão com o banco de dados
            df.to_excel('C:\\importacao\\' + nome_arquivo +
                        '.xlsx', index=False)
            
            dataframe = pd.read_excel('C:\\importacao\\' + nome_arquivo +
                        '.xlsx'
)
            # Iterando pelas linhas e inserindo no banco de dados
            i=0
            ficheiroID = uuid.uuid4()
            transImportacaoFicheirID = uuid.uuid4()
            for index, row in dataframe.iterrows():
                # Extrair os valores de cada coluna do DataFrame
                # Custo = row['nome_da_coluna2']

                ValorIva = row['valorIva']
                TotalCusto = row['VALOR DO CARTAO']
                Valor = row['VALOR DO CARTAO']
                IdDespesa = '44309AE2-B46C-49D0-9252-B3D92F7CFD09'
                Descricao = 'Comunicações MN Iva Taxa Normal'
                Totaldesconto = 0
                IdMotorista = row['idMotorista']
                Quantidade = 1
                NdocDespesa = row['Nº de Factura']
                IdFornecedor = '5AA9BFA4-6A03-4053-BF52-D0417B53F952'
                IdPais = 'C752C51A-29D6-4E3C-AA56-FC2845ABE599'
                IdUtilizadorCriacao = '281B923E-09A7-4B73-805C-ABBEEF62C162'
                IdIVA = 'A1D441A1-D5AF-49B0-85E4-22D8B43CAEDD'
                DataDespesa = row['DATA FATURA']

                # Construindo a consulta SQL de inserção
                LinhasDespesas = "INSERT INTO dbo.LinhasDespesas ([IdCabecDespesa], [IdFolhaServico], [IdViagem], [IdViatura], [IdMotorista], [IdOrdemCarga], [IdCmr], [IdReboque], [TotalCusto], [IdDespesa], [Descricao], [Valor], [Quantidade], [IdUser], [FechadoDespesas], [IdUserDespesas], [FechadoSubsidios], [IdUserSubsidios], [ExportadoERP], [IdUserExportadoERP], [IdOperador], [Afaturar], [Faturado], [NdocDespesa], [Idcabecdoc], [ARHP], [RHP], [Custo], [IdCombustivel], [Data], [AdBlueA], [AdblueNA], [IdFornecedor], [Km], [Litros], [PostoAbastecimento], [IdPais], [Paga], [Conferida], TotalDesconto, Observacoes, IdUtilizadorCriacao, IdIVA, ValorIva, TipoDocCCT, SerieCCT, NumDocCCT, IdLinhaRH,SujeitoPassivo,Projeto,TipoOperacao, [IdEquipamento],[IdOrdemReparacao], [IdDepartamento], [IdSinistro], [DataDespesa], [IdDespesaProvisional], [IdLinhaDespConvert], [IdArmazem],[IdProjecto],[CCusto],IdaWarehouseArmazem, IdImportacaoFicheiro, idArtigo) VALUES ( NULL,NULL,NULL,NULL,'"+str(IdMotorista)+"',NULL,NULL,NULL,'"+str(TotalCusto)+"','"+str(IdDespesa)+"','"+str(Descricao)+"','"+str(Valor)+"','"+str(Quantidade)+"',NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,'"+str(NdocDespesa)+"',NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,'"+str(IdFornecedor)+"',NULL,NULL,NULL,'"+str(IdPais)+"',NULL,NULL,'"+str(Totaldesconto)+"',NULL,'"+str(IdUtilizadorCriacao)+"','"+str(IdIVA)+"','"+str(ValorIva)+"',NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,NULL,'"+str(DataDespesa)+"',NULL,NULL,NULL,NULL,NULL,NULL,'"+str(ficheiroID)+"','"+str(IdDespesa)+"')"
                
                # Executando a consulta SQL
                print(LinhasDespesas)
                cursor.execute(LinhasDespesas)
                cursor.commit()
                nLinhas=i+1

            today = datetime.date.today()
            

            aPlatImportadoFicheiro = "INSERT INTO dbo.aPlatImportadoFicheiro (Id,IdImportacaoConfBase, NomeFicheiro, IdUtilizadorCriacao) values ('"+str(ficheiroID)+"','B0173A17-6F62-4192-BF3E-A3789B99995E', '"+str(nome_arquivo)+".xlsx', '281B923E-09A7-4B73-805C-ABBEEF62C162')"
            print(aPlatImportadoFicheiro)
            cursor.execute(aPlatImportadoFicheiro)
            cursor.commit()
            # Commitar as alterações no banco de dados e fechar a conexão
            
            transImportacaoFicheiro = "INSERT INTO [TransImportacaoFicheiro] ([Id],[IdFicheiro],[IdFornecedor],[NumDoc],[DataDoc],[DtCriacao],[IdUtilizadorCriacao],[DtUltimaAlteracao],[IdUtilizadorUltimaAlteracao],[Activo]) VALUES('"+str(transImportacaoFicheirID)+"' ,'"+str(ficheiroID)+"' ,'5AA9BFA4-6A03-4053-BF52-D0417B53F952', '"+str(nome_arquivo)+"' ,'"+str(valorMES)+"' ,'"+str(today)+"','E93D1F9B-1D32-4579-AC3E-C4A2506AC345' ,'"+str(today)+"' ,NULL,1)"
            print(transImportacaoFicheiro)
            cursor.execute(transImportacaoFicheiro)
            cursor.commit()

            aPlatImportadoFicheiroResumo = "INSERT INTO [aPlatImportadoFicheiroResumo] ([IdImportadoFicheiro] ,[QtdTotal] ,[QtdIgnorada]   ,[QtdFail]  ,[QtdImportada]  ,[DtFimImportacao]  ,[DtCriacao]  ,[IdUtilizadorCriacao]  ,[DtUltimaAlteracao]  ,[IdUtilizadorUltimaAlteracao],[Activo])     VALUES  ('"+str(ficheiroID)+"' ,'"+str(nLinhas)+"' ,0  ,0  , '"+str(nLinhas)+"' ,'"+str(today)+"'  ,'"+str(today)+"'  ,'281B923E-09A7-4B73-805C-ABBEEF62C162'  ,'"+str(today)+"','281B923E-09A7-4B73-805C-ABBEEF62C162'  ,1)"

            print(aPlatImportadoFicheiroResumo)
            cursor.execute(aPlatImportadoFicheiroResumo)
            cursor.commit()
            cursor.close()

            messagebox.showinfo(
                'Concluído', 'O arquivo foi processado com sucesso.')
            root.destroy()

        def sair():
            root.destroy()

        # Criar janela principal
        root = tk.Tk()

        root.geometry("400x200")
        root.configure(bg="#3A7FF6")

        canvas2 = tk.Canvas(
            root,
            bg="#3A7FF6",
            height=200,
            width=400,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )

        canvas2.place(x=0, y=0)
        canvas2.create_rectangle(
            200,
            0.0,
            400,
            200,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            15,
            anchor="nw",
            text="Importador de Despesas",
            fill="#FCFCFC",
            font=("Roboto Bold", 12)
        )

        canvas2.create_text(
            220,
            15,
            anchor="nw",
            text="Data da Fatura:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        canvas2.create_rectangle(
            15,
            35,
            80,
            40,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            60,
            anchor="nw",
            text="Aqui é necessário a introdução ",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            80,
            anchor="nw",
            text="da data da fatura",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            100,
            anchor="nw",
            text="correspondente a importação.",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        entry3 = DateEntry(root, selectmode="day", date_pattern='yyyy-mm-dd')
        entry3.config(width=25)
        entry3.place(x=220, y=40)

        btn_obter_valor = tk.Button(
            root, text="Enviar dados", command=obter_valor)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=100)

        btn_obter_valor = Button(
            root, text="Sair do Ecrã", command=sair)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=150)

        canvas2.create_text(
            39.999999999999886,
            340.0,
            anchor="nw",
            text="Aplicação de suporte a importação de despesas",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 14 * -1)
        )

        # Código da imagem em base64
        codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

        # Decodifica o código base64 para dados binários
        dados_imagem_icon_luzitic = base64.b64decode(
            codigo_base64_icon_luzitic)
        buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
        imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

        # Converte a imagem para o formato .ico
        icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
        imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

        # Define o ícone da janela
        root.iconbitmap(default=icone_temp.name)

       # Definir o título da janela
        root.title('Importação de Despesas')

        root.resizable(False, False)
        root.mainloop()


def processar_arquivo_Toll_Collect():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.csv')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        def obter_valor():
            valorFaturaEntry = entry.get("1.0", "end-1c")
            nome_arquivo, extensao = os.path.splitext(
                os.path.basename(filename))
            # Abrir o arquivo CSV para leitura
            with open(filename, newline='') as f_in, open('C:\\importacao\\' + nome_arquivo + '.csv', 'w', newline='') as f_out:
                # Criar um leitor CSV e um gravador CSV
                # Ler as linhas do arquivo CSV
                linhas = list(csv.reader(f_in))
                gravador = csv.writer(f_out)

                linhas = linhas[:-3]
                gravador.writerows(linhas)

            with open('C:\\importacao\\' + nome_arquivo + '.csv', 'rb') as f:
                result = chardet.detect(f.read())
                encoding = result['encoding']

            df = pd.read_csv('C:\\importacao\\' + nome_arquivo +
                             '.csv', encoding=encoding, delimiter=";")

            df.to_excel('C:\\importacao\\' + nome_arquivo +
                        '.xlsx', index=False)

            os.remove('C:\\importacao\\' + nome_arquivo + '.csv')

            df = pd.read_excel('C:\\importacao\\' + nome_arquivo +
                               '.xlsx')
            df["FATURA"] = valorFaturaEntry
            df.to_excel('C:\\importacao\\' + nome_arquivo +
                        '.xlsx', index=False)

            messagebox.showinfo(
                'Concluído', 'O arquivo foi processado com sucesso.')
            root.destroy()

        def sair():
            root.destroy()

        # Criar janela principal
        root = tk.Tk()

        root.geometry("400x200")
        root.configure(bg="#3A7FF6")

        canvas2 = tk.Canvas(
            root,
            bg="#3A7FF6",
            height=200,
            width=400,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )

        canvas2.place(x=0, y=0)
        canvas2.create_rectangle(
            200,
            0.0,
            400,
            200,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            15,
            anchor="nw",
            text="Importador de Despesas",
            fill="#FCFCFC",
            font=("Roboto Bold", 12)
        )

        canvas2.create_text(
            220,
            15,
            anchor="nw",
            text="Número da Fatura:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        canvas2.create_rectangle(
            15,
            35,
            80,
            40,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            60,
            anchor="nw",
            text="Aqui é necessário a introdução ",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            80,
            anchor="nw",
            text="do numero da fatura",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            100,
            anchor="nw",
            text="correspondente a importação.",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        entry = Text(root)
        entry.config(width=20, height=2)
        entry.place(x=220, y=40)
        entry.config(borderwidth=1, relief="solid")

        btn_obter_valor = tk.Button(
            root, text="Enviar dados", command=obter_valor)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=100)

        btn_obter_valor = Button(
            root, text="Sair do Ecrã", command=sair)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=150)

        canvas2.create_text(
            39.999999999999886,
            340.0,
            anchor="nw",
            text="Aplicação de suporte a importação de despesas",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 14 * -1)
        )

        # Código da imagem em base64
        codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

        # Decodifica o código base64 para dados binários
        dados_imagem_icon_luzitic = base64.b64decode(
            codigo_base64_icon_luzitic)
        buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
        imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

        # Converte a imagem para o formato .ico
        icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
        imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

        # Define o ícone da janela
        root.iconbitmap(default=icone_temp.name)

       # Definir o título da janela
        root.title('Importação de Despesas')

        root.resizable(False, False)
        root.mainloop()


def processar_arquivo_NORPETROL():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.csv')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        def obter_valor():
            valorFaturaEntry = entry.get("1.0", "end-1c")
            nome_arquivo, extensao = os.path.splitext(
                os.path.basename(filename))
            with open(filename, newline='') as f_in, open('C:\\importacao\\' + nome_arquivo + '.csv', 'w', newline='') as f_out:

                reader = csv.reader(f_in, delimiter=';')
                writer = csv.writer(f_out, delimiter=';')

                # Percorrer cada linha do arquivo de entrada
                for row in reader:
                    # Substituir todos os pontos por vírgulas na coluna desejada
                    row[4] = row[4].replace('.', ',')
                    row[6] = row[6].replace('.', '')
                    row[7] = row[7].replace('.', '')
                    row[8] = row[8].replace('.', '')
                    row[9] = row[9].replace('.', '')
                    row[10] = row[10].replace('.', '')
                    row[11] = row[11].replace('.', '')
                    row[12] = valorFaturaEntry
                    # Escrever a linha modificada no arquivo de saída
                    writer.writerow(row)

            # Exibir uma mensagem de conclusão
            messagebox.showinfo(
                'Concluído', 'O arquivo foi processado com sucesso.')
            root.destroy()

        def sair():
            root.destroy()
        # Criar janela principal
        root = tk.Tk()

        root.geometry("400x200")
        root.configure(bg="#3A7FF6")

        canvas2 = tk.Canvas(
            root,
            bg="#3A7FF6",
            height=200,
            width=400,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )

        canvas2.place(x=0, y=0)
        canvas2.create_rectangle(
            200,
            0.0,
            400,
            200,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            15,
            anchor="nw",
            text="Importador de Despesas",
            fill="#FCFCFC",
            font=("Roboto Bold", 12)
        )

        canvas2.create_text(
            220,
            15,
            anchor="nw",
            text="Número da Fatura:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        canvas2.create_rectangle(
            15,
            35,
            80,
            40,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            60,
            anchor="nw",
            text="Aqui é necessário a introdução ",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            80,
            anchor="nw",
            text="do numero da fatura",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            100,
            anchor="nw",
            text="correspondente a importação.",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        entry = Text(root)
        entry.config(width=20, height=2)
        entry.place(x=220, y=40)
        entry.config(borderwidth=1, relief="solid")

        btn_obter_valor = tk.Button(
            root, text="Enviar dados", command=obter_valor)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=100)

        btn_obter_valor = Button(
            root, text="Sair do Ecrã", command=sair)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=150)

        canvas2.create_text(
            39.999999999999886,
            340.0,
            anchor="nw",
            text="Aplicação de suporte a importação de despesas",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 14 * -1)
        )

        # Código da imagem em base64
        codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

        # Decodifica o código base64 para dados binários
        dados_imagem_icon_luzitic = base64.b64decode(
            codigo_base64_icon_luzitic)
        buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
        imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

        # Converte a imagem para o formato .ico
        icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
        imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

        # Define o ícone da janela
        root.iconbitmap(default=icone_temp.name)

        # Definir o título da janela
        root.title('Importação de Despesas')

        root.resizable(False, False)
        root.mainloop()


def processar_arquivo_VIAVERDE():

    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.csv')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        def obter_valor():
            valorFaturaEntry = entry.get("1.0", "end-1c")
            nome_arquivo, extensao = os.path.splitext(
                os.path.basename(filename))
            with open(filename, newline='') as f_in, open('C:\\importacao\\' + nome_arquivo + '.csv', 'w', newline='') as f_out:

                # Criar um leitor CSV e um gravador CSV
                leitor = csv.reader(f_in)
                gravador = csv.writer(f_out)

                # Iterar sobre as linhas do arquivo original e escrever no arquivo de saída
                for i, linha in enumerate(leitor):
                    if i >= 7:  # Excluir as primeiras sete linhas
                        gravador.writerow(linha)
                # Ler o arquivo CSV
            with open('C:\\importacao\\' + nome_arquivo + '.csv', 'rb') as f:
                result = chardet.detect(f.read())
                encoding = result['encoding']

            df = pd.read_csv('C:\\importacao\\' + nome_arquivo +
                             '.csv', encoding=encoding, delimiter=";")
            df['FATURA'] = valorFaturaEntry
            df['DATA ENTRADA'] = pd.to_datetime(
                df['DATA ENTRADA'], errors='coerce', dayfirst=False)

            # Formatar novamente a coluna no formato desejado
            df.loc[:, 'DATA ENTRADA'] = df['DATA ENTRADA'].dt.strftime(
                '%Y-%m-%d')

            df['DATA SAÍDA'] = pd.to_datetime(
                df['DATA SAÍDA'], errors='coerce', dayfirst=False)

            # Formatar novamente a coluna no formato desejado
            df.loc[:, 'DATA SAÍDA'] = df['DATA SAÍDA'].dt.strftime('%Y-%m-%d')

            df['DATA PAGAMENTO'] = pd.to_datetime(
                df['DATA PAGAMENTO'], errors='coerce', dayfirst=False)

            # Formatar novamente a coluna no formato desejado
            df.loc[:, 'DATA PAGAMENTO'] = df['DATA PAGAMENTO'].dt.strftime(
                '%Y-%m-%d')

            dados1 = df.loc[(df['OPERADOR'] == 'B2') | (df['OPERADOR'] == 'E1') | (df['OPERADOR'] == 'TM') |
                            (df['OPERADOR'] == 'P3') | (df['OPERADOR'] == 'VI') | (df['OPERADOR'] == 'B1') |
                            (df['OPERADOR'] == 'P1') | (df['OPERADOR'] == 'O1') | (df['OPERADOR'] == 'VD') |
                            (df['OPERADOR'] == 'N1') | (df['OPERADOR'] == 'I1') | (df['OPERADOR'] == 'IF') |
                            (df['OPERADOR'] == 'E2') | (df['OPERADOR'] == 'BP') | (df['OPERADOR'] == 'P2') |
                            (df['OPERADOR'] == 'L1') | (df['OPERADOR'].str.lower().str.contains('i. de portugal'))]

            dados1.loc[:, 'OPERADOR'] = 'Infraestruturas de Portugal'
            dados1.loc[:, 'DATA ENTRADA'] = pd.to_datetime(
                dados1['DATA ENTRADA'], format='%d/%m/%Y')
            dados1.loc[:, 'DATA SAÍDA'] = pd.to_datetime(
                dados1['DATA SAÍDA'], format='%d/%m/%Y')
            dados1.loc[:, 'DATA PAGAMENTO'] = pd.to_datetime(
                dados1['DATA PAGAMENTO'], format='%d/%m/%Y')

            dados2 = df.loc[(df['OPERADOR'] == 'BR') | (
                df['OPERADOR'].str.lower().str.contains('brisa'))]

            dados2.loc[:, 'OPERADOR'] = 'Brisa Concessao Rodoviaria, S.'
            dados2.loc[:, 'DATA ENTRADA'] = pd.to_datetime(
                dados2['DATA ENTRADA'], format='%d/%m/%Y')
            dados2.loc[:, 'DATA SAÍDA'] = pd.to_datetime(
                dados2['DATA SAÍDA'], format='%d/%m/%Y')
            dados2.loc[:, 'DATA PAGAMENTO'] = pd.to_datetime(
                dados2['DATA PAGAMENTO'], format='%d/%m/%Y')

            dados3 = df.loc[(df['OPERADOR'] == 'S1') | (
                df['OPERADOR'].str.lower().str.contains('scutvias'))]

            dados3.loc[:, 'OPERADOR'] = 'Scutvias - Autoestradas da Beira, S.A.'
            dados3.loc[:, 'DATA ENTRADA'] = pd.to_datetime(
                dados3['DATA ENTRADA'], format='%d/%m/%Y')
            dados3.loc[:, 'DATA SAÍDA'] = pd.to_datetime(
                dados3['DATA SAÍDA'], format='%d/%m/%Y')
            dados3.loc[:, 'DATA PAGAMENTO'] = pd.to_datetime(
                dados3['DATA PAGAMENTO'], format='%d/%m/%Y')

            dados4 = df.loc[(
                df['OPERADOR'].str.lower().str.contains('bragaparques'))]

            dados4.loc[:, 'OPERADOR'] = 'Bragaparques, S.A.'
            dados4.loc[:, 'DATA ENTRADA'] = pd.to_datetime(
                dados4['DATA ENTRADA'], format='%d/%m/%Y')
            dados4.loc[:, 'DATA SAÍDA'] = pd.to_datetime(
                dados4['DATA SAÍDA'], format='%d/%m/%Y')
            dados4.loc[:, 'DATA PAGAMENTO'] = pd.to_datetime(
                dados4['DATA PAGAMENTO'], format='%d/%m/%Y')

            dados5 = df.loc[(df['OPERADOR'] == 'AA') | (
                df['OPERADOR'].str.lower().str.contains('autoestradas do atlântico'))]

            dados5.loc[:, 'OPERADOR'] = 'AUTOESTRADAS DO ATLANTICO'
            dados5.loc[:, 'DATA ENTRADA'] = pd.to_datetime(
                dados5['DATA ENTRADA'], format='%d/%m/%Y')
            dados5.loc[:, 'DATA SAÍDA'] = pd.to_datetime(
                dados5['DATA SAÍDA'], format='%d/%m/%Y')
            dados5.loc[:, 'DATA PAGAMENTO'] = pd.to_datetime(
                dados5['DATA PAGAMENTO'], format='%d/%m/%Y')

            dados6 = df.loc[(df['OPERADOR'] == 'DL') | (
                df['OPERADOR'].str.lower().str.contains('aedl'))]

            dados6.loc[:, 'OPERADOR'] = 'Aedl - Estradas de Douro Litoral S.A.'
            dados6.loc[:, 'DATA ENTRADA'] = pd.to_datetime(
                dados6['DATA ENTRADA'], format='%d/%m/%Y')
            dados6.loc[:, 'DATA SAÍDA'] = pd.to_datetime(
                dados6['DATA SAÍDA'], format='%d/%m/%Y')
            dados6.loc[:, 'DATA PAGAMENTO'] = pd.to_datetime(
                dados6['DATA PAGAMENTO'], format='%d/%m/%Y')

            dados7 = df.loc[(df['OPERADOR'].str.contains('VV')) | (
                df['OPERADOR'].str.lower().str.contains('via verde'))]

            dados7.loc[:, 'OPERADOR'] = 'Via Verde Portugal, S.A.'
            dados7.loc[:, 'DATA ENTRADA'] = pd.to_datetime(
                dados7['DATA ENTRADA'], format='%d/%m/%Y')
            dados7.loc[:, 'DATA SAÍDA'] = pd.to_datetime(
                dados7['DATA SAÍDA'], format='%d/%m/%Y')
            dados7.loc[:, 'DATA PAGAMENTO'] = pd.to_datetime(
                dados7['DATA PAGAMENTO'], format='%d/%m/%Y')

            dados8 = df.loc[(df['OPERADOR'] == 'VE')]

            dados8.loc[:, 'OPERADOR'] = 'Via Verde Pot.(Ve) Espanha'
            dados8.loc[:, 'DATA ENTRADA'] = pd.to_datetime(
                dados8['DATA ENTRADA'], format='%d/%m/%Y')
            dados8.loc[:, 'DATA SAÍDA'] = pd.to_datetime(
                dados8['DATA SAÍDA'], format='%d/%m/%Y')
            dados8.loc[:, 'DATA PAGAMENTO'] = pd.to_datetime(
                dados8['DATA PAGAMENTO'], format='%d/%m/%Y')

            dados9 = df.loc[(df['OPERADOR'] == 'LS') | (
                df['OPERADOR'].str.lower().str.contains('lusoponte'))]

            dados9.loc[:, 'OPERADOR'] = 'Lusoponte Concessionario para Trave.Tejo'
            dados9.loc[:, 'DATA ENTRADA'] = pd.to_datetime(
                dados9['DATA ENTRADA'], format='%d/%m/%Y')
            dados9.loc[:, 'DATA SAÍDA'] = pd.to_datetime(
                dados9['DATA SAÍDA'], format='%d/%m/%Y')
            dados9.loc[:, 'DATA PAGAMENTO'] = pd.to_datetime(
                dados9['DATA PAGAMENTO'], format='%d/%m/%Y')

            dados10 = df.loc[(df['OPERADOR'] == 'BL') | (
                df['OPERADOR'].str.lower().str.contains('brisal'))]

            dados10.loc[:, 'OPERADOR'] = 'Brisal - Auto Estrada Do Litoral'
            dados10.loc[:, 'DATA ENTRADA'] = pd.to_datetime(
                dados10['DATA ENTRADA'], format='%d/%m/%Y')
            dados10.loc[:, 'DATA SAÍDA'] = pd.to_datetime(
                dados10['DATA SAÍDA'], format='%d/%m/%Y')
            dados10.loc[:, 'DATA PAGAMENTO'] = pd.to_datetime(
                dados10['DATA PAGAMENTO'], format='%d/%m/%Y')

            # Salvar os dados filtrados em arquivos CSV separados
            if len(dados1) != 0:
                dados1.to_excel('C:\\importacao\\' + nome_arquivo +
                                '_I_Portugal.xlsx', index=False)
            if len(dados2) != 0:
                dados2.to_excel('C:\\importacao\\' + nome_arquivo +
                                '_Brisa.xlsx', index=False)
            if len(dados3) != 0:
                dados3.to_excel('C:\\importacao\\' + nome_arquivo +
                                '_Scutvias.xlsx', index=False)
            if len(dados4) != 0:
                dados4.to_excel('C:\\importacao\\' + nome_arquivo +
                                '_BRAGAPARQUES.xlsx', index=False)
            if len(dados5) != 0:
                dados5.to_excel('C:\\importacao\\' + nome_arquivo +
                                '_AUTOESTRADAS_ATLANTICO.xlsx', index=False)
            if len(dados6) != 0:
                dados6.to_excel('C:\\importacao\\' + nome_arquivo +
                                '_AEDL.xlsx', index=False)
            if len(dados7) != 0:
                dados7.to_excel('C:\\importacao\\' + nome_arquivo +
                                '_VIA_VERDE_PORTUGAL.xlsx', index=False)
            if len(dados8) != 0:
                dados8.to_excel('C:\\importacao\\' + nome_arquivo +
                                '_VIA_VERDE_ESPANHA.xlsx', index=False)
            if len(dados9) != 0:
                dados9.to_excel('C:\\importacao\\' + nome_arquivo +
                                '_LUSOPONTE.xlsx', index=False)
            if len(dados10) != 0:
                dados10.to_excel('C:\\importacao\\' + nome_arquivo +
                                 '_BRISAL.xlsx', index=False)

            os.remove('C:\\importacao\\' + nome_arquivo + '.csv')

        # Exibir uma mensagem de conclusão
            messagebox.showinfo(
                'Concluído', 'O arquivo foi processado com sucesso.')
            root.destroy()

        def sair():
            root.destroy()
        # Criar janela principal
        root = tk.Tk()

        root.geometry("400x200")
        root.configure(bg="#3A7FF6")

        canvas2 = tk.Canvas(
            root,
            bg="#3A7FF6",
            height=200,
            width=400,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )

        canvas2.place(x=0, y=0)
        canvas2.create_rectangle(
            200,
            0.0,
            400,
            200,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            15,
            anchor="nw",
            text="Importador de Despesas",
            fill="#FCFCFC",
            font=("Roboto Bold", 12)
        )

        canvas2.create_text(
            220,
            15,
            anchor="nw",
            text="Valor da Fatura:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        canvas2.create_rectangle(
            15,
            35,
            80,
            40,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            60,
            anchor="nw",
            text="Aqui é necessário a introdução ",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            80,
            anchor="nw",
            text="do valor da Fatura.",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        entry = Text(root)
        entry.config(width=20, height=2)
        entry.place(x=220, y=40)
        entry.config(borderwidth=1, relief="solid")

        btn_obter_valor = tk.Button(
            root, text="Enviar dados", command=obter_valor)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=100)

        btn_obter_valor = Button(
            root, text="Sair do Ecrã", command=sair)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=150)

        canvas2.create_text(
            39.999999999999886,
            340.0,
            anchor="nw",
            text="Aplicação de suporte a importação de despesas",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 14 * -1)
        )

        # Código da imagem em base64
        codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

        # Decodifica o código base64 para dados binários
        dados_imagem_icon_luzitic = base64.b64decode(
            codigo_base64_icon_luzitic)
        buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
        imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

        # Converte a imagem para o formato .ico
        icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
        imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

        # Define o ícone da janela
        root.iconbitmap(default=icone_temp.name)

    # Definir o título da janela
        root.title('Importação de Despesas')

        root.resizable(False, False)
        root.mainloop()


def processar_arquivo_STARRESSA_ESPANHA_GASOLEO():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        # Carregar o arquivo Excel em um DataFrame
        df = pd.read_excel(filename)
        # Remove o caminho e a extensao do nome do ficheiro
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))

        duplicate_rows = df.duplicated()

        # Check if there are any duplicates
        if duplicate_rows.any():
            # Get the duplicate rows
            duplicate_data = df[duplicate_rows]
            print(duplicate_data)
            # Remove the duplicates
            df.drop_duplicates(inplace=True)

        # Calcular os Valores das colunas
        df['Valor Faturado'] = (
            df['Montante Operação'] - df['Montante desconto'])
        df['Valor Faturado S/ Iva '] = (df['Valor Faturado'] /
                                        ((df['% IMPOSTO']/100) + 1))

        # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
        df.to_excel('C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)
        # Exibir uma mensagem de conclusão
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_STARRESSA_PORTUGAL_GASOLEO():

    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        # Carregar o arquivo Excel em um DataFrame
        df = pd.read_excel(filename)
        # Remove o caminho e a extensao do nome do ficheiro
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))
        duplicate_rows = df.duplicated()

        # Check if there are any duplicates
        if duplicate_rows.any():
            # Get the duplicate rows
            duplicate_data = df[duplicate_rows]
            print(duplicate_data)
            # Remove the duplicates
            df.drop_duplicates(inplace=True)

        # Calcular os Valores das colunas
        df['Total c / IVA'] = (df['Montante Operação'] - df['Montante desconto'])

        # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
        df.to_excel('C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)
        # Exibir uma mensagem de conclusão
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_ILIDIO_MOTA():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:

        # Carregar o arquivo Excel em um DataFrame
        df = openpyxl.load_workbook(filename)
        # Remove o caminho e a extensao do nome do ficheiro
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))

        # Seleciona a planilha
        planilha = df.active

        planilha.delete_cols(1)

        # Remover primeiras 5 linhas
        for i in range(1, 4):
            planilha.delete_rows(1)

        df.save(
            'C:\\importacao\\' + nome_arquivo + '.xlsx')

        df = pd.read_excel('C:\\importacao\\' + nome_arquivo + '.xlsx')
        linhas_para_remover = df[df["Design"] == "OFERTA CAFÉ"].index

        # Remove as linhas do DataFrame
        df.drop(linhas_para_remover, inplace=True)

        df.to_excel(
            'C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)
        # Exibir uma mensagem de conclusão
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_STARRESSA_FRANCA_PORTAGENS():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        # Carregar o arquivo Excel em um DataFrame
        df = pd.read_excel(filename)
        # Remove o caminho e a extensao do nome do ficheiro
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))

        # cabecalho = list(df)
        # Calcular os Valores das colunas
        def calcular_valor(row):
            if row['% IMPOSTO'] > 0:
                if row['Conceito'] == 'AUTOROUTES A PEAGE':
                    count = df.loc[df['Conceito'].isin(
                        ['AUTOROUTES A PEAGE'])].shape[0]

                    return (row['Montante Operação'] + (df.loc[df['Conceito'] == 'GEST. DES SERVICES AUTOROUTES', 'Montante Operação'].iloc[0] / count)) / (1 + (row['% IMPOSTO'] / 100))

                elif row['Conceito'] == 'GEST. DES SERVICES AUTOROUTES':
                    return (row['Montante Operação'])
                else:
                    return (row['Montante Operação'] / (1 + (row['% IMPOSTO'] / 100)))

            else:
                return row['Montante Operação']

        df['Comissões'] = df.apply(calcular_valor, axis=1)

        # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
        df.to_excel('C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)
        # Exibir uma mensagem de conclusão
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_STARRESSA_ITALIA_PORTAGENS():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        # Carregar o arquivo Excel em um DataFrame
        df = pd.read_excel(filename)
        # Remove o caminho e a extensao do nome do ficheiro
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))

        # Calcular os Valores das colunas
        def calcular_valor(row):
            if row['% IMPOSTO'] > 0:
                if row['Conceito'] == 'PEDAGGI AUTOSTRADALI':
                    count = df.loc[df['Conceito'].isin(
                        ['PEDAGGI AUTOSTRADALI'])].shape[0]

                    return (row['Montante Operação'] + (df.loc[df['Conceito'] == 'GEST. SERV. PEAJES CONSORZIO', 'Montante Operação'].iloc[0] / count)) / (1 + (row['% IMPOSTO'] / 100))

                elif row['Conceito'] == 'GEST. SERV. PEAJES CONSORZIO':
                    return (row['Montante Operação'])
                else:
                    return (row['Montante Operação'] / (1 + (row['% IMPOSTO'] / 100)))

            else:
                return row['Montante Operação']

        df['Comissões'] = df.apply(calcular_valor, axis=1)

        # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
        df.to_excel('C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)
        # Exibir uma mensagem de conclusão
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_STARRESSA_SUICA_PORTAGENS():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        def obter_valor():
            valorFaturaEntry = entry.get("1.0", "end-1c")
            # Carregar o arquivo Excel em um DataFrame
            df = pd.read_excel(filename)
            # Remove o caminho e a extensao do nome do ficheiro
            nome_arquivo, extensao = os.path.splitext(
                os.path.basename(filename))

            df["Moeda"] = "EURO"
            df["Montante Operação"] = (
                float(valorFaturaEntry)/100) * df["Montante Operação"]

            # Calcular os Valores das colunas
            def calcular_valor(row):
                if row['% IMPOSTO'] > 0:
                    if row['Conceito'] == 'AUTOROUTES A PEAGE':
                        count = df.loc[df['Conceito'].isin(
                            ['AUTOROUTES A PEAGE'])].shape[0]

                        return (row['Montante Operação'] + (df.loc[df['Conceito'] == 'GEST. DES SERVICES AUTOROUTES', 'Montante Operação'].iloc[0] / count)) / (1 + (row['% IMPOSTO'] / 100))

                    elif row['Conceito'] == 'GEST. DES SERVICES AUTOROUTES':
                        return (row['Montante Operação'])
                    else:
                        return (row['Montante Operação'] / (1 + (row['% IMPOSTO'] / 100)))

                else:
                    return row['Montante Operação']

            df['Comissões'] = df.apply(calcular_valor, axis=1)

            # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
            df.to_excel('C:\\importacao\\' +
                        nome_arquivo + '.xlsx', index=False)
            # Exibir uma mensagem de conclusão
            messagebox.showinfo(
                'Concluído', 'O arquivo foi processado com sucesso.')
            root.destroy()

        def sair():
            root.destroy()
        # Criar janela principal
        root = tk.Tk()

        root.geometry("400x200")
        root.configure(bg="#3A7FF6")

        canvas2 = tk.Canvas(
            root,
            bg="#3A7FF6",
            height=200,
            width=400,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )

        canvas2.place(x=0, y=0)
        canvas2.create_rectangle(
            200,
            0.0,
            400,
            200,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            15,
            anchor="nw",
            text="Importador de Despesas",
            fill="#FCFCFC",
            font=("Roboto Bold", 12)
        )

        canvas2.create_text(
            220,
            15,
            anchor="nw",
            text="Valor do Euro:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        canvas2.create_rectangle(
            15,
            35,
            80,
            40,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            60,
            anchor="nw",
            text="Aqui é necessário a introdução ",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            80,
            anchor="nw",
            text="do valor de 100 Francos",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            100,
            anchor="nw",
            text="que corresponde a Euros.",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        entry = Text(root)
        entry.config(width=20, height=2)
        entry.place(x=220, y=40)
        entry.config(borderwidth=1, relief="solid")

        btn_obter_valor = tk.Button(
            root, text="Enviar dados", command=obter_valor)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=100)

        btn_obter_valor = Button(
            root, text="Sair do Ecrã", command=sair)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=150)

        canvas2.create_text(
            39.999999999999886,
            340.0,
            anchor="nw",
            text="Aplicação de suporte a importação de despesas",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 14 * -1)
        )

        # Código da imagem em base64
        codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

        # Decodifica o código base64 para dados binários
        dados_imagem_icon_luzitic = base64.b64decode(
            codigo_base64_icon_luzitic)
        buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
        imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

        # Converte a imagem para o formato .ico
        icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
        imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

        # Define o ícone da janela
        root.iconbitmap(default=icone_temp.name)

    # Definir o título da janela
        root.title('Importação de Despesas')

        root.resizable(False, False)
        root.mainloop()


def processar_arquivo_IDS():

    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xls')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        def obter_valor():
            valorFaturaEntry = entry.get("1.0", "end-1c")
            # Carregar o arquivo Excel em um DataFrame
            nome_arquivo, extensao = os.path.splitext(
                os.path.basename(filename))
            df = pd.read_excel(filename)

            # Salvar como XLSX
            df.to_excel('C:\\importacao\\' +
                        nome_arquivo + '.xlsx', index=False)

            df = pd.read_excel('C:\\importacao\\' + nome_arquivo + '.xlsx')

            dados1 = df['TRS_DATE']
            df['TRS_DATE'] = pd.to_datetime(
                dados1, format='%Y-%m-%d').dt.date
            dados2 = df['INVO_DATE']
            df['INVO_DATE'] = pd.to_datetime(
                dados2, format='%Y-%m-%d').dt.date
            df["FATURA"] = valorFaturaEntry

            df.to_excel('C:\\importacao\\' +
                        nome_arquivo + '.xlsx', index=False)

            messagebox.showinfo(
                'Concluído', 'O arquivo foi processado com sucesso.')
            root.destroy()

        def sair():
            root.destroy()
        # Criar janela principal
        root = tk.Tk()

        root.geometry("400x250")
        root.configure(bg="#3A7FF6")

        canvas2 = tk.Canvas(
            root,
            bg="#3A7FF6",
            height=250,
            width=400,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )

        canvas2.place(x=0, y=0)
        canvas2.create_rectangle(
            200,
            0.0,
            400,
            250,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            15,
            anchor="nw",
            text="Importador de Despesas",
            fill="#FCFCFC",
            font=("Roboto Bold", 12)
        )

        canvas2.create_text(
            220,
            15,
            anchor="nw",
            text="Número da Fatura:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        canvas2.create_rectangle(
            15,
            35,
            80,
            40,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            60,
            anchor="nw",
            text="Aqui é necessário a introdução ",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            80,
            anchor="nw",
            text="do numero da fatura",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            100,
            anchor="nw",
            text="correspondente a importação.",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            120,
            anchor="nw",
            text="Caso o ficheiro a ser importado",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            140,
            anchor="nw",
            text="seja novo, tem que se guardar",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            160,
            anchor="nw",
            text="em formato xls mais especifico",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            180,
            anchor="nw",
            text="'Livro do Excle 97-2003 (.xls).'",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        entry = Text(root)
        entry.config(width=20, height=2)
        entry.place(x=220, y=40)
        entry.config(borderwidth=1, relief="solid")

        btn_obter_valor = tk.Button(
            root, text="Enviar dados", command=obter_valor)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=100)

        btn_obter_valor = Button(
            root, text="Sair do Ecrã", command=sair)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=150)

        canvas2.create_text(
            39.999999999999886,
            340.0,
            anchor="nw",
            text="Aplicação de suporte a importação de despesas",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 14 * -1)
        )

        # Código da imagem em base64
        codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

        # Decodifica o código base64 para dados binários
        dados_imagem_icon_luzitic = base64.b64decode(
            codigo_base64_icon_luzitic)
        buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
        imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

        # Converte a imagem para o formato .ico
        icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
        imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

        # Define o ícone da janela
        root.iconbitmap(default=icone_temp.name)

    # Definir o título da janela
        root.title('Importação de Despesas')

        root.resizable(False, False)
        root.mainloop()


def processar_arquivo_MONTEPIO_RENTING():

    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xls')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        def obter_valor():
            valorFaturaEntry = entry.get("1.0", "end-1c")
            valorFaturaEntry2 = entry2.get("1.0", "end-1c")

            # Ler o arquivo Excel
            nome_arquivo, extensao = os.path.splitext(
                os.path.basename(filename))
            df = pd.read_excel(filename)

            # Salvar como XLSX
            df.to_excel('C:\\importacao\\' +
                        nome_arquivo + '.xlsx', index=False)

            # Carregando o arquivo XLSX
            workbook = openpyxl.load_workbook(
                'C:\\importacao\\' + nome_arquivo + '.xlsx')

            worksheet = workbook['Sheet1']

            target_cell = None
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == "Matrícula":
                        target_cell = cell
                        break
                if target_cell:
                    break

            if target_cell:
                # Obter os dados abaixo da célula que contém o valor "Matrícula"
                data = []
                none_count = 0

                for row in worksheet.iter_rows(min_row=target_cell.row + 1, min_col=worksheet.min_column, max_col=worksheet.max_column):
                    row_data = [cell.value for cell in row]

                    none_count = 0

                    for value in row_data:
                        if value == 'Total':
                            none_count = 1
                            break
                        dados = [row_data[0], row_data[1], '', row_data[3],
                                 row_data[5], 'Aluguer', valorFaturaEntry2, valorFaturaEntry, '', '', '', '', '', '']

                    for value in row_data:
                        if value == 'Total':
                            none_count = 1
                            break
                        dados2 = [row_data[0], row_data[1], '', row_data[3], row_data[7],
                                  'Serviço de Gestão', valorFaturaEntry2, valorFaturaEntry, '', '', '', '', '', '']

                    for value in row_data:
                        if value == 'Total':
                            none_count = 1
                            break
                        dados3 = [row_data[0], row_data[1], '', row_data[3], row_data[8],
                                  'Contrato de Serviço', valorFaturaEntry2, valorFaturaEntry, '', '', '', '', '', '']

                    if none_count == 0:
                        data.append(dados)
                        data.append(dados2)
                        data.append(dados3)

                if data:
                    # Criar um DataFrame pandas com os dados
                    df = pd.DataFrame(
                        data, columns=[cell.value for cell in worksheet[1]])

                    # Escrever o DataFrame de volta no arquivo Excel
                    df.to_excel('C:\\importacao\\' + nome_arquivo +
                                '.xlsx', index=False)

            messagebox.showinfo(
                'Concluído', 'O arquivo foi processado com sucesso.')

            root.destroy()

        def sair():
            root.destroy()

        # Criar janela principal
        root = tk.Tk()

        root.geometry("400x300")
        root.configure(bg="#3A7FF6")

        canvas2 = tk.Canvas(
            root,
            bg="#3A7FF6",
            height=300,
            width=400,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )

        canvas2.place(x=0, y=0)
        canvas2.create_rectangle(
            200,
            0.0,
            400,
            300,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            15,
            anchor="nw",
            text="Importador de Despesas",
            fill="#FCFCFC",
            font=("Roboto Bold", 12)
        )

        canvas2.create_text(
            220,
            15,
            anchor="nw",
            text="Fatura:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        canvas2.create_rectangle(
            15,
            35,
            80,
            40,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            60,
            anchor="nw",
            text="Aqui é necessário a introdução ",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            80,
            anchor="nw",
            text="Valor do IVA e Fatura",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            100,
            anchor="nw",
            text="correspondente a importação.",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        entry = Text(root)
        entry.config(width=20, height=2)
        entry.place(x=220, y=40)
        entry.config(borderwidth=1, relief="solid")

        btn_obter_valor = tk.Button(
            root, text="Enviar dados", command=obter_valor)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=180)

        btn_obter_valor = Button(
            root, text="Sair do Ecrã", command=sair)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=220)

        canvas2.create_text(
            39.999999999999886,
            340.0,
            anchor="nw",
            text="Aplicação de suporte a importação de despesas",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 14 * -1)
        )
        canvas2.create_text(
            220,
            80,
            anchor="nw",
            text="IVA:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        entry2 = Text(root)
        entry2.config(width=20, height=2)
        entry2.place(x=220, y=100)
        entry2.config(borderwidth=1, relief="solid")

        # Código da imagem em base64
        codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

        # Decodifica o código base64 para dados binários
        dados_imagem_icon_luzitic = base64.b64decode(
            codigo_base64_icon_luzitic)
        buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
        imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

        # Converte a imagem para o formato .ico
        icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
        imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

        # Define o ícone da janela
        root.iconbitmap(default=icone_temp.name)

        # Definir o título da janela
        root.title('Importação de Despesas')

        root.resizable(False, False)
        root.mainloop()


def processar_arquivo_VALCARCE():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.txt')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        def obter_valor():
            valorFaturaEntry = entry.get_date()
            valorFaturaEntry2 = entry2.get("1.0", "end-1c")
            # Carregar o arquivo Excel em um DataFrame
            df = pd.read_csv(filename, delimiter='\t')
            # Remove o caminho e a extensao do nome do ficheiro
            nome_arquivo, extensao = os.path.splitext(
                os.path.basename(filename))

            # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
            df.to_excel('C:\\importacao\\' +
                        nome_arquivo + '.xlsx', index=False)

            ficheiro = 'C:\\importacao\\' + nome_arquivo + '.xlsx'
            # Carregando o arquivo XLSX
            df = pd.read_excel(ficheiro)
            cabecalho = list(df)
            #  Altera as "," para "." para converter em float
            Precio = [s.replace('.', '').replace(',', '.')
                      for s in df['Pre.Clien']]
            Dto = [s.replace('.', '').replace(',', '.')
                   for s in df['Dto.Clien']]
            Cantidad = [s.replace('.', '').replace(',', '.')
                        for s in df['Cantidad']]
            df['Cantidad'] = [s.replace('.', '').replace(
                ',', '.') for s in df['Cantidad']]
            df['Preco'] = df['Pre.Clien']
            df['Desconto'] = df['Dto.Clien']

            # Converte as colunas em float
            valores_Cantidad = []
            for valor in Cantidad:
                valor_float = float(valor)
                valores_Cantidad.append(valor_float)

            # Converte as colunas em float
            valores_Precio = []
            for valor in Precio:
                valor_float = float(valor)
                valores_Precio.append(valor_float)

            # Converte as colunas em float
            valores_Dto = []
            for valor in Dto:
                valor_float = float(valor)
                valores_Dto.append(valor_float)

            # Calcula os totais
            P_D = []
            for i in range(len(valores_Precio)):
                total = (valores_Precio[i] - valores_Dto[i])
                P_D.append(total)

            df['P-D'] = P_D

            # Calcula os totais
            PxQ = []
            for i in range(len(valores_Cantidad)):
                total = (P_D[i] * valores_Cantidad[i])
                PxQ.append(total)

            df['PxQ'] = PxQ

            # Seleciona a os elementos do Cabeçalho
            df['Mes de Faturacao'] = valorFaturaEntry
            df['Taxa de IVA'] = int(valorFaturaEntry2)

            cabecalho.append('Preco')
            cabecalho.append('Desconto')
            cabecalho.append('P-D')
            cabecalho.append('PxQ')
            cabecalho.append('Mes de Faturacao')
            cabecalho.append('Taxa de IVA')

            # Selecionar as colunas a serem exportadas
            colunas = cabecalho

            # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
            df[colunas].to_excel(
                'C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)

            # Exibir uma mensagem de conclusão
            messagebox.showinfo(
                'Concluído', 'O arquivo foi processado com sucesso.')

            root.destroy()

        def sair():
            root.destroy()

        # Criar janela principal
        root = tk.Tk()

        root.geometry("400x300")
        root.configure(bg="#3A7FF6")

        canvas2 = tk.Canvas(
            root,
            bg="#3A7FF6",
            height=300,
            width=400,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )

        canvas2.place(x=0, y=0)
        canvas2.create_rectangle(
            200,
            0.0,
            400,
            300,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            15,
            anchor="nw",
            text="Importador de Despesas",
            fill="#FCFCFC",
            font=("Roboto Bold", 12)
        )

        canvas2.create_text(
            220,
            15,
            anchor="nw",
            text="Data da Importação:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        canvas2.create_rectangle(
            15,
            35,
            80,
            40,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            60,
            anchor="nw",
            text="Aqui é necessário a introdução ",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            80,
            anchor="nw",
            text="da Data, IVA",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            100,
            anchor="nw",
            text="correspondente a importação.",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        entry = DateEntry(root, selectmode="day", date_pattern='yyyy-mm-dd')
        entry.config(width=25)
        entry.place(x=220, y=40)

        btn_obter_valor = tk.Button(
            root, text="Enviar dados", command=obter_valor)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=150)

        btn_obter_valor = Button(
            root, text="Sair do Ecrã", command=sair)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=190)

        canvas2.create_text(
            39.999999999999886,
            340.0,
            anchor="nw",
            text="Aplicação de suporte a importação de despesas",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 14 * -1)
        )
        canvas2.create_text(
            220,
            70,
            anchor="nw",
            text="Valor do IVA:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        entry2 = Text(root)
        entry2.config(width=20, height=2)
        entry2.place(x=220, y=90)
        entry2.config(borderwidth=1, relief="solid")

        # Código da imagem em base64
        codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

        # Decodifica o código base64 para dados binários
        dados_imagem_icon_luzitic = base64.b64decode(
            codigo_base64_icon_luzitic)
        buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
        imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

        # Converte a imagem para o formato .ico
        icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
        imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

        # Define o ícone da janela
        root.iconbitmap(default=icone_temp.name)

        # Definir o título da janela
        root.title('Importação de Despesas')

        root.resizable(False, False)
        root.mainloop()


def processar_arquivo_SEGUROS():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xls')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        # Ler o arquivo Excel
        nome_arquivo, extensao = os.path.splitext(
            os.path.basename(filename))
        df = pd.read_excel(filename)

        # Salvar como XLSX
        df.to_excel('C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)

        # Carregar o arquivo XLSX
        df = pd.read_excel('C:\\importacao\\' + nome_arquivo + '.xlsx')

        df = df.dropna(subset=[df.columns[0]])
        primeiros_nove = df["Objecto Seguro"].str[:9]

        df["MATRICULA"] = primeiros_nove

        df.to_excel('C:\\importacao\\' + nome_arquivo +
                    '.xlsx', index=False)

        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_AS24_FRANCA_PORTAGENS():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:

        # Carregar o arquivo Excel em um DataFrame
        df = pd.read_excel(filename)
        # Remove o caminho e a extensao do nome do ficheiro
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))

        # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
        df.to_excel('C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)
        # Exibir uma mensagem de conclusão
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_AS24_FRANCA_COMBUSTIVEL():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:

        # Carregar o arquivo Excel em um DataFrame
        df = pd.read_excel(filename)
        # Remove o caminho e a extensao do nome do ficheiro
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))

        # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
        df.to_excel('C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)
        # Exibir uma mensagem de conclusão
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_AS24_ESPANHA_COMBUSTIVEL():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:

        # Carregar o arquivo Excel em um DataFrame
        df = pd.read_excel(filename)
        # Remove o caminho e a extensao do nome do ficheiro
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))

        # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
        df.to_excel('C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)
        # Exibir uma mensagem de conclusão
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_AS24_ESPANHA_PORTAGENS():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:

        # Carregar o arquivo Excel em um DataFrame
        df = pd.read_excel(filename)
        # Remove o caminho e a extensao do nome do ficheiro
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))

        # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
        df.to_excel('C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)
        # Exibir uma mensagem de conclusão
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_CTIB():
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.pdf')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        pdf_path = filename

        # Extrair as tabelas do PDF usando o camelot-py
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))
        tables = camelot.read_pdf(pdf_path, flavor="stream", pages="all")

        if tables:
            dfs = []
            for table in tables:
                df = table.df
                dfs.append(df)

            # Concatenar todas as tabelas em um único DataFrame
            final_df = pd.concat(dfs)

            # Salvar o DataFrame em um arquivo XLSX
            final_df.to_excel(
                'C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)

        # Caminho para o arquivo Excel
        excel_path = 'C:\\importacao\\' + nome_arquivo + '.xlsx'

        # Carregar o arquivo Excel
        workbook = openpyxl.load_workbook(excel_path)

        # Selecionar a planilha desejada (substitua 'Sheet1' pelo nome da sua planilha)
        worksheet = workbook['Sheet1']

        # Encontrar a célula que contém o valor "Matrícula"
        target_cell = None
        for row in worksheet.iter_rows():

            for cell in row:
                if cell.value == "Requisição:":
                    target_cell = cell
                    break
            if target_cell:
                break

        if target_cell:
            # Obter a coluna correspondente à célula que contém o valor "Matrícula"
            col_index = target_cell.column

            # Obter os dados abaixo da célula que contém o valor "Matrícula"
            data = []
            none_count = 0

            for row in worksheet.iter_rows(min_row=target_cell.row + 1, min_col=worksheet.min_column, max_col=worksheet.max_column):
                row_data = [cell.value for cell in row]

                none_count = 0

                for value in row_data:
                    if value == None or value == 'Motivo':
                        none_count += 1
                if none_count < 4:
                    data.append(row_data)

            if data:
                # Criar um DataFrame pandas com os dados
                df = pd.DataFrame(
                    data, columns=[cell.value for cell in worksheet[1]])

                df = df.dropna(how='all')
                df = df.reset_index(drop=True)
                # Escrever o DataFrame de volta no arquivo Excel
                df.to_excel(excel_path, index=False)

            def obter_valor():
                valorFaturaEntry = entry.get_date()
                valorFaturaEntry2 = entry2.get("1.0", "end-1c")
                valorFaturaEntry3 = entry3.get("1.0", "end-1c")

                # Read the Excel file
                df = pd.read_excel(excel_path)

                # Transpose the DataFrame to convert columns to rows
                # Obter o valor da célula
                valor4 = df.loc[0, 1]
                # Remove the header
                df = df.iloc[1:]

                # Remove the last row
                df = df.iloc[:-1]
                df = df.transpose()

                # Obter o valor da célula
                valor = df.loc[1, 7]
                valor2 = df.loc[2, 7]
                # Atribuir o valor da célula acima
                df.loc[0, 7] = valor
                df.loc[1, 7] = valor2

                # Atribuir um valor vazio à célula acima
                df.loc[2, 7] = pd.NA
                df[7] = df[7].str.replace('€', '')
                df['Motivo Inspecao'] = valor4
                df['IVA'] = int(valorFaturaEntry2)

                df = df.drop(0)
                valorIVA = (df['IVA']/100)+1
                valor5 = float(df.loc[1, 7].replace(',', '.'))
                df['Total S/IVA'] = valor5 / valorIVA
                df['Fatura'] = valorFaturaEntry3
                df['Data Fatura'] = valorFaturaEntry
                # Remove the last row
                df = df.iloc[:-3]
                # Save the transposed DataFrame to a new Excel file
                df.to_excel(excel_path, index=False)
                messagebox.showinfo(
                    'Concluído', 'O arquivo foi processado com sucesso.')
                root.destroy()

            def sair():
                root.destroy()

            # Criar janela principal
            root = tk.Tk()

            root.geometry("400x300")
            root.configure(bg="#3A7FF6")

            canvas2 = tk.Canvas(
                root,
                bg="#3A7FF6",
                height=300,
                width=400,
                bd=0,
                highlightthickness=0,
                relief="ridge"
            )

            canvas2.place(x=0, y=0)
            canvas2.create_rectangle(
                200,
                0.0,
                400,
                300,
                fill="#FCFCFC",
                outline="")

            canvas2.create_text(
                15,
                15,
                anchor="nw",
                text="Importador de Despesas",
                fill="#FCFCFC",
                font=("Roboto Bold", 12)
            )

            canvas2.create_text(
                220,
                15,
                anchor="nw",
                text="Data da Importação:",
                fill="#505485",
                font=("Roboto Bold", 12)
            )

            canvas2.create_rectangle(
                15,
                35,
                80,
                40,
                fill="#FCFCFC",
                outline="")

            canvas2.create_text(
                15,
                60,
                anchor="nw",
                text="Aqui é necessário a introdução ",
                fill="#FFFFFF",
                font=("ABeeZee Regular", 10)
            )

            canvas2.create_text(
                15,
                80,
                anchor="nw",
                text="do Valor da Data, IVA, Fatura",
                fill="#FFFFFF",
                font=("ABeeZee Regular", 10)
            )

            canvas2.create_text(
                15,
                100,
                anchor="nw",
                text="correspondente a importação.",
                fill="#FFFFFF",
                font=("ABeeZee Regular", 10)
            )

            entry = DateEntry(root, selectmode="day",
                              date_pattern='yyyy-mm-dd')
            entry.config(width=25)
            entry.place(x=220, y=40)

            btn_obter_valor = tk.Button(
                root, text="Enviar dados", command=obter_valor)
            btn_obter_valor.config(width=22)
            btn_obter_valor.place(
                x=220,
                y=210)

            btn_obter_valor = Button(
                root, text="Sair do Ecrã", command=sair)
            btn_obter_valor.config(width=22)
            btn_obter_valor.place(
                x=220,
                y=240)

            canvas2.create_text(
                39.999999999999886,
                340.0,
                anchor="nw",
                text="Aplicação de suporte a importação de despesas",
                fill="#FFFFFF",
                font=("ABeeZee Regular", 14 * -1)
            )
            canvas2.create_text(
                220,
                70,
                anchor="nw",
                text="Valor do IVA:",
                fill="#505485",
                font=("Roboto Bold", 12)
            )

            entry2 = Text(root)
            entry2.config(width=20, height=2)
            entry2.place(x=220, y=90)
            entry2.config(borderwidth=1, relief="solid")

            canvas2.create_text(
                220,
                140,
                anchor="nw",
                text="Fatura:",
                fill="#505485",
                font=("Roboto Bold", 12)
            )

            entry3 = Text(root)
            entry3.config(width=20, height=2)
            entry3.place(x=220, y=165)
            entry3.config(borderwidth=1, relief="solid")

            # Código da imagem em base64
            codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

            # Decodifica o código base64 para dados binários
            dados_imagem_icon_luzitic = base64.b64decode(
                codigo_base64_icon_luzitic)
            buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
            imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

            # Converte a imagem para o formato .ico
            icone_temp = tempfile.NamedTemporaryFile(
                suffix='.ico', delete=False)
            imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

            # Define o ícone da janela
            root.iconbitmap(default=icone_temp.name)

            # Definir o título da janela
            root.title('Importação de Despesas')

            root.resizable(False, False)
            root.mainloop()


def processar_arquivo_TRIMBLE():
    # Abrir a caixa de diálogo de seleção de arquivo
    def obter_valor():
        valorMES = entry.get_date()

        valorData = str(valorMES)
        # Carregar o arquivo Excel em um DataFrame
        # Configuração da conexão com o SQL Server
        server = 'SBs2019-ISAAC\ABMN'
        database = 'trimble'
        username = 'Bds'
        password = 'olivettiBDS1'

        # Criação da conexão com o SQL Server
        conn = pyodbc.connect(
            'DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD=' + password)

        # Chamar a stored procedure com os parâmetros
        cursor = conn.cursor()
        cursor.execute(
            "{CALL ImportacoesTrimble (?)}", (valorData))

        # Recuperar os resultados em uma lista de DataFrames
        tables = []

        while True:
            if cursor.description is not None:
                columns = [column[0] for column in cursor.description]
                rows = cursor.fetchall()
                df = pd.DataFrame.from_records(rows, columns=columns)
                tables.append(df)
            if not cursor.nextset():
                break

        # Fecha a conexão com o banco de dados
        conn.close()

        # Verifica se existem tabelas a serem salvas
        if len(tables) == 0:
            messagebox.showinfo(
                'ERRO', 'Nenhuma tabela encontrada para salvar.')
            return

        # Gera o arquivo Excel com as tabelas retornadas
        excel_file = 'C:\\importacao\\Trimble' + \
            valorData+'.xlsx'

        with pd.ExcelWriter(excel_file) as writer:
            for i, table in enumerate(tables):
                table.to_excel(
                    writer, sheet_name=f'Tabela {i+1}', index=False)
        df = pd.read_excel(excel_file)

        df['datafatura'] = valorMES.strftime('%Y-%m-%d')
        df['ndoc'] = ''

        df.to_excel(excel_file, index=False)
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')
        root.destroy()

    def sair():
        root.destroy()

    # Criar janela principal
    root = tk.Tk()

    root.geometry("400x200")
    root.configure(bg="#3A7FF6")

    canvas2 = tk.Canvas(
        root,
        bg="#3A7FF6",
        height=200,
        width=400,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas2.place(x=0, y=0)
    canvas2.create_rectangle(
        200,
        0.0,
        400,
        200,
        fill="#FCFCFC",
        outline="")

    canvas2.create_text(
        15,
        15,
        anchor="nw",
        text="Importador de Despesas",
        fill="#FCFCFC",
        font=("Roboto Bold", 12)
    )

    canvas2.create_text(
        220,
        15,
        anchor="nw",
        text="Data da Importação:",
        fill="#505485",
        font=("Roboto Bold", 12)
    )

    canvas2.create_rectangle(
        15,
        35,
        80,
        40,
        fill="#FCFCFC",
        outline="")

    canvas2.create_text(
        15,
        60,
        anchor="nw",
        text="Aqui é necessário a introdução ",
        fill="#FFFFFF",
        font=("ABeeZee Regular", 10)
    )

    canvas2.create_text(
        15,
        80,
        anchor="nw",
        text="do Valor da Data",
        fill="#FFFFFF",
        font=("ABeeZee Regular", 10)
    )

    canvas2.create_text(
        15,
        100,
        anchor="nw",
        text="correspondente a importação.",
        fill="#FFFFFF",
        font=("ABeeZee Regular", 10)
    )

    entry = DateEntry(root, selectmode="day", date_pattern='yyyy-mm-dd')
    entry.config(width=25)
    entry.place(x=220, y=40)

    btn_obter_valor = tk.Button(
        root, text="Enviar dados", command=obter_valor)
    btn_obter_valor.config(width=22)
    btn_obter_valor.place(
        x=220,
        y=100)

    btn_obter_valor = Button(
        root, text="Sair do Ecrã", command=sair)
    btn_obter_valor.config(width=22)
    btn_obter_valor.place(
        x=220,
        y=150)

    canvas2.create_text(
        39.999999999999886,
        340.0,
        anchor="nw",
        text="Aplicação de suporte a importação de despesas",
        fill="#FFFFFF",
        font=("ABeeZee Regular", 14 * -1)
    )

    # Código da imagem em base64
    codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

    # Decodifica o código base64 para dados binários
    dados_imagem_icon_luzitic = base64.b64decode(codigo_base64_icon_luzitic)
    buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
    imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

    # Converte a imagem para o formato .ico
    icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
    imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

    # Define o ícone da janela
    root.iconbitmap(default=icone_temp.name)

    # Definir o título da janela
    root.title('Importação de Despesas')

    root.resizable(False, False)
    root.mainloop()


def processar_arquivo_WTRANSNET():

    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.pdf')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        pdf_path = filename

        # Extrair as tabelas do PDF usando o camelot-py
        nome_arquivo, extensao = os.path.splitext(os.path.basename(filename))
        tables = camelot.read_pdf(pdf_path, flavor="stream", pages="all")

        if tables:
            dfs = []
            for table in tables:
                df = table.df
                dfs.append(df)

            # Concatenar todas as tabelas em um único DataFrame
            final_df = pd.concat(dfs)

            # Salvar o DataFrame em um arquivo XLSX
            final_df.to_excel(
                'C:\\importacao\\' + nome_arquivo + '.xlsx', index=False)

        # Caminho para o arquivo Excel
        excel_path = 'C:\\importacao\\' + nome_arquivo + '.xlsx'

        # Carregar o arquivo Excel
        workbook = openpyxl.load_workbook(excel_path)

        # Selecionar a planilha desejada (substitua 'Sheet1' pelo nome da sua planilha)
        worksheet = workbook['Sheet1']

        # Encontrar a célula que contém o valor "Matrícula"
        target_cell = None
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value == "Código":
                    target_cell = cell
                    break
            if target_cell:
                break

        if target_cell:
            # Obter a coluna correspondente à célula que contém o valor "Matrícula"
            col_index = target_cell.column

            # Obter os dados abaixo da célula que contém o valor "Matrícula"
            data = []
            none_count = 0

            for row in worksheet.iter_rows(min_row=target_cell.row + 1, min_col=worksheet.min_column, max_col=worksheet.max_column):
                row_data = [cell.value for cell in row]

                none_count = 0

                for value in row_data:
                    if value == None or value == 'N°CS' or value == 'N°chassis' or value == 'Tipo' or value == 'Nº Cliente':
                        none_count += 1

                if none_count < 4:
                    data.append(row_data)

            if data:
                # Criar um DataFrame pandas com os dados
                df = pd.DataFrame(
                    data, columns=[cell.value for cell in worksheet[1]])

                # Escrever o DataFrame de volta no arquivo Excel
                df.to_excel(excel_path, index=False)
            server = 'SBs2019-ISAAC\ABMN'
            database = 'aTrans'
            username = 'Bds'
            password = 'olivettiBDS1'

            # Criação da conexão com o SQL Server
            conn = pyodbc.connect(
                'DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+password)
            for index, row in df.iterrows():
                cdu_seguro_carga = row[4]

                # Criar uma consulta SQL
                sql = f"SELECT (SELECT matricula FROM viaturas WHERE id=IdViatura) AS Matricula, Num as 'Ordem_Carga', CDU_valor_seg_carga AS 'Valor_Seguro', '' AS desc2, '' AS desc3 FROM OrdemCarga WHERE CDU_seguro_carga='{cdu_seguro_carga}'"

                # Executar a consulta e obter os valores
                df_query = pd.read_sql_query(sql, conn)

                # Juntar os resultados da consulta ao arquivo original
                # df_merged = pd.concat([df, df_query], ignore_index=True)

                # # Execute the query and retrieve the results
                df_query = pd.read_sql_query(sql, conn)
                # # Adicionar por linhas de matricula
                if not df_query.empty:
                    # Extract the value from column of df_query
                    matricula = df_query['Matricula'].iloc[0]
                    Ordem_Carga = df_query['Ordem_Carga'].iloc[0]
                    Valor_Seguro = df_query['Valor_Seguro'].iloc[0]
                    desc2 = df_query['desc2'].iloc[0]
                    desc3 = df_query['desc3'].iloc[0]
                else:
                    matricula = ''  # Define um valor vazio caso não haja resultados
                    Ordem_Carga = ''  # Define um valor vazio caso não haja resultados
                    Valor_Seguro = ''  # Define um valor vazio caso não haja resultados
                    desc2 = ''  # Define um valor vazio caso não haja resultados
                    desc3 = ''  # Define um valor vazio caso não haja resultados

                # Assign the extracted value to  column of df
                df.at[index, 'Matricula'] = matricula
                df.at[index, 'Ordem de Carga'] = Ordem_Carga
                df.at[index, 'Valor_Seguro'] = Valor_Seguro
                df.at[index, 'desc2'] = desc2
                df.at[index, 'desc3'] = desc3

            # Salvar o arquivo atualizado
            df.to_excel(excel_path, index=False)
            conn.close()
            messagebox.showinfo(
                'Concluído', 'O arquivo foi processado com sucesso.')


def processar_arquivo_VIALTIS():
    # Abrir a caixa de diálogo de seleção de arquivo
    filename = filedialog.askopenfilename(
        initialdir='/', title='Selecione o arquivo', filetypes=[('Arquivos do Excel', '*.xlsx')])
    if (filename == ''):
        messagebox.showinfo('Erro Sem Ficheiro',
                            'Nenhum arquivo foi selecionado.')
    else:
        def obter_valor():
            valorFATURA = entry3.get("1.0", "end-1c")
            # Carregar o arquivo Excel em um DataFrame
            df = pd.read_excel(filename, sheet_name="Details")
            # Remove o caminho e a extensao do nome do ficheiro
            nome_arquivo, extensao = os.path.splitext(
                os.path.basename(filename))

            df["ORDEM DE CARGA"] = "Adicionar Ordem de Carga"
            df["NumeroFATURA"] = valorFATURA

            # Exportar o DataFrame para um arquivo XLSX com as colunas selecionadas
            df.to_excel('C:\\importacao\\' +
                        nome_arquivo + '.xlsx', index=False)
            # Exibir uma mensagem de conclusão
            messagebox.showinfo(
                'Concluído', 'O arquivo foi processado com sucesso.')
            root.destroy()

        def sair():
            root.destroy()

        # Criar janela principal
        root = tk.Tk()

        root.geometry("400x200")
        root.configure(bg="#3A7FF6")

        canvas2 = tk.Canvas(
            root,
            bg="#3A7FF6",
            height=200,
            width=400,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )

        canvas2.place(x=0, y=0)
        canvas2.create_rectangle(
            200,
            0.0,
            400,
            200,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            15,
            anchor="nw",
            text="Importador de Despesas",
            fill="#FCFCFC",
            font=("Roboto Bold", 12)
        )

        canvas2.create_text(
            220,
            15,
            anchor="nw",
            text="Número da Fatura:",
            fill="#505485",
            font=("Roboto Bold", 12)
        )

        canvas2.create_rectangle(
            15,
            35,
            80,
            40,
            fill="#FCFCFC",
            outline="")

        canvas2.create_text(
            15,
            60,
            anchor="nw",
            text="Aqui é necessário a introdução ",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            80,
            anchor="nw",
            text="do Número da Fatura",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        canvas2.create_text(
            15,
            100,
            anchor="nw",
            text="correspondente a importação.",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 10)
        )

        entry3 = Text(root)
        entry3.config(width=20, height=2)
        entry3.place(x=220, y=40)
        entry3.config(borderwidth=1, relief="solid")

        btn_obter_valor = tk.Button(
            root, text="Enviar dados", command=obter_valor)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=100)

        btn_obter_valor = Button(
            root, text="Sair do Ecrã", command=sair)
        btn_obter_valor.config(width=22)
        btn_obter_valor.place(
            x=220,
            y=150)

        canvas2.create_text(
            39.999999999999886,
            340.0,
            anchor="nw",
            text="Aplicação de suporte a importação de despesas",
            fill="#FFFFFF",
            font=("ABeeZee Regular", 14 * -1)
        )

        # Código da imagem em base64
        codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

        # Decodifica o código base64 para dados binários
        dados_imagem_icon_luzitic = base64.b64decode(
            codigo_base64_icon_luzitic)
        buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
        imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

        # Converte a imagem para o formato .ico
        icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
        imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

        # Define o ícone da janela
        root.iconbitmap(default=icone_temp.name)

       # Definir o título da janela
        root.title('Importação de Despesas')

        root.resizable(False, False)
        root.mainloop()


def processar_arquivo_BOMBA_PRÓPRIA_ABLUE_PARQUE():

    def obter_valor():
        valorMES = entry.get("1.0", "end-1c")
        valorANO = entry1.get("1.0", "end-1c")
        # Configuração da conexão com o SQL Server
        server = 'SBs2019-ISAAC\ABMN'
        database = 'aTrans'
        username = 'Bds'
        password = 'olivettiBDS1'

        # Criação da conexão com o SQL Server
        conn = pyodbc.connect(
            'DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD=' + password)

        # Criar a consulta SQL
        sql = f"SELECT Matricula, Empresa, Codigo_Movimento, Data_Movimento, Quantidade_Litros, Preço_S_Iva, Preço_C_Iva, Preço_Liquido_S_Iva, Preço_Liquido_C_Iva FROM LZ_Movimentos_Hecpoll where Codigo_Movimento not in ('Entrada de Stock') and   mes = '{valorMES}' and ano = '{valorANO}'"
        # Executar a consulta e recuperar os resultados
        df = pd.read_sql_query(sql, conn)

        # Fechar a conexão com o banco de dados
        conn.close()
        # Fechar a conexão com o banco de dados
        df.to_excel('C:\\importacao\\BOMBA_PRÓPRIA_ABLUE_PARQUE' +
                    valorMES+'_' + valorANO + '.xlsx', index=False)

        df = pd.read_excel('C:\\importacao\\BOMBA_PRÓPRIA_ABLUE_PARQUE' +
                           valorMES+'_' + valorANO + '.xlsx')

        df.loc[df['Quantidade_Litros'] == 0, 'Quantidade_Litros'] = 1
        # Formatar novamente a coluna no formato desejado
        df['Data_Movimento'] = pd.to_datetime(
            df['Data_Movimento']).dt.strftime('%Y-%m-%d')

        df.to_excel('C:\\importacao\\BOMBA_PRÓPRIA_ABLUE_PARQUE' +
                    valorMES+'_' + valorANO + '.xlsx', index=False)
        messagebox.showinfo(
            'Concluído', 'O arquivo foi processado com sucesso.')
        root.destroy()

    def sair():
        root.destroy()

    # Criar janela principal
    root = tk.Tk()

    root.geometry("400x300")
    root.configure(bg="#3A7FF6")

    canvas2 = tk.Canvas(
        root,
        bg="#3A7FF6",
        height=300,
        width=400,
        bd=0,
        highlightthickness=0,
        relief="ridge"
    )

    canvas2.place(x=0, y=0)
    canvas2.create_rectangle(
        200,
        0.0,
        400,
        300,
        fill="#FCFCFC",
        outline="")

    canvas2.create_text(
        15,
        15,
        anchor="nw",
        text="Importador de Despesas",
        fill="#FCFCFC",
        font=("Roboto Bold", 12)
    )

    canvas2.create_text(
        220,
        15,
        anchor="nw",
        text="Mês: (em valor numérico):",
        fill="#505485",
        font=("Roboto Bold", 12)
    )

    canvas2.create_rectangle(
        15,
        35,
        80,
        40,
        fill="#FCFCFC",
        outline="")

    canvas2.create_text(
        15,
        60,
        anchor="nw",
        text="Aqui é necessário a introdução ",
        fill="#FFFFFF",
        font=("ABeeZee Regular", 10)
    )

    canvas2.create_text(
        15,
        80,
        anchor="nw",
        text="do Valor do Mês e Ano",
        fill="#FFFFFF",
        font=("ABeeZee Regular", 10)
    )

    canvas2.create_text(
        15,
        100,
        anchor="nw",
        text="correspondente a importação.",
        fill="#FFFFFF",
        font=("ABeeZee Regular", 10)
    )

    entry = Text(root)
    entry.config(width=20, height=2)
    entry.place(x=220, y=40)
    entry.config(borderwidth=1, relief="solid")

    btn_obter_valor = tk.Button(
        root, text="Enviar dados", command=obter_valor)
    btn_obter_valor.config(width=22)
    btn_obter_valor.place(
        x=220,
        y=180)

    btn_obter_valor = Button(
        root, text="Sair do Ecrã", command=sair)
    btn_obter_valor.config(width=22)
    btn_obter_valor.place(
        x=220,
        y=220)

    canvas2.create_text(
        39.999999999999886,
        340.0,
        anchor="nw",
        text="Aplicação de suporte a importação de despesas",
        fill="#FFFFFF",
        font=("ABeeZee Regular", 14 * -1)
    )
    canvas2.create_text(
        220,
        80,
        anchor="nw",
        text="Ano:",
        fill="#505485",
        font=("Roboto Bold", 12)
    )

    entry1 = Text(root)
    entry1.config(width=20, height=2)
    entry1.place(x=220, y=100)
    entry1.config(borderwidth=1, relief="solid")

    # Código da imagem em base64
    codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

    # Decodifica o código base64 para dados binários
    dados_imagem_icon_luzitic = base64.b64decode(codigo_base64_icon_luzitic)
    buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
    imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

    # Converte a imagem para o formato .ico
    icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
    imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

    # Define o ícone da janela
    root.iconbitmap(default=icone_temp.name)

    # Definir o título da janela
    root.title('Importação de Despesas')

    root.resizable(False, False)
    root.mainloop()

# MENU DE SELEÇÃO


def selecionar_opcao(event):
    opcao = combo_box.get()
    if opcao == "STARRESSA ESP GASÓLEO":
        processar_arquivo_STARRESSA_ESPANHA_GASOLEO()

    elif opcao == "RESSA PORTUGAL - GASÓLEO":
        processar_arquivo_STARRESSA_PORTUGAL_GASOLEO()

    elif opcao == "ALTICE":
        processar_arquivo_ALTICE()

    elif opcao == "TOLL COLLECT":
        processar_arquivo_Toll_Collect()

    elif opcao == "VIALTIS":
        processar_arquivo_VIALTIS()

    elif opcao == "RESSA FRANÇA - PORTAGENS":
        processar_arquivo_STARRESSA_FRANCA_PORTAGENS()

    elif opcao == "RESSA SUÍÇA - PORTAGENS":
        processar_arquivo_STARRESSA_SUICA_PORTAGENS()

    elif opcao == "RESSA ITÁLIA":
        processar_arquivo_STARRESSA_ITALIA_PORTAGENS()

    elif opcao == "MONTEPIO RENTING":
        processar_arquivo_MONTEPIO_RENTING()

    elif opcao == "VALCARCE":
        processar_arquivo_VALCARCE()

    elif opcao == "VIA VERDE":
        processar_arquivo_VIAVERDE()

    elif opcao == "ILIDIO MOTA":
        processar_arquivo_ILIDIO_MOTA()

    elif opcao == "NORPETROL":
        processar_arquivo_NORPETROL()

    elif opcao == "IDS":
        processar_arquivo_IDS()

    elif opcao == "TRANQUILIDADE":
        processar_arquivo_SEGUROS()

    elif opcao == "HECPOLL":
        processar_arquivo_BOMBA_PRÓPRIA_ABLUE_PARQUE()

    elif opcao == "TRIMBLE":
        processar_arquivo_TRIMBLE()

    elif opcao == "AS24 ESPANHA - PORTAGENS":
        processar_arquivo_AS24_ESPANHA_PORTAGENS()

    elif opcao == "AS24 FRANÇA - COMBUSTIVEL":
        processar_arquivo_AS24_FRANCA_COMBUSTIVEL()

    elif opcao == "AS24 ESPANHA- COMBUSTIVEL":
        processar_arquivo_AS24_ESPANHA_COMBUSTIVEL()

    elif opcao == "AS24 FRANÇA - PORTAGENS":
        processar_arquivo_AS24_FRANCA_PORTAGENS()

    elif opcao == "STARRESSA ESP PORTAGENS":
        processar_arquivo_AS24_ESPANHA_PORTAGENS()

    elif opcao == "WTRANSNET":
        processar_arquivo_WTRANSNET()

    elif opcao == "CTIB - CENTRO TÉCNICO DE INSPECÇÕES DE BRAGA":
        processar_arquivo_CTIB()

    elif opcao == "WTRANSNET":
        processar_arquivo_WTRANSNET()

    else:
        processar_arquivo_PorFazer()


window = Tk()

window.geometry("862x519")
window.configure(bg="#3A7FF6")

canvas = Canvas(
    window,
    bg="#3A7FF6",
    height=519,
    width=862,
    bd=0,
    highlightthickness=0,
    relief="ridge"
)

canvas.place(x=0, y=0)
canvas.create_rectangle(
    430.9999999999999,
    0.0,
    861.9999999999999,
    519.0,
    fill="#FCFCFC",
    outline="")

codigo_base64_isaac = 'iVBORw0KGgoAAAANSUhEUgAAAQkAAAAxCAYAAADJNlwQAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAB7qSURBVHhe7Z0JnFxVlcbvfbV0dxKydGeBrOw7atiUZRAQXBAQEAXZUYFBcR9AR3TEFWT8oY6oBBGFkV0EZJVVVBAGBVmGRSQknRCQdHfCknTX8u58/1d9i1cvr7orncSE8X2/fOlX9ba7nnvOuefesmbNYUNxF/Et4haDnyeIk8QOsV/sFl8UnxWfEB8U7xdfETNkyPD/EG8VzxX/JroRsiTeJX5CnCxmyJDhDY6i+BHxETGt068KERiXiTuKGTJkeIMhEI8R54lpHXx18zpxKzFDhgxvANBZ/yCmdeY1STSLb4htYoYMGdZR/Ku4TEzrxP8o/kXEGZohQ4Z1CPgefiKmddq1waXie8QMGTKsA2DK8hYxrbOuTZbFo8UMGTKsRWD/3yqmddJ1gVXxSDFDhgxrAMMFU3H+EnG1dkI91HXZwK1vAjPJWjNef6fqM9Iop7MTrbVFY60kgFvsQtNvnKvo3PMutItNaBbquwU6Xm6cTz8aBabHHdGnDBkyrDYMJyT+TTyndjgyME+6jc2HOwd58yabN1vYnNlY7DCWU02hXi/B4EigLRjrcom0IkDmu6p7wlXNI7ryMVfpMWFup7tMP1OyqwIiRN9UO4xAZOiVtcMMLWK0+P7aYUsIRcp5kfiAiNDPsI4g2fGCbjN2vGtrm3BzZWCTy9zy6zWiF7p02Xoa8dv1l9q7IRxgJB+yk0+3QXhs0G4OCdrMZBukXout8KyrVh90FatObv6mDr9IWoK0B7fUuBzngW5G8wgnSruYbnNuc5uzsyVwdpLg6TTxZ7vlztm79ffq18r2qq1Mz0jCu88WT6sdRvi7OKV2mKFFzBRHKqwp78+Il0afMqx1NAiJ+WZspynk31cy9ojAmncUmmgadN5HXaV6Q1iyt4QlM9dV6x21qNOn50aFx+c6gmb3SwC4n1X7wyvCfts9jLAZCjx8e2kph+XaEEa2TUKkdiZK41L9f261tN7ZG5l5jFKtIhMSq45VERIAzWJXkXU8GdYyUjuxcPI4Y897d1B0B0gT2DUoSOVveq152lVDhMX10jBOzHW4DwZtsg7Scb8rhydVXrU9Lmz6vJFgls2Fc/JjzNY23yB0nHEPLyuV9t3SvLp48KvhcJZ4eu0wQjMhsb64uaiiSQWylHc+JtLoh8JG4vjaYVTO3Dc/+tQ63izG8z5XXFI7HBLTxU1E6ixZJ+RhQPxfsY8vWkSakGA9D9PWyXfwmcV/Pv8ePxZPrh1GZbyd6JXLNMSf2yP697P+Z5qYfG8S5JO6puyT7/HPGArcw3sxmXChpWE9cWsRcyytrCGOeMpqoZiGDUTaHXXNM/xzfJr5S12xTCKt3RHKQBpYbJk2QPs08H6/BmsFtItcwMmI440NDw/aqr/Ij60+W+wKu4tdrhnnp3zn+XSxM+RZ8WevTk6USfKU3rHCuwud39b5VoEmEX8uq1TjoFIuElvNBx1sqAZGJ31ZjN/DO6nEVvE+MX4/vF0cCruJpC15Xxpp9OeJw3U0j1li8hnvFJthlPiMGL9eZmMddIz4ueH4c9Hjc2LaNc3IyuTPinHgm0u7No2viReK8ahgOuP3RSKG0+5J4+/FqaIHZc9zV6XdnSQycKRdn0aExC5pkoSZjHji9FRnLw8HgiMrLwezS73mc5VXw7vDcohzcfCSOoZqRcudw9fwqg6JmLxWPF88U6RSeO+HRBr8O8S9xfcOfsc57NSvixeL94graAbLnEv1eIXGLh88XB3YUzxO9FlldHxY/HOMfxU9CGH/ZO0wFayaZYSJg5GLvLYCRqXv1Q4bQBl+oHa4AhgIrhFbXQuDlvExcY/o0+oHEbyP1w7rWFsmHkL7O+Lx0acaWhWOAIH3YZFFjx5HiKxqbqZ1pgEh/p+1wwiUPc/1aaGzD9fueKcHEco/FMdFn1rDxuKF5vli51bdbV17dxc7DxU/eWqu44ktbA51IylVViBawWHSMC5uQcPw1HWn6t5mQPqiejLDsJf4bvFt4rYi+1AkhRoq/7vEL0+3wWOXKR3J980vdN61yEyhI7WK4TSJj4r+HCNssoN7fFd8dJDNOjz588+izBlt4s+eLQ4HzCN/z/PiVbHPmCxjxCR2F/018Kciy/zRAGYMEhOIxh1vC4Tkt4KV1SSoR9T0+PV3ih5JTYKZk52GII3bI65JoCWQz50TxP/xeRHflb/216JHXJOgPdCBk9xf9KYKjDte45HK1OunROrWlzXEBDhIfE7018ZNDsref0+dNGt3DDq+3f0HXwwCAePvhz8S6VvxOseMYXD5nRhdpw631P591MQNBkruQGfdAdYyeluiK808Vw1vDkvmvrB8952u/Et9xSCNFkADY4Si09Yx1lj3zqDo9g+KZveg0OBETEJvv/gpWzlz34Gl2MJvF3cQKSQSOpTUphKfFpGg90ok/uaR/IRpQRAcq6dK26ilHfS68JUuF5w2rdJzgR5IobaK4RyXaDbxBoBaTJoA76GcMB/QXuB94h/FJBjNqchNo081VfIEkef5EZt7aYBUWBooM8rCj1BoXLeJT4qdfCGQHzpAHAhWImg9EAg0zjTsKzI6Ui8Pia04JGl4yeehhqcpeoAGn/RjoWV6oeSFiMfXxJdqhw3wbYf68ZomQsKPyNTTUGt+fiEiGAExN/vUDhtCARA0+FzSQJlStoAtDvyzmEL3Wh1aMG2+GWh71Bmgv3lhcJh4ee0wQlq7Q6ulj6S1O9rAt2qH0XNpH83qg/R5c6/Rr9VlglPfHRQrXwlG3/RgcdxB3fnOty5qG7fRotFTJi8sdm25ID9hF7VWGmRexBxg5KPRRFJnkBecZjqmdxe6juoudF4n7WR5cnSHc6V5/CS/XlWCpaoHtmRnjTG2upPNV0/ItVfn6N5HihMaNYdC5zPiD9CMVAKojSPBcJoEwhEHVfya4fgVMQnMLH+e53mhS8eP266MAGmgQ1CR/rrfDn4HEDb+exxyW4px0JD9eTjSsmqGNE1iZYlZ54GQSLumGdFEPeKaBMJzKMwR/bVxn05ck+gVT0nhF8X4wkfqwCOu3f2KL4YAZp2/Nj6Fjwm6su2ONuaBkPDfN3OKerB/i7+21zcqjxtE/ADg5Vk2mHtK0NF2YNBWGGUtoz7oc8bdqBuvXVrK3bqteYkRAk0ASYlayB4TdSw2nWOXFyx+hiPUhJHMCJgG9JowvK5aMteEA+ZhV0nzk5gzc6Mrx+fa8z7BS4wLnwor5nFXNQ+5ivmLKz8814X/olNU1KogqUkwYiV3yKJTHSqinpEfkkW6GQ3x0qPio/aOFQESvkv0Gg334X32zi124mK/DCoFYM/6gC7ez+iXnF2gnOMOOjQRNBOeQTq+KnpTgwaPRuCR1CQYGRkh04BKiibBczFfcGYNhzRNolXQMVCTUZs9kpoE/izKtBm+LPrZobgm8ZSYFJhxXCBiToJmmkQrQJVHcPjZBYQE7QXcKGKaNAN+BAZfENckgG93m4kM1rQ53/ZwdHMtHdy3OzRavicdcU0CLavBEkgAE8xPP68wq8Wo6SVInEs3s8EfvpYfdedThc7fa9QeqI3enctk81+3oNB53EIzZqKuGxIb2dxlJ+c6qncXxq/gO/DkHNdMsgEdqp6GfaRxHB20V/cICtUpiXMxxm2wkSKpSWBuxEEFs5fGvWK8oyVBJ44/x1cc+I0YPzccmVmIg4pvVlfN+EHRg8YfP0eDTgMObBqJv67VjpKmSaBNYXM3IwMJjbNuMsaQ1CQQsq2iVU2CgYB9Vv21mG0eKzO7AVHjvdkBMD38OdT3ZoIKoY4vxl+L1uKxn4i2iBmxMu3O++Pwg8S/bxYRy6CHP81f16BJMAK2Mhf+Sqexv/1orv2V44L24hgb7KmHdOlpVePcPdaaa4PAXTutvy9tnp+ZiWjV5g4yG47Itdv9g6IdleK/UCm7W8OSu7jab/7oygGpbQHYsNiyq4JknERSk8DrjaPPAxsTh2EyifgSvO2Kyo+Up/EkbUtGSK9hxEFnQfsAnEczwScAGKm8vY4jjGcky5D0UKd+JFog4vFmdKLzYybGtTrq3k99AxoX6Y9fwwwTDWg4pGkSdBqE40iQ1CSIAaFMm4Fy8v6AuCZBOaa1cUZihHg8r8wEfLx22KBJUI5eq46D+/EB+HO80zvp/11ksyQPyvgF0ZsP1B31TXnHZ0AQCt7sQsNB0/HAuRivL49m7Q73QHJtE2URb7vkn7qLC+p4uUdS7FiRzPkgiuH4sp56xYdt+7mPFSacN6/Y9ai0gUekXVzbXeg6er6ZnhwVSAA7Y9efMcbY8EhpCDcWxjXVLu6QdoEWMbp5jAUSF0mZxLDaTQpQyeLPTmoSzD2vrG3o4zRoSL5SIKrnCgJyEHTSeD2gudCYGW3jmlTcNEqCAKu4f8M7xACdp5lGlkbiGFotTxzQyfuHmt0YDisbJ4Gm5xHXJFolnS8+QxLXJIYKcmPA8Nd5wQQQ9mgx/lwrRJjHHZwIdgastGubMR4fRNu5Xky7rhlpH5+KN1A/4nhnCTYxoz4OmFbmVskUPg3sr5vFZrEJPOsMkcAO/84ILAQ7Ktdm3he02fVStItXpapcUx1wPwv7zV9dlQwQcMKojDeb93tsI2LbYU/HnVitgMYcd5pRHt6W82C2A98NvoKkVz4O/COMBl6Cow0cXDuMwGjFyNQMTNcdKPqyIAqRjs+UMKCMfyCiTTQDqryPh2BkQRPwtjLlhBo7VEwC9xDXQt226u9BGCZnVNC+EDQjAQKTNgNovMMBLQYnJGBk9X62ocBzsdW5F40HX5sHgpmpQeoBO58yTwN9hZkigMCImwXMZpEOpj45TsLni7qhnBhAGFDiQKPlGZgrw7U73p8MqCP9aBRM+Q7Vp0kLghKT6/F4R6RjkCkitVBTUfFwDpFQ7F8cW9iCOAeHSiCgw5JJBMZNYprAoCFhJ2NDMa1aT4vMD3dgUHRH5drNmxNh1oAcLHHhA+Nt8K0ZpZ5f60bSQ4fC1qZTecFAYePkyZAhw2oAdgwedjo0/RAyB4+XHI+p78R4RbGPGFni1zYjIzHzz+z30Ey4oNp9SWTet+H+rW2+elZ+dL/MjiVJMwTOLXS+cGqu45XJ6c5MBF6GDBlWAXFNwgMTAA8+3k86NtNfAI2CKDSEA+ozAoJruQYVmr/DmSU8479F1j6keZpJD+ohsQFoGd4zOzDa2CcOCIrbHZlrt29J0S5wdN4SyhSpDpgHXNmfZ28CNIyVAZpNPO4+w7qHtHabYc1gYLjCppNisyIwsIX8vDs2D/YOQgOzAg8oq8uwd5jKgjibhgLOJeb2iUaL238eCBwcqQSXNETJSbsIjx7Cd/Gkq4YXVfvNjeHA1UuNYzZhZYBd6mcVMmT4Z0dfsoMx9UcUHyMw6noczFSw1oBgDjQNfAqA6/5HxHOKucAMBuYDgiUZhJQGHEHMIzO98ye+SIA04jRCWBwg1qep8F1Iu3DSLszsdN/FEuvMT9vL5iuTWt+ABs0jPg2VIcM/M1ZwijNLQKdnmoeIN1T/FTqfgDpOh/2ZyPQj96wOIiROFBtmPWJgao35ZuaYG+5lUdoZuVHVPxdqodp/1N/fF8b/eEGxa6hY/QwZMowALDhi+grtgOkYpkL+S2T2I83xyKiLP4KVbis7j9uMjPoERRHunQZMGxYzYbI03Cs1I9xSAkPqBxJwJHESGTJkWAmwbJtAEkJBERiM4ET7pa7BEPiONQLM569s2HAzPigOpV0w74ypgl8jfh8RaRkyZFgDaObZR9X/tEjnY7oRrYHOiZ8izYZH60D7IPAkHmU4UuK7GEq7YD0DYcN+GhUfRoYMGVYRabMbdK6jRAKh2EciLQyVEFGmPXFiElxFB8ZxebVIlFYyrh6/Bv4NnJmHiAicVQG78BBRhw8l6ZAkT0RNsoqtlT0e63jabNrWUexLXV1og3CzatXuFVhzvnNRfnmT1b+C/h4yY6CHlZwRWCLvrN2tXMhts8lrf68vNZ9f7DrNGnd4LggPntq/JNqXYUFh4uHOus/MKPVEU7Usc5eIU7lHP0MSwTnbUy7mduRZ8zq6pgZV86wz7mXrTIg01PvH6O8Zs0q932WJvp73IxWCBLevXqck24dnlnp24ZPSd4xORfnQFXq8UbZszobh/mFg36HjM/ha98fbxy+Uxo8oD0S5bqvThHvXYM3VM0u9HxvMy3lKG+Ze9AShr1AubL+BeXGFGaz5ha7b9Aa1n+hCEqmyt1dUy2NO95sXdxc7Vb+2rJOsP4gQBPar0wd6kovezILixM1DEz4knXevmZVenO8R2D5Aadx2Rqk3iqSdX+w8X+k7Qq99jUzynbX2uXodFDsfUHltbp3rjwrB2na9/88zy71shGQWtHV9PHTubN0ozTnKqf7ZF4LywI7TYnupzi90Xq97d6vm7bYbLltcXwMxv63zE9bZcwbzLVh+WuaefFD91AYDS1mX4tPIdLz+DL4Dbd6Z82aUe6M1IboGrf5IfVcLVuTnapx7qaNsd8BRv6DQeazu+qaub9jGTvddrCcelLfVN/v3AdqOXvU91R8b0FIfvHphmlOSRVhUCLHnhKjS2TA5fLgpQDOgkig0BMYXRBLCWnnWOhALQTiwX7uBqYIGgibC7AeNlS3CRrqceHsRk4Z04Ash/NqDrPGLYyslIMDm5pmBUimcAvWUOSqkX/vPF/T3LVTB5dUYetRYpkQs9U7Wy55UcdZ/p6O7OAH/DJ3ovmK5Gl/Uo/Lm50PM1tVqcD97c/Cdaj1QpeBjiaB3Km/uJjXoUZ4mcGd3vBZEodR6Bi2mzeXsW1T56ysd66vBXaQHD2pY7m1K+2+nl3o71OjbYUWdRjdRZhGU5pz+W0IeomcoH3xW95ut469zT75UwJ/Tphe+Z/A50XZsek7BheYLundynRIQtee6XZX+S3ie0h2dkxD4YagMcD4Ja11BN80plZzK2K1ftZZOfEC+8CrtaRC2TRk/lrR6pgmIGqKyCfXgq54360dLoRcWJn5IAvskPaeu7UoIqh7NHYPlFz1T+avH00hAtKmcz/TndcPB6uxEBUeQZN5DT/ml8hiVr8p6rN48xhXb/UY/EgQT3qVrtlOF/iFXcQ07k6mM8iqrB2p11NvuSuWNdW2l4vLxhX87Kx1f8u/gr+r5XOWNkGqPXFTeg3nIlYqbKZ2bLWsLollFlVteea+3LTA330n7wCVwYznMx9eXUH+zNfD8WPU3ifoj/6q4i9OEhF9Ywtp11j8QE8HKSBZm4R/AqRkPdUYoYAbwYkK52a+SOIMrRM6RcTQOHxhFg2G1HIKHqVLWM7DwaCRx/cRt0HjxmbBHIs8kDSPGJqZvKVQyB9SIS/7zmSpzFZZGXTceKT/IOWqVm6ufR5JcqoRGluAcZ4Iz1OBPVUaP0ejasAWdKv46dZwfuMDeEUnu2qK1OgIXEv5+iEahRRoBn+evDe3Jtm25j1GJEFTdOd3Frp/MK3ZdqPe8V4zWrmgIYPiPRoLoQiFnLOZhHVS6GvXY1/PRNUcdYaLyVl//UjaFKF3SCRrSJ6FmbWC+JS2gW/c17LilBvqcse6jpNtTXx9TMssaGmoj3IAv441KPU8q8Xco4bSJOpSXi1RWC8jv4FdDwD6ifNxQKZQvXVSYsG3VuLNUa9+pF4aHNe+M8gD5KYkEVE8f8OWjTJ+mDlQbrQdBGQ8eclySljZxemlxtFNU1A5cwCD4RZUJo/5RElZ1IT3Y6dC2omTNNC/3Ks2XJPOtdzbUm164QjYCa9/u01kpDkT7UARBNXWFrG62ucCeK03sa2oUp6p97yfNNdKOIlj7Qmjdp33bkyayIDD2kDQh4YF6FNcWWJrM0lK2DqMwCNlm0U18bTyrIwmQYjSls7JPAcIBzQKBgTnC9m/eCUmm44KHDsXoO9wOQmlgRyeW8zLrsQZhy0FoH1LCt1R//IBU0lOXl8dHm79slu86Ud8rHeGJEvEqO1dyNoxvngKsRpCvS109Tg3jfHWqU+Id2pQrt6s17FkNqruIu/FXjSVfdQUWY9Wh+69ThV6u+y8LAnNCR+n15e00hsHDVNC6lfYK+dARW9Ufoud9cmm5Nz6SpUKNhrR+o1qubFcuhfE9E0w5by/Nh3bPSlDddTDtMjHthvl8bqiNXuppvTLyY7kt9Iq6ys7L9M7jSuVwm3wpn7bStwGU5YxS52fVATor1t6rTvTxwNl5yQLRCHwXeYAzzMsrLh93Rpqjnav7jlb5lwOXa7oJsNIYSNhchSnIZ7UDli1srXycoHL9IfeHNvyur5cGqTsIdcStlah6vluFhPZiPfVRJeIgPXwrSZ59ZizvTV002C1zUNe8TXbE4a7ChkW2X8fn1spdnTrnLg1csHvh9ba3m8qpK1l2rQAtAVMC7YAIS0YJRnF8GAiB5K7HgMAr4iq4h4bFe1lpx/VEbaaZBqx05HrY0EGGAHYovo+hVkW2BEnSc5TKqVK76kJHEvZ4Jf2rM8s9M1gG7wrLr9HIN1W25IEd/W7pK8XcXzXqn6Eij0YUqaVjpLZersweq+dcLVv3CxqhpNL3RBvAzM937aTGfJ3es0jnI3NBo/PPdf32qvRo52NGbjX0vXXd+6cP9N4+v6Nzmq3aBcV8sMGUZS8x29QApVFanz1CmsLtek6tfp2RyuzeLjUy0ubYJEgN9SypqOsvNBuMqhZKv9Lz+XnWg6f390U7VJE/W1y+TNbpnjMri1nJGkGj+f1q7LTzuiCX+fHArErPj9RRbtJzZLrY1/1Y1uyXd+EOG5T7VmgXKo+7dQFmzQNqjLQROtZEqWN7Ti31Rutu9L4Bpf9O5aeeV+fCW2aV+9BUG0BMjBJ2ocp39+7ixM3UOQ/W8bfn5btOlvZztI4jVV3PvFDP3EdpQ2uJykidZdmscm+0f4Tq4GHV289VJ+cO+oh+pc/fnlbq/aYuZs3QBbphL4289wzej/Z2rDr6VqVS5YVcMfe0yu1LwWA7qDo7KjDhlYFzx08v912pfH9a+f287rpZz6X/zFC7kBngTlKdsGSBNNynczl9Hy+32UrLS2p/aO34FuZIZJdkFpwyt9i1pS6+Vam5a3ppwkmBTGe1hY/oHWcp35Oi+iwsf8La8PtqUdG+JGpbBQnQy1Vfp88oL77gmULXbUVMYmsjf5nypCSYfaICGgKEV9NZcWCmje7MKLDqko5MwnG2cR2dH7LEOAnUZkK8uQeNA1uREG+uZ4FZfDceD0YiLzBYKp0GnIksRItvKz5iqGHtqJFklDoIS24jzGubvHHgKtvT4fn8mATk2ELXoSYIe50pPB0/54E6R2VHHbwwXjZqvjPe6WqOSDt7Rmkx4e3RZ1sxewaqwOgCQabLElU0q13DuWZWe6HwymFhedSVM82CBhUYPGLGTVhesPstcmH9x25UKZWNTf6hLQadefPbJm3qXHX2rFIvgj1y2LYX+g4NAvci6eQ7Nf7cwkLnUf3Fws2bvvZifU8Ndb73Whc2xJ+ELniOPD3XNm6jXJjfTY0vGpmABM+LMwf6UndSwm6XNhOZh6E0m9DapdVS9XeYHtEFAvuSqPwaNF417MenVnrQQBvwNzNhXLFodp1R6mOrgjoQHs6EG/vvn8tP2C1vrd+AOEJogtLM8mIif41MuAMlqJ6eJvOHz9RbYILtw7y5feby3oUvjJ48pVoK99Yzi94eUD5fUD5vnds2bsO8y+2YbAdz1Q7UAfPTBnpve67YuXXemZ24VxmTHDCvVm31oQ1jTkScsBJcO9N2uI4CjcooqNzrr6vtOWvDGZXeaKs52k6u4vZV+f9pZrnvsRfbJm0yELqdlK/LGVxylWCPqTpWO1L11kBZmCA3fkO1v6h9h9VdeSfndBG/r9ew6UwaqBz2H0Ad81t0NxvVGQkYdfk9B/wYJARJ+k2RUTJNILF4jBkPKodOjsccpyPr8pvtwZcM9uI9aA5oKhkyZFhLQCVioRcbh+B3YC9AfBPNNnRBW2BhFYu3cIbRkdlliZ1y2MAjTWCwEQeayyUiQomOz2pTfuqtmTOSGRd2ZmLKNkOGDOsIUIOJQyBOATUU9Z6dm+J7TsThtQViGtAWEBhMfeL9xUZsUCUHgVDCJMEJilBC40JFZ7Ylcg5lyJDhjQHCsFmhSVAHjiVsJeZeiVtopi3gwyAWA20BgYGjjJ+pIygrTWAglHB2Et1JlCcmBvs9xn9GLUOGDG8A4OzwYdgsCsPDjT+DIJRm2oI3YZj6QWDEg7Tqzq8YEEqsG0ELWdm9KzNkyPAPBAu/2IWYfSXS1nogFNAMEBJoCnR+hMdQK0kxYdAWMGEQGIQz+yCtuqc/Q4YMbxwwqhMshU8CZyNOx7RdgDE74mHYdH5CqodaSYoJw07STL8gMNA01nCQVIYMGdYUmMlg9SVh2zgmcVCyiMvvjRkHAoPZDR+GTecn1JaVpJgfSaB1ECK+pn7yPkOGDP9AYGJgGjAjgROTqU+i4viNUL9uIwkfhk08BUFU7HRFgFWGDBnWEaTNQKwO8FxWRjLFyV+26yfMlig5lpSvECko4IgkUOp3Yn15dYYMGdYmjPk/JIjDot6atdMAAAAASUVORK5CYII='
codigo_base64_luzitic = 'iVBORw0KGgoAAAANSUhEUgAAAGQAAAApCAYAAADDJIzmAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAALEoAACxKAXd6dE0AAAZxSURBVHhe7Zh7bBRFHMdndu8BbS3vSilRkJcWaTGSYAhR4S9FCtGAEQ2pBh9FRGzLUaTRUlMI9O5aoIRoSawSUxJAidgCvoKa1EcQiAgV2qCGQLCIWFoKvbvdGb+zN3d9knDXrhgzn2SyM9+ZnZ2Z38xvZpYoFAqFQqFQKBQKxb8MlU/bOD/3/gTd0TJf07QUztlXt+9t/ElmKXrBVoNcWjAhjTHyJT4zSaQ5JyanvGDkR41+q4CiB5p82oJpEk/EGAJKia5xsr5pYfpIKSm6YatBNELvk9EOKHURw5gqU4pu2GoQRshpGY3CCTeY3lNXhLHVIEGqe2GBJpkU1jARtqTuafhNKopu2H7K+nPepFHMwbIp4an43MEReuOndDcMo+gV2w1iFx6P5zaTcx+W+A6fz1cn5T6Tl5c3A48nQqFQYUVFRSCsdrB8+fIRLperyOFw+DZu3Pi7lPsNW12WnbS3tw/AbHoa0XFhpX/glD6Eg8cziYmJCSKdn5//St7KlaVWJoAxxiB/MQyWIaV+pU8Gufv9+mFTqk69llnV4M94tyGL7Nqly6yYWPNJZlphTfrqwtr00tdrJs+S8i3h3NmzpZyxyRs2bPjbEiidC8PfYcWB3+8/zExzbFlZ2T4p9Stxu6zM906nMU7rUMGdlsBxDyd8+4nnJuWgE9zSboJVe6dMcbnMr9GUISKNyyMmKS8qeay+RCSF1huW63C7f8X3l8Fl7ZCyRUFBwRjDNNdj4Dzl5eXnhZabm5um6boXg/0WY2yY7nAsswp3glH6gcZYO6JPBoPBVVgNJejLAnQtRCmtQ2PqKOe7oW1CPWthlAaUpaj7QU3Tnoc+GWVaMcvfQZuqrUpjJO4VYjKyMmoMAVqMe8cLU6tOT5OKBTpD4ZeXYdnvQ8PHSzmK02GujRhDgMujsObqgs/vSpZSzAQCgaF4LHI6ndE6MGCD8FiEZqZwl+sy4/xYJKCRg9DO+TBGE/LHYUY8jvIatO/wzmWEi+jrIej1iCeKepA9Ak+CfmXD0PvR5uuopxAdeNsg5JLIi4f4DIJBxsDdI1MdQDWoNlGmLGCMIVTT1qNDWehEoZSjoJ57ZTQKbJLgNAeOkcl+Z3Np6S/lfr9XBAzAATRiBtr3JtzREVlE7BUmVsBORM+hW2cw47eXe72HwrlhioqKEjAWZQhby/z+l/D+gTKvd+cmn+8zWSRm4jOI5ZLoUZnqBLwW1X+WCYu2trZWNLgGQZxI9oTVztDDMtIB51d0Nz8jU7aBFTsUffkQnTmIwSyT8k3T0tIyHcYajIm2GUlU03fidlkhLcEPdy+WsNUQ+H4DV3PviWfHHRfpCJWVlaGkpKTF8Nvp6HStlKNwjRThZcvPC4S/hvZG8az6q1KKG9M0MfF7B7NbeKWtaLzh1PUcSDEPKN5PQns5nm1S6jNxG+RU9ui/giZ7ABvnYqyX1YxqDx8/W71GZnehuLiYYXO9LpNdWDfn5BnDzTIxU5fAqB6q8enr5tRXyOy4wABZxjQpHW0JYbr80Lx67doKfPNRbNILcZ+4IuWeiFMGIb2eHk2n8yRWCDUYmyelPnPDGfRfJ3LKwmD5cDqKrjwH50GsjEbsWw0Y8Aad0hyszmSUq0RnM1FkNgyGQxITfn4TxlPsExEuImQxQkrw3niv19ua7/FUwSgz8e4juAxexv1jMOoVp7uZ4kKKTX038mdpmFD47o/E7R4oVuYWr7cxXGVsxL1CbjU4QQlv0YLByMXp6ItIwKmpOjk5mWPAchAyTMaO4AgiBq0Uz3NM03QM2MswUABhKcp/Ewmo70WEIMq2ys/gLMy2oVwy8o8GDeNVoaDMFdRh/f4JBQJLYKAalNmB09ZxzTB+cJjmUyIvHtBme2mvHT5RJ/pSdCIV3vaAK/WPajqNhGR2X7lR+zE+1qY9EKshBXvYBbjNICRR3soDvb3bOS8St+rBBBjW3NzcJPZESF3yBbjRJ6KPQ3VdbxYrS8oxc6MO9QuBj1MyiI7LIyVJIi16gEZXuedeXIIPd+mQIoytLovqJD9iDIGwPnx2Nq8dNSGsKLpjq0HgyHtc7mAUzWShsTKp6Ibdm/q38hkFLuuagzuiN2JFV2w1SMDU/dgpjoWP8mLT4O1YIivovAtx/+v5v2Prpi5g+8e7Td48m3E6HKvj+wFZl+I6nysUCoVCoVAoFAqFwg4I+QfWBJ64HHCcOQAAAABJRU5ErkJggg=='

# Decodifica o código base64 para dados binários
dados_imagem = base64.b64decode(codigo_base64_isaac)

# Cria um objeto PhotoImage com os dados da imagem
imagem = ImageTk.PhotoImage(data=dados_imagem)

# Decodifica o código base64 para dados binários
dados_imagem2 = base64.b64decode(codigo_base64_luzitic)

# Cria um objeto PhotoImage com os dados da imagem
imagem2 = ImageTk.PhotoImage(data=dados_imagem2)

base64_image_data_terminiar = "iVBORw0KGgoAAAANSUhEUgAAALQAAAA3CAYAAACl6WBEAAAThklEQVR4nO2deZQVxb3HP7+q7nvv3HuHYYYBFEQFBAZH0GCCaxZcMDEuYBafS/Sp8Rk9UdAY4/M8l7i9xEQUoiGaRKPxaDSJe0z0oSbuGmUbERABURDDwCx35t6Ze7u76v3RPQvDIuBDZ579OQfOvT3VVdV9v139q6pf/UqstZYdxAYBohSIRAcs+aUraHrhdVoWLib/1nKKa9ZS/HA9QWse6wc7WlRMH0YcjdOvHHdAf8pG7E5mzAjKP1dL/0P2Jz1qz66ExmCtRbTe8bJ2RNDWGESkU8j5JctZ9/BT1D8+h5b5S/DzzRiKKFxAI2gEtcOVjOn7WAxgMHhYfBQp3Gw/yifUUv3VSQw8/giye4+KEttQ2Gr7NbN9gu5R0Pon/8H7v7yLhidfolisR+GiSCGiwzRRenb8JRDz/4Wo8RMRUIINDNb6GNox+LhuJYOOO5yhZ59M9VFfBsAGBlFdDec2FbOtgraBQXQo5ObXFvDO5T9n/VPPYiihSaMcF0ws4JjtQKRL4L6PTyuKBIOOPYoRl51PxYGfAyKLYBtb620SdIeYg3wby398E6tu+g2+n8eRLKIU1phYxDEfD5FISxbfNuMmKxl27imMvPpinPLMRg3qVrP5KEFbP0AcTcuCJbx5xkU0znsNl3JEO9gg7uTF/N8jWmMDH49mKvc7gHF3zyQ7bnSnFrd67tYE3fFUrHvoKerOnIbX1ITjZLG+AeIWOWYnIoJoje+34Pbrx7i7b2XQ8UeEI2tbGQXZYhtu/QDRitW/upf53/ouQVMeR2ejobdYzDE7GWuxvo+jswS5NuZ/8yxW/+reUMyB2eJpmxW09X3E0ay57T7qzp2GGI1SidjEiPnEsUGAUgkkUNSdO401t90HWm1xTmMTk6PDzKh/ZA7zvnkWYhVYDWbLT0VMzE5HKZAAS8B+f/wNg6ZO3mxHcWNBGwNK0bpoGa9MPAZTaEMkATYWc0wvQCmsKaHLsxzw4kNkx43p1Gxnks5Pka6D1jx1J5+PX8ihdDIWc0zvwRiUTuK1NFJ36nSCQlt4vFub3CloGyn9nctn0LjwVRynPLaZY3odNghwnHIaF77GiqtmRq12l6DFWms7ZmJyc9/ktUOnYEsGjMSTJTG9ExFQBrRw4MuPUT6htnM2cSOLevnlN+G1NSPoWMwxvRdrERz8UitvX/KT6GDo76E6XEAbnnmJdU/8FVdVxKZGTK/HBgGuKmf908/SMOdFREk4xNfhyfTerLsw4m+XZ1NMzKeKKAwlVt10R+d3JUqRX7Kc9Y89jWMzYecwJqYPYE2AQ4aGZ16g8PZKREVe9+seegrPNCLaiW3nmL6DBdEOxfb1/OvBvwGgsJb1T/wdwfmUaxcTsyNYlLisf/RpsBansHwVLXPfRJHq/eaGEoRNbXwLO31qXhwdLmDorfdIJFzdAeEiC7Ntb1pRKnKw75sDAdZYtE2Rm7+Y/NIVOM0vzsUr5FCS6PXmhjHtWGwkaaHL609QJHdiyZaSvwFFCk0Zvc3bUJTCGA8/aAMsigSKBGzm4e9xJr5pxZgSLv13fkV3BtYiovHbmml+eS5Obu6bGNrQOtW7n1IRsqNq0GUprDVYP3JMiVqXwuLlYespkdC3pDnZzIRRz2PdRnqsNSjHYcxlV9Dw7MtseP55HJXuaqm7rXjfbFlb+tvW6rM9eYjgmwKJ8gFUHzwJlUqSf2sZbe++h/UsnTdCov+ifEQpfJNn0JGTyY6vYdWMX4OVrnZiW+reW1AKExRpWbgEJ790OUJim19RnzgiYAPESbDPnT8jXTMCay1OeRZTKGADQ3FtPa8eMBXTVgBrEFTX0rAOOhft+uHEUUfeAtaExzrWrRnjI4BoFxP4qGSS4Zedh0omWff8UzgqG7UMgjFeVJ7eSOQb/U1rbE8fXiXhc2eDyIzq0ZpuksemK4RCURYYfPRX2fv26yltaMJ6HulRw1nz2/tZfNF/4ahMeI1B6Kkm4oRFKcGYItVHT2LoOSfx3qw7sZ4H6HBM1wRYTNe96s1Yi+DSWrcUeWHUJJtf9g4ibq9/Gp1+WYz4pHcbxgGvP8r8KefQ8MpLOJKh2FSPxZDIVOLnWzEUcSSLtQEWi8UL8yirwBZLGONjCf8l0gPwCy1YDBYfJ1EBgcEPWtGSxtoAJ53BFEvYwCds93wMRRJlAzDtRQLbjiYFIljrE9BGIjUAW/LwTDOaTLdQDgYTle2mKrGe30OsgiXKoyzKI2jGkfJuVlb0oCdcvrTqZVbffi/LrrsBTYr08OHobJqmN15H6zR+0IIihZPJUMo3IKjIRLOIdlDJBH6hEDXONgpBkcQpSxO0FeltJtYmiGBtiWxNDY7X0MTG9mjvxcvlMLTjpUPHKb8ph9fchE8r5aNHUzPrGsqG747X0MDK637J2scfIZkeRM2tP6b+L8+w29knkRo6mFf2P5aRl12A96/1pEcNZ9CUybS/v5a675zPwOOOYrezT0a08M7lM1j74MNoSTPqhkvZMOcFVj/8e8ZcciXiOojW7HryFEypyLJLf8q6J55ExEVns9TOuIGqrxxE6cN68itX4dU38vYPrgvNl1uuYv1f/86up04lO3Yv/vmlEyk1rA8bFcASoNNpam++gcovTsS0F1lzx/28O+s3kW1so3QG7SRwB1VRXLuOYmkdSappWboYsGiVxgua2fW44xl55TR0NkPujTexvs8Hdz/E2jkPMuzYU6ia/EXeOu9HpEeOZsQVF/DhPQ8x+sb/omXhYupOvQAR3bvlYS2gKX5Yj/KbW8KWozdXOEKUg4iLuG74SnZdRIREv0omPHE3xTUf8saU0/jg7gepvfvnVNTuS+C3M/DYwxl769U0vfw6Sy68mqCYp2LCPtTcejVefQNzv30GKpXgC3//E1WHH0LdudNZ98gc9rnrRhIV1Rhbovrrk8iMGUFAG+lRe7LXtT8gMbCKeaecRfMr8xh/3yzcbH+MaWPf+39Jv4n7UnfWdBZfciWZvfZk4DGHYQlACVWHHUzNrKtoX7WGJdOvxGvORa/28EcwtsS4O2dQPmEf5p/8PZb+6FpGXDGN3U7/NzyawhEXaxFx8QpNrLz2FsbOvpYv3PMAlQcfiMWgSOKbFgYfeRT7PnQ7a//wOPOmfpfG519j11OnkNptMAEFMmNGMPDoSeFDVJZiyGlTGXnNRbx/++9ZNeO3UZ16/+yxIAStBZT1ffpChYHwSewe90MJvm1l4NFHIMkEi866hLbFK1kxeybNr85nyGnfwis1oMqSrL79PhZd9UPq/+cZDB7iaj743Z9ZfP2VbPjn86y+/T5UNk3did+n4cUXWHndrVhryY4djaVE0FrAFEtAODzW+Pw/qTt/Gs1vzGX5j2eC4+BWVdF//OcZcOShzPvaGWx47jnqX57D6tvui84FU/LQZSk2/O05Flx8HvVzno3WyAmgMLad7J6jqD72MN48ZRq5uQtZ++TDrLr5DnY9dUoo/I7+jrUoSbDsyp+y4NvnkakdxcTnH2T/R+4iOXQgBo8RV17IB3c+wJKfXUHrW4t5d/ZtFN5eGXX6BFMs4be0hp89DxsEvPvzX7Pslp+QmzsPofeboiGC9T0ccRz6lKi7IRLamqk9hpIY0J8D5/4FcV1MqZ3M6JE0vzQPhYtpK9L86nwSegCuU45XbEK0Jr9sJUolUTjodBnF99cSFNtJqCrEDZeddSyb74jhZ7FIIkHpw/WIOGFsEscBE4AxZGr3om3VaoofrCPhVoMR3Kr+nQIicqJpeul1XF2BqyswpVJURth5S+05BEQY94dfIKKxgUdi8EBaFixGU9ZjLFyh3Sxr//hn1v7xYSoPPoAJj91JzayrmfeNs0gOGcTq2+8l4VaibQbXcZGE0/0mdnaGlesSFNrJvTqflDMEZfvSOlKLOC6Ozmbwmpo2O2HRq+lsqYWgNY+3vpG6k6YhKRexQtDWTmnDBpxEf4znIUThp8R0nq+S4eiOJQyUI64TCsh4XSaY7VZeBx1Cj0Ypur8xSusaSQ4ZjE6n8Aqhze835zpHGZQ4WD8IndKD0F5moyIEv7EFUcLbF/837R+uQTsZTFtbZJp09/gVjC0ReI1oMmhJsvalh3nnP/dm5PU/RND4uRbSI3bH88KOacnPb9oJ7XFtYWsXdN2r3o6Ew6tuRRaV3GUgEPQ5LztxHay1aDLUP/40bnUlVYcfRG7BfFoXL0MlXbyGJkSFw2Y9r08cHYah6kCFcSC6/7jiOl0x2RwnHGojDISySRQfpXDKM2x47jlMvsDY2dfglGdIVQ5j0AlfI1FdiU6lOx+ccB1cj1e5NWhJkVu0iLaVqxnynakU3llBbsECrAkImluw2K4hPdop33tv9pk5g/Ka0VCuqRoxkWEX/Dstb7xJkXr+9cAT7PHDsxl0wGRsmWWXY75O2cg9cCrKAQFR4bVB1C9x+pwWOgY13Kr+OMmhg8ktWYQjwseIrPvJYixeYw7r+WhSFFatpO7Ui6iZdQW7Tz8DcVxK9Rt4/YsnYo3Bb26Jxlg7fijBb8kTtLVHxwTTXsLPtXSVYS1+Y67THPNzLZj2EoIiyBew0m3s2Fi8xmYQhe+1MO+4s6m94wYOnvcU1vNZe/+jJIftwtAzTmTV7NsIWvLYyB7fCAsiGuO3seCE7zH+gVv40tKX8At5dLqM+cf9B80L5qPE6UqvhcovT2TIaSdQamjEzWRpfXsFb517KQlVybs3ziY9ane+8MwDFFa9R9uy93hv5l0MPfPbvPuLXxG0tOLnWsPsggC/MbfVuBe9EREhsD5lI4YhS6ZfbVfcPAtX9+879pISnEyaoNCODXwQRWDzuOkqMjV7YQpttK18P+yIiaAzaUx7Eet3iNqi02lsEGCKxTDLRAJxXYJ8ISxDBJ1NY6IydCaD9TyCUhFdlkaAoK2NDrtYZ9KYfBvWWALyuIkKMrWj8eo30Lp6KYnkQFQygZfL4WTLMcVij4esO2HrqyQVhpjVmvZV74fCs5umAyEzfCRuZX9KGxoorFoJgJYUxobj5dnho3HKs+QWLsJQJFU5GK8ph3LdzusWpVCZsvAe9NaJts0gjsbzGxlxwQXImjsesHVnXoiWZN9poQltz9CejMyAyJ/BELaigttpb/ZMGx4LW6Eum9RuMjPW/byuz6rT7u1uz3ZPG9YlFJKg0boME3iARdCbrU9PwpnOAEMRCyjczc/aRVPnhlKUr0aRjGb7IlcAkcgPxqBJRfUrRR6W3a9703vQJ5CwLzH+dzNxKg7aH6esAtOWB+k7awlFNvbdDoOwOzjihnampctvQTb185ZNrlU2Sdf9e9fn0BkmLHTzacO6aBzJYK0N47F1O2dz9elJKEaFknRXWZs7x9roJZEMO75ROtttaA9rUSoJSBgl35huM8Pdr3vTe9DriWZm3XQ/Kg6ZgMqMGU6//cYSSHun+2GfYAs/rjUmfF12//uWhPBRx7aUx+bEtZnvG4UZ7n7OtgrG2nDU4aPCFVu6XFu3lM7YjV1st3ZtfQhRCkM7/SbsQ3rkHuGawurjDsfYLdlzMTG9G4tP9dFfAYmMwMEnfJVkqhobeLGmY/oOItjAx1GVDJw6GQBljSU9ejhVhx2CTwFRfaxDEPOZRZTClzyDjjucTM1IrDGojth1e1x4ZujJFceyi+krWIOyLsPOPz36blGiNdZYqo44lOrDJ+GZ1o+1T1xMzCeBaI1ncgw6+mtUHXZwOHKjdcdAatizHX3DpTiJDJY44ExMLyZySnPL+rPXNdM3+pOCaBA/MJRPqGXPaefgBbm4lY7ptYjWeEEzu59zOuUT9una0Ri6BTyPxkmD9iKvHfQNmhfOC/dU6SvT4TGfCURr/KCVivH7MfHFP6HTZZ2zodA94Hl0QKfLGHfPDNzySkxQ3Cg6ekzMp4qEq7vddD/G3zsLnc1Ex7vM443VGpke2XE1jL/rZtCABLGoYz59lAIVgCOMv/dWMrWjwpX0PbS5iVIl2mFo4NTJ1N7yU/ygNVxdHIs65lMi3Dc+wA9aqb3lJww8/ohwp7bN7Cy7WZWKE+4FN/R7JzFu9kysNqF3VtxRjPmEEa0xpoRVAeNmz2ToOSeBH3QtSuiZfus7yYa7dtY/MoeFp30fL9fctfdKH3NiielrCOIofL8Vt38l4+64eYtbuW101rbu9d1a9zZ1p02jcf6ruFTEe33H7DS69vpuoepzE6m9cwbl+9Z8/L2+O+h4KoKWPO9ccSOrZ99DqdiIIxVdjuRxix3zcYhWn1tj8G0rjpNljwu/y8grp6MzZR/ZMndmsy2CBjp3GQJofmUeK67/BeseexJDCYeOpfy208E8JuYjiRb7ogTjewQUUCSonjyJva65mIqJ+wJss5hhOwQd5hyuhOjIfP2T/2DNr+9l3aNP43mNKBwUKUScME2HwKNzYz7DRNFPJZoEscZgbYChHYNHMjmQqqMOZth5p1N91JeBjpU/sl1uGNsn6IieBeXfWkb9I3Oo/9uztMxdhNcaxqMQHBQuoHrEk4j5rBEGwgyAAIMXBoPMVNBvv7FUH3MEg6YcSaZmZJQ4bAh3ZKh4hwTdWckgiF4ZXQUXlr1L04uv0zLvLfJLV9C24j28DU34uZbeHX86ZqchjkZnMyR3qSY5dFeye48kO34s/Q/9PJkxI7sWlUTL1j7O8PD/AjLwq7aEq13tAAAAAElFTkSuQmCC"

# Carrega a imagem a partir dos dados em base64
button_image_1 = load_base64_image(base64_image_data_terminiar)

button_1 = Button(
    image=button_image_1,
    borderwidth=0,
    highlightthickness=0,
    command=fechar_aplicacao,
    relief="flat"
)
button_1.place(
    x=556.9999999999999,
    y=250.0,
    width=180.0,
    height=55.0
)

base64_image_data_email = "iVBORw0KGgoAAAANSUhEUgAAALQAAAA3CAYAAACl6WBEAAAU2ElEQVR4nO2ceZRUxb3HP1V36Z7pHmZlGPbdBSOogLigqMEFCRIENW5xicmLHjXReNRoTGJ8ajTPqFHjkkRN1CdqjBLi8kxEkLhgcEERUTZh2AaYfeu7Vb0/qnt6HAFFzYn2ud9z+sz09K2q3+9X3/vb6vaIg37WoYkRo0Ag/9MCxIjxRSImdIyCQkzoGAWFmNAxCgoxoWMUFGJCxygoxISOUVCICR2joBATOkZBISZ0jIJCTOgYBYWY0DEKCjGhYxQUYkLHKCjEhI5RUIgJHaOgEBM6RkEhJnSMgkJM6BgFBfs/LUCMzwcBCJF/rwG9nW+JCmGuVZ/zG6RSmLkitWN5pDTrbE+Ofze+UoTObd7n3ZQvG3KE3BUCiCyxwgj8wPyUElwLXPujxBZAEEKoIOl8Pjk7fPBDTUmR6NqP7uuECto6NMUJYeT4gvbq09roU6cc4pMv+UxjP2ne7p+Hyhj0s8zzWdf/LGPFp5hXdPsZRuAFn36NnIds7tC4FuwzWDJtP4tJe0p69xK0dJrPhTAetSOAsUMlFx5tG2J8BpJJARkfDt5NctEUB5klshcYByOE2Z+yYsEPpziM7CPwgjwRd8XOPa8VmHWC8JPn+URCS2H0D1X+/ScJkfMeuWsj9dFxIjevzhq+x7y5sVJApPPepm+54PuTbRLd4kpuXE/5hNj5vuXW/9i4nelFXi4wsn1Mr2y4DdVHbdDTLpHO22ZYteSsSTa2lSeAlHm9uqcVUkAQgS3hkqkuD1+Q5GfHO5x6sM1FU1wePj/BVTMcXCvvtTt9zaQ9bcYPs2j3dFf6sSsQArxQc8ReNjMn2MhspPzWgTb7DpEEkZF1QKXgnMMdBlaKLjlzeuTm2d7cObtq/VG7WhK8EKaPs5g0yiIT7piD8AmEFgLaPaN8OmkM1OZtP4cKVT4c+IG5o9o9o3TSNZ5VZ+9kBbRmjGGLE0bgTj+vUKcHmcCs5dqQcGDNVsVBIy327C9ZUacJQqNsm2dkSSeNDJkgu+mhkWVHIaotY8ank+CHWT2zZNmRXkGUlStj/pawjdw5WyltvKYlIZUwHi0nT6TMez+ENk+TsE34X71Fceiekt37SlZsVoaEAlo6NEob+bzQjJPSyJC04e5zEhw2yuKmpwNOus3j5NsyzLw5w0UP+hy+l8Ut3050rZt0BEN6wzu1Ci8w84Tqoze80vlULlKgVPZn1umEEaSTgpuf8Zl1s0dDuyZSmlMOttncpKlv1RS58O56xeTrMix8X5FOQFO7Rqk8f/zw46T2AmPHNs/I5ljZ34WxdYenOWGCTVO7ZmvLzsPLDnNoKaDD1xy1t81pE21SSUFrp+bhl0PmvRvxixNcnlwcsni1wg81Jx9kM6hKcN2cgIuPdWnPaBatUlw8xaF/hWDuGxG/fyFAZIuFcyc7HLGXRcIWbG3R3P73gLc+jOhXYULa/QsCxg6zOH2izd/eDNnSDLP2t0HAL7/lMn9ZxEsfREwfa3PqwTYlRYKmDs2Nc32WrVcM7yP5n1MT/HlRyB8XBpSnRFcY9kM4YYLNzAkWriWob9PcuyDgzQ8V153k8KeFIe9t1AQRfOcwExFufibgiukum5oUa7dpvnOYQ99ywROvRdy7IMC1zQ1y8bEuh+5hkXRgY6Pi10+HLF0fMaq/5LzJDrc8G3DsPjYzxts8/HLI9HEwdR8bL4RfnZLg2bcj3l4bccIEm+PG2lSmBRsaNb/6W8DabQrQ3Hx6EseCE27J0NCmGT/cYkSNoK5J89pqxdl3e/zloiTfHGfzwD+N7tWlgnnLIixp7H/5NIeXV0Q8/65CKc3UfW32Gyq5+vGA0w+xmfw143UzviHj3Dci/vKviJ8e7zJnccigKosTJtiUFQvOnGSzqk5z7/yAWfvbjKwR3PpsiNYwc4LNjHE2FWnBtlbN9XN8VtVpXCcfoS8/zuHtdYrWjLFrabHg0VdD7lsQcOBIi2n7WQzpLZm5v80+gyUPvxLuMPxu10Ob8AIDKgRXz3L4y79CzrvXY/57EUGk6VUMR+4tyQTGe2tg0p421b0k7R50epoLjnY4/RCbe+YF3PpswDmH2xw12mJLi+bHx7kcP97mvgUhlz/is6lJcceZCfpVSLa2aEqLBbeflWDMIMktzwQ8vzSiulTQt1ywsVHh2tDUoTlmtMW1J7n8Y2nEhX/0qG/V/HymixTg2oIOX7O5WWNLI6MUxmOOrJFcdpzDHc+FXPgnj0UrI5PSlAkO38vuikpSwOGjLEqLBe0ehErzg2NcvjnO5q7nA+6dH3DukTaH7CGpb9Ncc0KCI/e2uP25gB8/4uOFglvPcOlXJtjYqKkoEdx5doJ+5ZKbng5YuDyiTy9Bv3JBbb0iYcO2FsWJB9hcOs3l8ddCzr/fI4g0V0x3aO3UTBljiHfxgyY0/PbsBD+f6bDvEMnVs1yumO7w+pqIp98Kmby32d5UAirTgg+3mOiRdARHj7G6omyoYNKeFgMqBKHSvL9RMef1iPsWhCQd2HugxdJaRUVacPRoSZT17rv1lTR3aDr9vPc9dA9Dvs3NmlMn2vxoqsOji0KufMRn3rsRHf5H0yc/BBD8ZIbLMWNsbnraZ/YrARcf67DPYElbxqRknZ6mtVN3RY9d9tC5Fo8QZsKn34q447kAWwom723R2A7r6hVCQklSUFMmeHJxiFawuVnjh3DP8wHvb1QoBd8cZzFmkOSNNZLjxlqceVeGV1doUkn4+eOKA0ZaTBljcf0c44kGVVpc8YhPu6dxLMH7mwKOG2tx5z9CFiyLqCkT3PrtBDfM9blxbkAqAS9/IPmvrzsUJQSrthhPFSlIJQUqm8fnKn6AUQMEr67U3PZ/AULASQfabGhQ1DWbdCiVgNIieG+jQgrY1Ggq/F/+NaCuWfHS+4IZ4xSjB1ms3aY5ZA/Jib/J8N4GjWvDJQ96PHN5kkP3sPjd/JC6Zk1ZseDKR3yU1kgpuGdewDH7WNw4N+Dd9YqBlYLzjnS4fLbH/QtCihOCN9YoZu5v4dowfazNP5ZGLFmruOecBOkknHW3x9ZWzV1nJ8hkU7vaes3XBkpsCwZVSbSGdfUagalFvBA+2KywLbCloE8pvLLCFCzLNiieX6r5xn4WowZYnH+/x5trFZO/ZtHSCZsaNUtqI6btZ1HkCH7wR4+aMkFZStC/UvD3dyLKi+HsSQ43zvV5cGFIVS/B4jURqYTAyXY/LAleoNnQqPACyXVzfOpb4Y01EcePVxyxl83FD3ocNspizCDBxQ/5lCahV/GOk+jtElprk7vWNcNVjwVcNs3h6DHG6E8uDhlWLdjWounI5jnFCahIweo6Q4Qxgy3eXa9YVaeo7iVo7DA5byYwFXltg2LFZk11qblhMj5sbtIkXUE6CeOGWjz1ZkhrRlNTKmjugHHDJAJoaNP0KxcM6S2oLBGMH2Zx19mSdBIm7m7xm2cDOjwoLc4XT7kcWGmTj9fWK6590ucHxzh8Y1+bG/4a8MySkGHVkvo2I6dlQXlKUJEWrNmisS3TKXhrrWJri6KqRNCaMd6t09eMHWqxuk6xbpumd0m2kI5gW6tpYRW7MKqf4KGXA0Klqe4laGiH/Ydb2ZaapqZMsFtfSTop+PpeFhNGWJQWCfYfIbnhrwFFjmBYteDJZ0KG9xFMGCE58y6PjA/VvQTVvQQvLIuwLUE6KWjPQGsnjOwjaWzXtHmaUMGIPoLmDk1ju9k/x4a+ZZK123ySrqDT18wYb3PTaS5XPOLz2ipFKgGDKgWN7ZrGDk2/Msme/SQvLlfUlAlKiwSWhMqU4J1axejBFrYFL3+g6FsuuiJDrkGgs/thWYJ9h1i8uDxiW4uxS4dv9s4LTddk7FBJbb2mvBgq0iLr1bePHRaFSptJn3ozZPqvMyxcHnH1LJfeJYIhVZK19Zq2jKapQ3PQSIsiV7CqTpFOwuAqeHudMulHAAlbMGqAxbvrI2wLHJlvw7R2QpELQ6slyzdElKcE5Sl4u1aRsI0BMoFmRB/JtlZNQ5vGC/OVcWO7pjUDa7ZozrjL44nFEUkHtrZoMj4fK+e1hs5AM/uVkOk3ZVhZp7j+ZJfKEsHgKsHqLYp2T9PSAQfvZqGUKUhLiqB/hfGWmQA6PEgnYEQfQ/JiF1xHECoIFLRmzE01uEry/iZF31JJSbFkaa3u0ssLNMOqJRsaFI1t2cIvK29Th6bdM9Hh5NsyPLskojQlSDiCjQ3mps4EUN+mCBVUpEzq8sEmk5LtN1TyTq2i04fd+wlW1mmaO0zIHl4tWV9vQnhrBkYPkpQWC5Zv1HiBZmSN5GczXX7ymM9ji0LSCbNXu/ezqGvWtHRq0klzw7+3IUJpk5L1rxAkXcGqzeYGyHVCckWtl01Rmzs0bRnzWbELNaWwZJ0iytqtPGVkfGed4UNlWrBkbUQQsVMy75TQkYJjRlvsP8KitRNeWGY2raTIeLmaUkGRA5P2sPjeEQ4tnRo/glRC0LtEcOBIyR59JY4FV0x38SPNi+8plq1X9Ck1bSopobIErv+Wy6YmzfNLFbv1NQ351XXGK2rMK+kISosFqaSgfzlsbNRsbdVIAbc+63P/iyFVJcajFrlw6TSXCSNM7tWzzTNjnM24oSZten5pRCphIpJjwcAKSXECDtvT4oxDHZozGktAyhX0rxAcPspiaG9J0oGfHu9S16J5baViyTrF4CrBjPEWWptNuvGUBMs2KBYuj9h7kCSKNLX1CtvO61WcgPK0WXNApWDNFkVju8YPBTfO9Xn45ZDBVZLylPH4kdYUJ0xKUZ4SHD3aprQYzjvKQQhYWac44xCHgRWmsOpVBKVFhhS9ioxtXQeG9paUpwRjh0h+OMWctnR4mt37Se48O8FL70csWqEYO1RSnjZEHVEjWF1n0knbMlE5nRRUpY0ew6olbRlNuw8rNikiDd85zMKR0KfU2M8L4XtHOJx4gE0mMOnqgArJoXtYDKgwN9avTknw3kbFopURqaRJ/Xr3EpSnBL2KxU4P1nacQws4ZozFfkMkG5uMR/j9CwEbGzVPvh7yi5kuj19UREO74ro5PpdNczlqtM0zb4WUFgsWr1bcflYCW5r+5fn3e/gRrN6qufJRE+6P3ccinRTUNmgu/V8PP9QMqbLY0KBpbDfFXKQgnRA893bIlDGSB85NsKlJcf79Pj+e7fPfJzjMu7IILzQe7/t/8OldKvj2IcZgC94zRiHbXXEsmLqvxYgaSX2rKUB//ZRPYxs8uijimlkOT1xUxPoGxY1zPS6e6jJxd4vFa0weva5ec/d3jV71rZofPejhOrC0VnH9nIDzJjuccqBNWUqwsk5x9eMBkTIF9tptmg4frKznKkkK5iwOmbi7xewLkqzeovjhAz7XPhlw1QyHo0cXIYRpfZ17r0dLp6auSXPQbhZ/eyPipqcCzjvS4fjxNo8tCqkugVtOT1BdKrh8tk9tgyaVFPz5XyHXzHJ59MIkv5sXcO/8kIm7WTx0fpKGNs3NTwdcMtUUZamkcRpjBktmX5BASpi/LOKGuQEpF5ZtiCh2BU0d8NBLIZdOc2mZrLnmCZ/ytODDbYoghKZQc9VjPpdNczhwpNnnP8wPWL0l4rtHOLy6MuKBf2qGD5N4gWZbK/z+ey4JR7C5yXDEtgSdHsx+JeS7RzjMGG9z34KA2a+ElBWL7R6/ix39f2itTTU5sMp4xs1Nms1NmoRj3H5VicnZ1tWbcFldKmjLaMYMllx7osv0mzKUpwXlxWZj/dB4duMJoCwlGFIlaPM0q7cYL+pkj21z3YguISHr/U0I39piUh0vhCIHhveRBJFmfYNptUlhcq2mdk0Y9XjWIdu4H1hh7vaNDZq6Zk1RwoTEfmWCshSs2mJSqppSk38esZfkR1Ndpt6YoW+5oKwIlm8yPdZEtgXV7mn6lEr6V5gc9cNtClsKbGny1J56SWHWLC0WDKiAzc3Q2qnxAuiVTVe8wOilNLR0ak492OaSqS6n/TbD4tWKvmUCIWBLi6ZvmWBgpeTDrYqWTk0qIdDk9SpPwYYmTWMblKehb6lgfaNJ4ypTpmJWynhf2+p2rhCatKE421vPPadhDoWMcWsbzKklZGsQafa5osREvZZOzZYWsz+VaUEmMHw6faLNiQfYzLrVY3i1KRhXbDbFq2tnD6kiGFQlSNgmMkV6J6e0n/QPz/3QTJrblNzhSBiZl2Pnla9v05x/lMuxYySn3OERRWZsrkebCxUyO97Pki9hG9bqXHNf50/KugTNHhIEoSnYbJk/sPBDM96x8jloriDsTuZImRvptIkm5ofKjHFsurogfmSuS2RlDrMHIhOGS/pXSJ5eEhJm9UrY+TwRzPV+aNa2pNE71y3S2RNPq1uR2mULlZfXktmDjyh71JvTK2s/S8LX9zJ97iVrTT4fKpPPCwHNnZr6VrN2zoN118vJkjWMzJqOZV5BlCdJz5Ceq1dUtxO8nF1zR/Y58uXsoPVH9ZDSrNv9pDMTwMTdJOUpybNvh112SjjdHqTKdqa8MP9IwYMvhTS1mxZkT3ziw0m5ybs/7KJ1/i4m68ltaV5JR/P6h4rOAHol85vZ3Ui5jSmy8vPRjezbO7bW2oRqy82/zxk46eYuyo/b3oMxloT2jOaRV8KujdE9xuWOhXMElNkxvUsclqxT3DMvoCybx2n9cTm7Hyt3/7zrb9uxce5ou+f13R/Iye4tGrjz7wHf2NfiuLE2I2pkV6G1fKPin8sVb65VWD1s2FOvnmvu5DS5S+6eOvSUkR6f71CPrCydvqam1GVDY8gDCwPK06a92tOu3edQ2uzH9sgMn8JD7ypsmX3+4gud9YtDrp22q0jYRq8gzLed/hPIHfi0+5oiV5BOmL8HkTkmjrJP1H1Jzf8x5OwaZqPsp4Ft7fjSL/zx0SCiK0x8GSGgK9fbFeSOzRMOXwq25Iqids+8zz0Xk/PCXxXk7Orugl13dtkXTujtPU31ZcNn2e/P8szyvxO5/NjuFnq3lwJ92fFF2/Ur9YB/jI/jq0bgfzfi7xTGKCjEhI5RUIgJHaOgEBM6RkEhJnSMgkJM6BgFhZjQMQoKMaFjFBRiQscoKMSEjlFQiAkdo6AQEzpGQSEmdIyCQkzoGAWFmNAxCgoxoWMUFP4fkRGU59mKAgQAAAAASUVORK5CYII="

# Carrega a imagem a partir dos dados em base64
button_image_3 = load_base64_image(base64_image_data_email)

button_3 = Button(
    image=button_image_3,
    borderwidth=0,
    highlightthickness=0,
    command=abrir_email,
    relief="flat"
)
button_3.place(
    x=23.999999999999886,
    y=464.0,
    width=180.0,
    height=55.0
)

server = 'SBs2019-ISAAC\ABMN'
database = 'aTrans'
username = 'Bds'
password = 'olivettiBDS1'

# Criação da conexão com o SQL Server
conn = pyodbc.connect(
    'DRIVER={SQL Server};SERVER='+server+';DATABASE='+database+';UID='+username+';PWD=' + password)

# Criar uma consulta SQL
sql = "select Nome from aPlatImportacaoConfigBase where Activo = 1 order by nome asc"

# Executar a consulta e obter os valores
cursor = conn.cursor()
cursor.execute(sql)
valores = [row[0] for row in cursor.fetchall()]

# Criar um botão "Selecionar arquivo"
combo_box = ttk.Combobox(
    window, state="readonly",   values=valores)
combo_box.config(width=50)
combo_box.place(
    x=490,
    y=182)

combo_box.bind("<<ComboboxSelected>>", selecionar_opcao)

canvas.create_text(
    45,
    127.0,
    anchor="nw",
    text="Importador de Despesas",
    fill="#FCFCFC",
    font=("Roboto Bold", 24 * -1)
)

canvas.create_text(
    490,
    127,
    anchor="nw",
    text="Selecione o Fornecedor:",
    fill="#505485",
    font=("Roboto Bold", 23 * -1)
)

canvas.create_rectangle(
    40,
    160.0,
    100,
    165.0,
    fill="#FCFCFC",
    outline="")

canvas.create_text(
    40,
    180.0,
    anchor="nw",
    text="Aplicação de suporte a importação de despesas.",
    fill="#FFFFFF",
    font=("ABeeZee Regular", 14 * -1)
)

canvas.create_text(
    40,
    230.0,
    anchor="nw",
    text="Será selecionado o tipo de importação através da",
    fill="#FFFFFF",
    font=("ABeeZee Regular", 14 * -1)
)

canvas.create_text(
    40,
    250.0,
    anchor="nw",
    text="seleção do fornecedor desejado para fazer importação.",
    fill="#FFFFFF",
    font=("ABeeZee Regular", 14 * -1)
)

canvas.create_text(
    40,
    290.0,
    anchor="nw",
    text="Verificar sempre os resultados obtidos antes de efetuar",
    fill="#FFFFFF",
    font=("ABeeZee Regular", 14 * -1)
)

canvas.create_text(
    40,
    310.0,
    anchor="nw",
    text="a importação no Atrans.",
    fill="#FFFFFF",
    font=("ABeeZee Regular", 14 * -1)
)

# Código da imagem em base64
codigo_base64_icon_luzitic = 'AAABAAEAAAAAAAEAIAC2IwAAFgAAAIlQTkcNChoKAAAADUlIRFIAAAEAAAABAAgGAAAAXHKoZgAAAAFvck5UAc+id5oAACNwSURBVHja7Z0JeFT1vf5/WdSLLJkJoFIRI9kn20wCAvqgyL22Uh8lyYB6a7VqrXW3tqxZMGpt1db2/6/3XlEkk4grdQGSmUmoii1arbgrkAUEKySTAKJVIJkE5r7fM+fkIuIC2ebMeT/P8z4nRAkzc/K+5/vblSKEEEIIIYQQQgghhBBCCCGEEGI2QrOV+vDCiWpHYZJqL0xT7UXpKlCUpgLFaapNV8AtV3y/OEV1fv8c9fHssfzgCDEjgcIU1TozpcfcbTD7vmsT1I6i045FAIxpL0rLRAA4EQCT8N+nQPkIAAcC4BQEwJAW/P0298HhkKqJEBKhiMk16ab954xxMW3FqWPawwa/AXoQqofegpqgrVAL1Ap9BDVD70IvQB5oLjQNP2NcwJ0S3xMms1LU9vOzVKiCnzkhEWD8VDytMyBcL9JMmgidA90HvQZ9Cu2HQkeoA9AX0DvQUugiVAxjPpmTgH8zXe2emaV2TzmHQUDIYPDx7Cy1E+31Vjfa7YXJEgAnwaQ/hV6Edh+F4b9Ne6DXoflQWmjSeVoQfOb6ndZvQAgZqKe+tMuL07Vr66wMG67X6E/7ff1g/EMl1cQGqKTNnTa2HQHUrjc9CCH9SGjaNNV0+Ylh4xdlxKPNjzZ62ooBMv7hguBV6JK2wrQhWijNzFSthRm8UYT0NduLMlTgklNQ8qfKcJ089RfqHXihQda/oPvb3Onj2twy+oDXV8QmASF9RktRqtpemKratZ749GQYrhrqjADzH9xhKKMHp7e4Hdr8ggCbBIT0nvbZySqAJ+su6ekvTnfqnXyhCNX70A+2FqfFBLRJRem8gYQcLbtg/p0Xj9fa1jDWGdAbEWx+QzK/wL3V7dRCoH0WKwFCjhjp8Ns6C+YvgoncaXkw1ToTmP/gEJjRUpSh2otSVFsh+wQIOSIaf2pXgVkp0u4fr7evQybT+jZ32uS2WWi6FCbzhhLyXTHm77cVab39j5jQ/IZeCrjTTtXej5tNAUK+vd0/QxbxSOmcGafPx+8wcQDI6MADbcWpQ2WqcoD9AYR8My0XpYeX5halnQ3zbDex+Q19gffyY1mz8P5lOap5BvsDCDl86T/LKP3TEmCcZ6PA/IbWBYpTk9oRbB2zM3mjCTkcYpBAeF79VYM0vbc/mwK/3lGUESvrBrjZCCGHPv1180MnwCxro8j8hja1Fadmy2zGzy7I4w0n5MsBkKraCrUVfj+B9kZhAEgVcHv7zORY6Q8ghBxc/uPJuCPc9vdFofl75gagCZAkw4JsBhDS0/ZPMcb+z8J1ZxQHgAxpXimbkoZct/PGEyLsQABsnDFZQuC3UWx+Q08h6IbIaAchRCljI8/RMnPOAgGwua04NZ07CBHSEwCyyUdqPkzRboEAkH0MCjk1mJAvVwBXHOXuvWbU3a3uJNU681TefGJtpDf809lJSt/GO2QR1QSKxse3zhzHXwBi8QCYNVa1upNjYYrnLBQAsr34qDbuGESsTgDtf8iu77BrlQDYBGWxI5Cw/R9u/yfp++lZJQAC7cVpZ7QzAIjlA0CO9irShsUaLRQAn+jHl/EXgFg9ALQZgLkwwxYLBYCcOTiDAUAYANYMgD0MAEJUzyQgqzUB5PDS6QwAwgAoShdZrhNQP+eAvwDE4gHg1nbMTdRP97VKAGyGsjkdmDAALstRbRdrOwCvsFAArGsrShvNFYHE8gQuy1UtF2bLXIA/WCgAatvcGfHtbm4QSogc+mlsBGqVxUD3ts3OVAwAQlTPbMACaIcFzB+Eitn+J+TLASC7Af/NAgEg8x0yOAJAiM72WRnqt/MuUVppHP0B8OeAO+X4QDFPCCJEo8WdrvQDNKfp8+SjdjegQHHaT9vcGbzphHypGSAHg7jT5DTg+igOgI14n3LcuWq54Hu86YQYtM9OgSkcxmhAR5QGwF3ts9Pj2tzcCISQQwIg2Tgb4KQo3RxkS1txap6sfWideQpvOCGHshNt422FWhBcE4VVwN07i8fH7SxOQdhl8WYT8pV+gKJ0FSiG3OmyRVhNFJn/rUBxanJ7YZraf1YybzQhX8euwvGqXXYJKk77d33VnNnNv1c7Dgzt/n+6k9SWK5J4kwn5Oj5GeRxAALRcmBwP45ToB2mY+UTghwPu1OHSvxHgqcCEfIemgDYzUNsoRJYJP2HiAFiL9zA+/F4484+Q70zorEStKRAoSksz6RThJmhqqztD7Srmk5+QI2I9mgKdlySrdre2TmAC9I6JzL9Nzv9rv3Cc2jErQ+3kwh9CjpyWC9NUa2Gmai8+TULgbOhdk5j/R9uLHTEyosFpv4T0gkBRpgpICPxnslEJvBzB5m+ALmy/KC0m4A4PaRJCehsCMzO0EAgUpuKJmi5LaZ/S19VHkvmln+LMz2aNN/Y55I0jpK9on+1Qe52pakdhhiwaGgmz3RYhG4jIIR9LoPFaSM1KVa188hPS98gkmh+GwisHdxZlHtNWnH4ejFc3iNXAm9DlbcUpQ2WYr/2i8dqaBkJIPxIusdP1uQLpo2C+m2W67QBNGpLJPc3QHe3utNPatY4+lvyEDCg7ijNRaidrG4q2X+xQ7cVpY2HKG6C/Q5/3g/H3Qe9BiwLudMf2i5NjZWbfF1fn0PyEDF41kK6dMCxHbbdeKkGQLnsLng/9D/S23kbvjek3Qo/C5Jeh4hj3WXFKbKtM652Vot46Z7gKXVPAm0DIYNNelKaFgL7BqNpWnBIXcKeOCxRr24zNhZZBL+lHkMnGnC3QTmgX1ApthTbow4wyyrAIOq9NVvHNSjkuUBTetmwH9NF5p9L4hEQqrcWpqgUKFBuhkCJTimPaCpOPh4lPxdM8D9dJ+uQiCYgpkBP/fzKuwxEcsQFjn0L5GbNS1T//I1+FKvjZEmI6ZP+94I8LVdvMZGPz0R5zH6yAvgjpwyvHqdbZ3LWXEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEELMy5TlH0PblKuqSbk8jSq/qlG7ujxN2tdOfJ2Ha+6y91TuI+/yAyMRyZz6XDUXKvM5VJk3q0flPv3rWqgmW9361Bnq1uVTrPkhuSqboc26wcPKrG5VEx/dHIsAsOPPp8L06bjmIgBy5WsEQBICIDGven3c1D9tUPmecCiIcpZuVLmVDfztIwNKSU2mKlmVCWMbZs9Wd9WNUwv9OXEIgER8LwlKh3LLfJB8XZt1KgLAPm+FK7bsqXRV3vN3HaqkLk+V+nOi88PKfaRB5VaHn+hieOcTT8t1NJQPXQHdBz0LvQq9DzVCW3Q16t97DXoO+gN0Fcxf4KxqPEFd91w4SFAh5FVv1ERIf1Dmx9MdJjWe7JfdM0yVeh0n4OsJ0FXQH6AV0GvQ+1AjtEVXo/69V6FnofsQGj/BNR8BMOpuj+3/qoZV+DdW5UbD0x7lfHWDZvyJD4v5mxPw9Zkw7F3QS1A7tB8KHaHk7+yA/gb9DmEwHQFgy1/ajKqgQeWjKnBVsSogfUM5SvdFvvBTvmxlmpjUDp0N3QuthXZA+6HQEUr+Thv0kvysUm/WOaW+7ISy+szwv1Urcpjxif+uytfLfCdMCdOPgkl/jD/7oZ3QgaMw/TfpE2g19DOXp2GMq3qTFj55jzaqggcb+RtMjooFtTnqdl+G9lReuApteW/2ifj6Smg19MlRGP7btAvywfxXQCeU+TPxb0oY5KgKn0lOa9ZMX6WX+1WNI/QSfy20p49Nfzh16s2InyMARrqq5bU0aK+HkCN66muddtnhq9dhgzF/ppfv+/rB+Idqn15ZXLnIm5kgnYflNQ6tEzFimVaxRo1//lMtAPKqG+P0Uv/P0N4BMP7hgsCHquP7OZ7GeOksLFywTWUtX8/fbPKNXPNGgXr4vrGqFCV/SW1OfJkvazqMuArqGADjHy4InsNrOLukJi1OKpGq352mvcaIInvpBpW/fBuM3yw99GjnN86Ftg+C8Q9VG7TI6WkYKU2CSX9ar3LRPCDkcJSszFa/X5miFq1ySABIO78MCgyC8Q/Vdmie9D0skqbIqhRVsSIvMj60giqYvqpJFUjJX9kwHoargjoiwPyGuqDlaAakuao2agEgQUXIl8xfm6lKpMyug/m9WWkw2+NQMALMb6gTeqzUmz1+od+hFvmc2qjEoOL0bNB62/Mq5Srj9o0vRJDxD9XLaJacPuEhea3NqqCSIUB08/uzVQlK/tt8edLOnqC3v0MRqhfLvdnO8poMVebL0fopBoWJ1RvR3l+Pp7/W2z8FWhfB5jf0HjQtq/oDBECTmoAAI9bmtlUZWnu/YqU2qUeG9t6JYPMbegM6o9yHSgABUOEf4M5B6UzLXbxZm4kHQzlNYv6DQ2BSPqqAYXet06YhE2tSUVOgzmpF6R+e3HM69K4JzG/ozXJvlmthLSqXJzMGtmPw3F9vxRNUM38KtMZE5jf0Sr6nIcO1dIOatTVEJ1iU36w8D8bPlIk+mTDUKyYyv6E1ULJ0Wv7p4aSB+dDyPcaUXm3e/mMmNL+hP7sqG0fJ6IAsQiLWQmb3aQt1arNH4fqMCc1vaFlpeGai9n76ldPR7pfe/vylm+NgoAX6eLtZA0BGB+4oWNpwTH547gJdYRFK68NTbEt8jmNgmtuhLhMHgIwOlNxWkx6/yOtAsyat/z64bH0Fnkvm3nsaW01sfkOynmCGVgU8xACwzNN/hcOoAH4I7TSx+Q21ltVmTS/1ZquF3n6aHyC95jLkB0npXxMF5je0GqE2WoIt38MFRNGOVir7NI3G13+JAvMbWlVe67BDqrQ/hgZdSzcq5zJZ2NP4c5OX/ocqCN1SIO/PwyrAEgEQ1i9MXvofKpmqfE1FTa5a6Mvsa/M3ap1/0BiY5R9RZH5Db7s8TeNcqHIKlkZ3CCxevFj94he/UL/85S/Vr371qy9p3rx52vejtvSv6VnLf6pJxvuPVK+V12aPgbTJTX0aAM7HZF2/LLeNqGm+fVkF3JT0zJaoDICKigp1yy23qLlz56o5c+ZoKioqEtPH6Yq94YYbvhIK0dfzn6PmvZAsAXBzlD39D64Crr5jVWYfB4C2405DYoRP9e2t/u70NJ3oiqJmgBj/2muv1cws5sYT/jh8nYOvL8a1HPpvaCl0P1QCFUMZ0DFGCBjBET3t/2zZwefvUWh+Qy+U+bIS+2xI0Klv1AmdC30axQHwOZoBF4b3MzD/OgExrzzp9af6cGgm9CS0FeqGQodRF9QMVUE/gIbowaH9LDNTYmzM6c2+ENfPozgAPkUAnKu9176YIuysblLZjzfHwBh/jGLzG1qcV7053uwBYJT5paWlYn4X9AT0+deY/uu0G3oISjeqADNXAhIAC/xZsqZ+cRSb39Afy17KiSmr64NmQLj8bxqjb8oZ7QHwNt7rODM3A4x2/MKFC2NwvQBaf4TGP1SvQ9NvvPFGrRLAzzVn+R8e+hsXpZ1/h+pVvNcx8p77JgA8TVOivPw39IWrqnGaTAwya5tfzH/99deL+d3QR700vyFpFvzHzTffrBYsWGDKkQJ9/H8arl9YIAB2l3qzJpf3th9g9vKQcoa38r7hKHfvNaPm5Fc2mO4AEin5L7/8cm2ID2Y9E9rUR+Y39A6Mn2OMJJjqs5HDOmqyJQTmWsD8xm7D18uBJLOXzz76D278ktfVuY/vkAB40CLmFy1zVq2P+eGvQqYLAF2jYVZ/H5tfEwLgsVtvvXUYpERm4cH/SlALvdkxMMWjFgkA0YM/e6JYLX4x4eg/uEnVW9TEqg+PhSnqLBQALzk9G493VprrgBEpy8WUMOrPoWB/BAD0L2iWVAFSaZiFhbUOBIDjeJjirxYKgLpSb86xUG/b/9rsv7csFADva8eTmWhdgJhRzI8QGAWDru0n8xtaqQ8rmubzQXtYJEd1fWChAJANQ8Ys6k0/gB4AmVCThQJgi6uqMc9lorMErr76amP23vnQZ/0cAO1oZkyWf++mm24yRwCERwCcMMVWCwVAE5RR1gcBIFt+bbVQALTA/JPMFABiRinLca3oZ/Mbk4VulX/zvvvuM88IgDdrEtRioQCQsHP2KgD0BUCTNVNYJwDk6LKzzTQXQNr/0FCYcvkABIDoIRlqNEtHoAyHlYfP8ttpoQBoKfdmT4J6UQHINOAqbQ5Aq8UCYFq+x1wVAJQIvTJAAbACGmaWfgA9AKbpZ+9ZJQBaocl90QSQI7z/aaEAkLCb7DJZAOjDf28PUACsRsWRaJYJQXoTYIpuCqsEwEeQq5dNgAaRA2bYZKEA2OqqbHI6K5vMVgHICMC6AQoAH2QzSwWgB4BLN4VVAqC5rDbL0asDRfUAOAWmeNdCAbABSnaZrwmQ0F8TgA6jx/D0P9Y0FYDPIYd+pMAUGy0UAO+Ueh1joaP/4MLn/jUPifJ9AL56hFhlw3DIbLMAZQ3/fw1QANwlC4Ouu+46cwwDavvkOUbAFC9bKACeL/Vl/FupL/3oP7iTS15Ukx/boFyehioLBcBT+UsaYqfebp5ThI2tvGDMq/Rhuv40vywrLtQXHZni81nwRIGsBYiFKZZbKAA8t/0lXd24ZOrRf3BybJaMh+vHfR+wRgA0LMqv+lAVPGieJoDMAdCHAjP7YRHQVxYFQSebaSagHPtdXpsu/QC3WcT8B6A50v6vWJPUuw9PHwmYpi2Vjf4A2AedZ8b9AKQKQBDI/n5/6kfzH4DK0NyINduegWXaOQDZ5+n75kV7AMiS57P7ZFswfT8A2STjHQsEwEazdQAahEIhozNQdgDa3I9P/2Rjj0FTBUB4LoB0BDZYIADeRtiNK/P2wY5A+bJVdmVzPEyxxApLgfM9m47L9zSbsgKQzTruuece2QzkFqijj80vawwukw5H2RjEbHsClHkdCADHcRZZEvxQeV1WfHlf7AgkARCuAhqKo7wZIOX/pXmVG9XJi19WZkRCQJ8UNEIfEejuI/N3Qnfi58quwqbaC6BnJODVKarMny1DgpfCIPuivPwv0ioeXx/tDAzzK1dlw0kwyOtRHADv6nMetJ2QzIqxnbe+NPj+PqgEpNf/bj1UTHteQEVFz7Zgp+D6XhQHwOvQSX16UrBzaaM67ffrY/TRgO4oNL+McNyGpk5sNJwLcNDhHmLaedC2ozS/jChcizAZYmw4Ymbk8Mz5XmesPhpwIArN3y3bns2/H0FX04cBMGXZe6qgqlHlVzVJB9kHURgAzZDD5WlSpz+w2fQBcPAJP7fccouMDEyFlsla/u/Y079d3xK84MYbb4yR9r4MNUoTw8zctTxTyRHaqAIc+nr5aAuAD0p9WcmlaOqUvlDQl/VTSJ1e/ZH65fqQ9AcsjLINQuXpf0fBo82xBY80q6zl61U0ICEgxj2oSXA8rpNkGA+qhd7Sn/Af6bv+vgk9B82RkQQ87Y8zevulczEauOaNAlVWlwmlSRVwZ5RVAbIR6MJ5a6apcm+u1uTpUy5auQUVQIOcoHuyNl02egJgndPTkCRLn6es/JeKNqRsnz9//pfO/MPTXMw9FsoSs8PkDly/h+sxBwVGdJ4P2PA9Y22AbBG2LooC4GW8p5PLfHj6rzu3P3pRQirHs0XJKjlXVeNsGOezqDgHoLLpxwV4T6c985oqePANFe0YJ/wcHAiG2Y2qIZqpWJ6l5jxpl0lB0il4GbQnCsz/GYzvLqt1qLmrp/T9079nSLBqozY12FndIAuE/r/JmwJS+j/krGweJm3/vOoGRaxBqTdTWyAEwwyTMXOTNwWk9P9/5d7sIeW1MvTn6N8PT58arFyVjWNxXW3iAPhbvqcxyaUdfNJEV1guBMKHhZZ6HaeZfLvwepj/ZL2iGZgPL3txk8pfJqfoNk7Qt9E2m/kboDNzKzeqCVWb6AaLUlaTiSemZpwzTbpXwPtltVkTS/Ae7ngufeA+uNwlzSp7abOa8HCDVAPfhz40kfm3QRdMqsJrr2rW9jwg1g2ARdB8X56EwPnQxyYy/xbo+7d7HWrRykxVsTJ9YD+8vGqpABrU+KdflhAo1vbTj3zzb4d+lFu1PsbpadDOPSTWZpGsEpQqoN4lR4f9yCQhIFt+u+8On3eoSmuzB+fDkwAQuZ7QOgd/EOGThBqhmTlLNsc6KxFeHnb6Eb0SQADIvPnb6x0SAhdE+IrB9dCM0hccMfKayweq3f+1nYJV2tkBKvvRjdIxOEnfPizSNg95Baafmu1Zrw1jmmnDTzIwaGaCFj2XKXsHTIXJ1kbY6IC8ljVywMmi+gyZzagpIihY2qgKHtog+wfCXA2yd8D/RMjKwb3QUldlU3LBsk3KKX0WND/5hkqgxJujbvM5ZDhNJgotgfZGyAq/B6BTw+cc5gxe2f+1lcDDG5RzyfrwEGF141BZWgv9YxCrgbdRnVyFJspweU05SzaoXLxGQr6JhTU5agEU3kAkeziul0NvDeJT/03oJ2U+x1D9NWlBFbHoOwihaSAjBE2n4s936ottBioIpDPybmdVU7I0TZyyp2FVI3+zyZFVA7Xh9rXMrivzZifDfHdDmwewWbBZX69wWvg1OAZunL/XIVDZrDI8e7RyO+ehjbHO8OEiFdB6feONvjZ9UB/bvwvB48x7dENcnqdBFTyxFQHAkp8cZQjgabtg5QRtdt2CVVmygCgb+rXeSdgfewt26p18t+Mpnzv/eWesPO0X+Qoi+6l/OGRVXd69G1Ue2t3OygZ1umeTrLUf5/I0XIbrk/rcgc5eml6e9k/ne5qug1Jcno1x8tTPe2SDSlryV21XY0J6w63Lp6i71sapEu1sAVQF9RlxZTJ70Jd1FYz6Z308PthL08vT/knoam1mYn1GrLTzS+pz1Zy1E7T1C6ZFTJh1/zsq3xPeWkx7MlfJWoKGdJTmhVKuQzX6bkNyBFkA+kTvRNwD7da/t1lW7kFe6F597kGGq6r5eOfSBmPbMjVxyTaV+8i7/M0lfUrFmiw19/EMVeYPl+KlPm1fgeNLvVmZUDG+dy/k1VcYiqHboN36YiPpxPsECkCb9F17avRmxUz8nHT8vCGL9Om8slpx4TMFvd/SO9KY5tmi1CMvSgDo24yhjV65RRUs2xyPMBgNE2chKM7AdTo0Q9d0PNXle9kuT9PovOoN8c7KDdrQo2Z6KfHP/DWf9mTgwqCmQKmQCg/FhXcbVvN9mWqBLyu+tNZxAiqEbPy3M/DfpsuYva7p8j2ERVZ5bdaoBSuc8XN9OT3DeRIoVd4z1U2+FGt9mBMf3wJ9ZBxA0lMpaOHgMb7X2NO5mPvIek2ERBLz0UafKzvxGB12ejh8SbJWXwIDTYm5T09U854t4AdHCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGEEEIIIYQQQgghhBBCCCGmJLRmtHr5RqW6VttV0A/5RDbVJfLaVFCkfS9BdUOhl2NUqH4YPzhCzIhmcn/Y5GLqLl+i6qpLjOmqtx2P7yfB7E78t8nQ2dA0aAq+58L/m4IAGNHtt8V2+Ub+X1B4E1TH8yep0HJ+toREJF119rA000P1I+NxHacbfC60DPor9AG0FWqBdkK7oFboI2gj9DK0HLoNOg9K6fKOOM4IgyDCYN8LYxgGhEQCnXXyhLdr2vVAvATACTDq+dAD0DvQF1DoKNUBNUCPQpdLoHTVDIuVMOj2JqpgVaYKvcF7QMiAs2f1KBWqGYoAsKs9vmPk6TwWuhH6O/R5L0z/ddoHvadVBl5bVqh2dGzQa1ehNXEInQTeEEIGrI2PJ3Cn364ZDwEwCqa8BXobCvaD8Q/VAagZujPoTxgf9IebBdI8IIT0I6E1Si1QIb2Nbz9Gb6PXD5DxD6e3On22Kzu9tmHymjq9I1WnL5E3ipC+5oB/pNqyeKwK1o6UYTt56t+hd+KFBll7oIdRlSR31Q1XQa0zkk0CQvqu5K9NUF218oQdIZ18GXovfVcEmP9gyejB1OAzieH5BWwSENJ7Or12BADM/xcZl7edDr0SYcY/WI3QzAM+W6xMMBIRQo627PcN12bpdfmHK308/70INr+hbXj6X3pgpS0mHALsEyDkiAn5lNrlS1Ud0rnmS5gIY71rAvP3hABUuNd3HAKMHYOEHDm+kGacLl9CGsy01kTmN9QU9NmndvlG4H2M5P0k5Luija1rT35bIq5PmtD8B3UM2scH9enJhJBvK/3r5IlpUx01NpnLXzqIY/x9paV4P8MlADp8dt5gQr6JffosPxjm36GAyc0v2gtdJcuM9/hPUKH6E3mTCTkcHbV21eWXXn+bHaapiQLzG3obTYEUqWz2P8P7TMjhA8A7QnXXaeP910CdURQAonu6a21x3V6btkEJIeRg86Ptr3X++W0nwSyvRpn5RVu6fLY8mdX4xePH8YYTcmgAHFilPf2v0tfhh6JQv9lfZ4+TZg4h5CD07bts+uq+UJSqIei1acOCbAYQotNVp2/j5bOdA30SxQEg/RpXd6IZEPrEwRtPiNC9GoZoiZUA+F0Um9/Q00F/4lDZtJQQ0lP+a3v5/c0CAbAFygxyUhAhegCE9+YvgDF2WCAAZGZjMacGE2IEgLavntb7v98CASC691/P2dX+v3CRELE4MjW28x+p0gT4o0XML/LuX50Yf+B5BgBhAMh+f3EwxQoLBcC6oN9+AjsCCcv/8Ow/Wfb7moUCYDOUzX4AYnlk5V+ndl6f7X0LBYCscjyDAUBYAYSHANP1DTWtEgC7oekMAMIKILzzT64+Pm6VAJCzBGYwAAgDwJoB8AUDgBDrNgE+0dc98BeAWL0CsIvYCUiIVQMg6LPbo3QTkK/TJiiLAUAsT2j5WNW9cnQszPCshQLg9U6/bVQnNwYhDICxKrTqeOkHuM9CAVDT7R0Vv7/2JP4CEKJ3BP7EQouBfruvNkHtr+UW4YQYAZAPtVnA/J1Bv31m0JvAG0+IFgDh/QBGwxwvWWEdQKfXlt7Jo8MJCdONCmC3VzsN6DcWCIAng94RQyDeeEKELv3wzKDfNhXXnVFsftnu/Ao59XhHzUTeeEIMOtEmhhJgEF8UB8B6hFySti24j/eckB46auyquz5BKoHL9AM1o838B6CKbt+JsUFUAISQg9hXo58K7LdLZ+DaaJz91yWz//wJas/aY3nDCflKX4DXhhDQqoAroX1R9vS/M1Rji91fk6hNfiKEHIKMBkgIdHnt0hfwTDTtAdjlS0iSswC6nkvljSbka/sC5JRgn1YFnAVtj5K1/5d26h1/7Pwj5BsIrRyluuXsvFXD4hAEc0x+UrCU/v8d9NuHSgCgCuANJuTbCBrzAsLDgtUmDoAXEWLjguFdj3hjCfmudC5PVjJcFvTbToOBnjfpmP+kbm+i2ruas/4IObKmwBql9q4aqWTKbKfflqcdpmEe82+VPf86a/HaV9vUvnqW/oQcMftX21XH6gTV5R2p9C203jCJ+d1dtfYYKf27ePoPIUePjAp01NtV96phqstnl0rghQg2/3vQDzpeOC4mGJ7PwBtISK/7A+pRAdSPUh3heQLjYawqbV19ZPX2r4Ym7NP6LRI0EUL6qk+gNkF9UTNaobSWELDBbAug1ggw/2fQ/UFfwindvgS1b5VNcakvIf0RAmuU6loxViut9622y4nC0/RThQdj2vB+bSdjv+2SLv+IITJ56fOaUWpvDY/8JqRf0cbV5VDRcJNA5gr8TN9WfCCCoBv6AFoI04/trLVpMxc7OcmHkIFtEuzzac0BreSGIU+Efqp3Eu7upym9r0NzoZTQhwiiOpvqWJWHK0t+QgapGkjQDxexqQM1Egb2RP3Yrd/rVcHuo9xt+IBu+rehh4JemxsVx5jPVg6VyUmqyz9Uda0bpkIVvAeEDDpak8Bv75lG/KnvxJigzz4GX0+GrocehOqgN6Emfby+Re9I/Ahqht7RZx16IFmHcDaMP67TPzze+LnS7Oh4aSSNT0ik0oGKQBuO6zGtXYXuPkZ1eEccq4dCBioHJ66T9IBw4fsOmHssyvl/66gdqk3gCa9HsOmrE9nGJ8Sc/QV4Wu95fZTqqB0hTQQ9GBKMjUj1P4c7FYP1w9WuJwq0EQdCCCGEEEIIIYQQQgghhBBCCCGm438B2y7hUHxPW7YAAAAASUVORK5CYII='

# Decodifica o código base64 para dados binários
dados_imagem_icon_luzitic = base64.b64decode(codigo_base64_icon_luzitic)
buffer_icon_luzitic = BytesIO(dados_imagem_icon_luzitic)
imagem_pil_icon_luzitic = Image.open(buffer_icon_luzitic)

# Converte a imagem para o formato .ico
icone_temp = tempfile.NamedTemporaryFile(suffix='.ico', delete=False)
imagem_pil_icon_luzitic.save(icone_temp.name, format='ICO')

# Define o ícone da janela
window.iconbitmap(default=icone_temp.name)

# Definir o título da janela
window.title('Importação de Despesas')
# Cria um código base64 de exemplo
codigo_base64_isaac = 'iVBORw0KGgoAAAANSUhEUgAAAQkAAAAxCAYAAADJNlwQAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsMAAA7DAcdvqGQAAB7qSURBVHhe7Z0JnFxVlcbvfbV0dxKydGeBrOw7atiUZRAQXBAQEAXZUYFBcR9AR3TEFWT8oY6oBBGFkV0EZJVVVBAGBVmGRSQknRCQdHfCknTX8u58/1d9i1cvr7orncSE8X2/fOlX9ba7nnvOuefesmbNYUNxF/Et4haDnyeIk8QOsV/sFl8UnxWfEB8U7xdfETNkyPD/EG8VzxX/JroRsiTeJX5CnCxmyJDhDY6i+BHxETGt068KERiXiTuKGTJkeIMhEI8R54lpHXx18zpxKzFDhgxvANBZ/yCmdeY1STSLb4htYoYMGdZR/Ku4TEzrxP8o/kXEGZohQ4Z1CPgefiKmddq1waXie8QMGTKsA2DK8hYxrbOuTZbFo8UMGTKsRWD/3yqmddJ1gVXxSDFDhgxrAMMFU3H+EnG1dkI91HXZwK1vAjPJWjNef6fqM9Iop7MTrbVFY60kgFvsQtNvnKvo3PMutItNaBbquwU6Xm6cTz8aBabHHdGnDBkyrDYMJyT+TTyndjgyME+6jc2HOwd58yabN1vYnNlY7DCWU02hXi/B4EigLRjrcom0IkDmu6p7wlXNI7ryMVfpMWFup7tMP1OyqwIiRN9UO4xAZOiVtcMMLWK0+P7aYUsIRcp5kfiAiNDPsI4g2fGCbjN2vGtrm3BzZWCTy9zy6zWiF7p02Xoa8dv1l9q7IRxgJB+yk0+3QXhs0G4OCdrMZBukXout8KyrVh90FatObv6mDr9IWoK0B7fUuBzngW5G8wgnSruYbnNuc5uzsyVwdpLg6TTxZ7vlztm79ffq18r2qq1Mz0jCu88WT6sdRvi7OKV2mKFFzBRHKqwp78+Il0afMqx1NAiJ+WZspynk31cy9ojAmncUmmgadN5HXaV6Q1iyt4QlM9dV6x21qNOn50aFx+c6gmb3SwC4n1X7wyvCfts9jLAZCjx8e2kph+XaEEa2TUKkdiZK41L9f261tN7ZG5l5jFKtIhMSq45VERIAzWJXkXU8GdYyUjuxcPI4Y897d1B0B0gT2DUoSOVveq152lVDhMX10jBOzHW4DwZtsg7Scb8rhydVXrU9Lmz6vJFgls2Fc/JjzNY23yB0nHEPLyuV9t3SvLp48KvhcJZ4eu0wQjMhsb64uaiiSQWylHc+JtLoh8JG4vjaYVTO3Dc/+tQ63izG8z5XXFI7HBLTxU1E6ixZJ+RhQPxfsY8vWkSakGA9D9PWyXfwmcV/Pv8ePxZPrh1GZbyd6JXLNMSf2yP697P+Z5qYfG8S5JO6puyT7/HPGArcw3sxmXChpWE9cWsRcyytrCGOeMpqoZiGDUTaHXXNM/xzfJr5S12xTCKt3RHKQBpYbJk2QPs08H6/BmsFtItcwMmI440NDw/aqr/Ij60+W+wKu4tdrhnnp3zn+XSxM+RZ8WevTk6USfKU3rHCuwud39b5VoEmEX8uq1TjoFIuElvNBx1sqAZGJ31ZjN/DO6nEVvE+MX4/vF0cCruJpC15Xxpp9OeJw3U0j1li8hnvFJthlPiMGL9eZmMddIz4ueH4c9Hjc2LaNc3IyuTPinHgm0u7No2viReK8ahgOuP3RSKG0+5J4+/FqaIHZc9zV6XdnSQycKRdn0aExC5pkoSZjHji9FRnLw8HgiMrLwezS73mc5VXw7vDcohzcfCSOoZqRcudw9fwqg6JmLxWPF88U6RSeO+HRBr8O8S9xfcOfsc57NSvixeL94graAbLnEv1eIXGLh88XB3YUzxO9FlldHxY/HOMfxU9CGH/ZO0wFayaZYSJg5GLvLYCRqXv1Q4bQBl+oHa4AhgIrhFbXQuDlvExcY/o0+oHEbyP1w7rWFsmHkL7O+Lx0acaWhWOAIH3YZFFjx5HiKxqbqZ1pgEh/p+1wwiUPc/1aaGzD9fueKcHEco/FMdFn1rDxuKF5vli51bdbV17dxc7DxU/eWqu44ktbA51IylVViBawWHSMC5uQcPw1HWn6t5mQPqiejLDsJf4bvFt4rYi+1AkhRoq/7vEL0+3wWOXKR3J980vdN61yEyhI7WK4TSJj4r+HCNssoN7fFd8dJDNOjz588+izBlt4s+eLQ4HzCN/z/PiVbHPmCxjxCR2F/018Kciy/zRAGYMEhOIxh1vC4Tkt4KV1SSoR9T0+PV3ih5JTYKZk52GII3bI65JoCWQz50TxP/xeRHflb/216JHXJOgPdCBk9xf9KYKjDte45HK1OunROrWlzXEBDhIfE7018ZNDsref0+dNGt3DDq+3f0HXwwCAePvhz8S6VvxOseMYXD5nRhdpw631P591MQNBkruQGfdAdYyeluiK808Vw1vDkvmvrB8952u/Et9xSCNFkADY4Si09Yx1lj3zqDo9g+KZveg0OBETEJvv/gpWzlz34Gl2MJvF3cQKSQSOpTUphKfFpGg90ok/uaR/IRpQRAcq6dK26ilHfS68JUuF5w2rdJzgR5IobaK4RyXaDbxBoBaTJoA76GcMB/QXuB94h/FJBjNqchNo081VfIEkef5EZt7aYBUWBooM8rCj1BoXLeJT4qdfCGQHzpAHAhWImg9EAg0zjTsKzI6Ui8Pia04JGl4yeehhqcpeoAGn/RjoWV6oeSFiMfXxJdqhw3wbYf68ZomQsKPyNTTUGt+fiEiGAExN/vUDhtCARA0+FzSQJlStoAtDvyzmEL3Wh1aMG2+GWh71Bmgv3lhcJh4ee0wQlq7Q6ulj6S1O9rAt2qH0XNpH83qg/R5c6/Rr9VlglPfHRQrXwlG3/RgcdxB3fnOty5qG7fRotFTJi8sdm25ID9hF7VWGmRexBxg5KPRRFJnkBecZjqmdxe6juoudF4n7WR5cnSHc6V5/CS/XlWCpaoHtmRnjTG2upPNV0/ItVfn6N5HihMaNYdC5zPiD9CMVAKojSPBcJoEwhEHVfya4fgVMQnMLH+e53mhS8eP266MAGmgQ1CR/rrfDn4HEDb+exxyW4px0JD9eTjSsmqGNE1iZYlZ54GQSLumGdFEPeKaBMJzKMwR/bVxn05ck+gVT0nhF8X4wkfqwCOu3f2KL4YAZp2/Nj6Fjwm6su2ONuaBkPDfN3OKerB/i7+21zcqjxtE/ADg5Vk2mHtK0NF2YNBWGGUtoz7oc8bdqBuvXVrK3bqteYkRAk0ASYlayB4TdSw2nWOXFyx+hiPUhJHMCJgG9JowvK5aMteEA+ZhV0nzk5gzc6Mrx+fa8z7BS4wLnwor5nFXNQ+5ivmLKz8814X/olNU1KogqUkwYiV3yKJTHSqinpEfkkW6GQ3x0qPio/aOFQESvkv0Gg334X32zi124mK/DCoFYM/6gC7ez+iXnF2gnOMOOjQRNBOeQTq+KnpTgwaPRuCR1CQYGRkh04BKiibBczFfcGYNhzRNolXQMVCTUZs9kpoE/izKtBm+LPrZobgm8ZSYFJhxXCBiToJmmkQrQJVHcPjZBYQE7QXcKGKaNAN+BAZfENckgG93m4kM1rQ53/ZwdHMtHdy3OzRavicdcU0CLavBEkgAE8xPP68wq8Wo6SVInEs3s8EfvpYfdedThc7fa9QeqI3enctk81+3oNB53EIzZqKuGxIb2dxlJ+c6qncXxq/gO/DkHNdMsgEdqp6GfaRxHB20V/cICtUpiXMxxm2wkSKpSWBuxEEFs5fGvWK8oyVBJ44/x1cc+I0YPzccmVmIg4pvVlfN+EHRg8YfP0eDTgMObBqJv67VjpKmSaBNYXM3IwMJjbNuMsaQ1CQQsq2iVU2CgYB9Vv21mG0eKzO7AVHjvdkBMD38OdT3ZoIKoY4vxl+L1uKxn4i2iBmxMu3O++Pwg8S/bxYRy6CHP81f16BJMAK2Mhf+Sqexv/1orv2V44L24hgb7KmHdOlpVePcPdaaa4PAXTutvy9tnp+ZiWjV5g4yG47Itdv9g6IdleK/UCm7W8OSu7jab/7oygGpbQHYsNiyq4JknERSk8DrjaPPAxsTh2EyifgSvO2Kyo+Up/EkbUtGSK9hxEFnQfsAnEczwScAGKm8vY4jjGcky5D0UKd+JFog4vFmdKLzYybGtTrq3k99AxoX6Y9fwwwTDWg4pGkSdBqE40iQ1CSIAaFMm4Fy8v6AuCZBOaa1cUZihHg8r8wEfLx22KBJUI5eq46D+/EB+HO80zvp/11ksyQPyvgF0ZsP1B31TXnHZ0AQCt7sQsNB0/HAuRivL49m7Q73QHJtE2URb7vkn7qLC+p4uUdS7FiRzPkgiuH4sp56xYdt+7mPFSacN6/Y9ai0gUekXVzbXeg6er6ZnhwVSAA7Y9efMcbY8EhpCDcWxjXVLu6QdoEWMbp5jAUSF0mZxLDaTQpQyeLPTmoSzD2vrG3o4zRoSL5SIKrnCgJyEHTSeD2gudCYGW3jmlTcNEqCAKu4f8M7xACdp5lGlkbiGFotTxzQyfuHmt0YDisbJ4Gm5xHXJFolnS8+QxLXJIYKcmPA8Nd5wQQQ9mgx/lwrRJjHHZwIdgastGubMR4fRNu5Xky7rhlpH5+KN1A/4nhnCTYxoz4OmFbmVskUPg3sr5vFZrEJPOsMkcAO/84ILAQ7Ktdm3he02fVStItXpapcUx1wPwv7zV9dlQwQcMKojDeb93tsI2LbYU/HnVitgMYcd5pRHt6W82C2A98NvoKkVz4O/COMBl6Cow0cXDuMwGjFyNQMTNcdKPqyIAqRjs+UMKCMfyCiTTQDqryPh2BkQRPwtjLlhBo7VEwC9xDXQt226u9BGCZnVNC+EDQjAQKTNgNovMMBLQYnJGBk9X62ocBzsdW5F40HX5sHgpmpQeoBO58yTwN9hZkigMCImwXMZpEOpj45TsLni7qhnBhAGFDiQKPlGZgrw7U73p8MqCP9aBRM+Q7Vp0kLghKT6/F4R6RjkCkitVBTUfFwDpFQ7F8cW9iCOAeHSiCgw5JJBMZNYprAoCFhJ2NDMa1aT4vMD3dgUHRH5drNmxNh1oAcLHHhA+Nt8K0ZpZ5f60bSQ4fC1qZTecFAYePkyZAhw2oAdgwedjo0/RAyB4+XHI+p78R4RbGPGFni1zYjIzHzz+z30Ey4oNp9SWTet+H+rW2+elZ+dL/MjiVJMwTOLXS+cGqu45XJ6c5MBF6GDBlWAXFNwgMTAA8+3k86NtNfAI2CKDSEA+ozAoJruQYVmr/DmSU8479F1j6keZpJD+ohsQFoGd4zOzDa2CcOCIrbHZlrt29J0S5wdN4SyhSpDpgHXNmfZ28CNIyVAZpNPO4+w7qHtHabYc1gYLjCppNisyIwsIX8vDs2D/YOQgOzAg8oq8uwd5jKgjibhgLOJeb2iUaL238eCBwcqQSXNETJSbsIjx7Cd/Gkq4YXVfvNjeHA1UuNYzZhZYBd6mcVMmT4Z0dfsoMx9UcUHyMw6noczFSw1oBgDjQNfAqA6/5HxHOKucAMBuYDgiUZhJQGHEHMIzO98ye+SIA04jRCWBwg1qep8F1Iu3DSLszsdN/FEuvMT9vL5iuTWt+ABs0jPg2VIcM/M1ZwijNLQKdnmoeIN1T/FTqfgDpOh/2ZyPQj96wOIiROFBtmPWJgao35ZuaYG+5lUdoZuVHVPxdqodp/1N/fF8b/eEGxa6hY/QwZMowALDhi+grtgOkYpkL+S2T2I83xyKiLP4KVbis7j9uMjPoERRHunQZMGxYzYbI03Cs1I9xSAkPqBxJwJHESGTJkWAmwbJtAEkJBERiM4ET7pa7BEPiONQLM569s2HAzPigOpV0w74ypgl8jfh8RaRkyZFgDaObZR9X/tEjnY7oRrYHOiZ8izYZH60D7IPAkHmU4UuK7GEq7YD0DYcN+GhUfRoYMGVYRabMbdK6jRAKh2EciLQyVEFGmPXFiElxFB8ZxebVIlFYyrh6/Bv4NnJmHiAicVQG78BBRhw8l6ZAkT0RNsoqtlT0e63jabNrWUexLXV1og3CzatXuFVhzvnNRfnmT1b+C/h4yY6CHlZwRWCLvrN2tXMhts8lrf68vNZ9f7DrNGnd4LggPntq/JNqXYUFh4uHOus/MKPVEU7Usc5eIU7lHP0MSwTnbUy7mduRZ8zq6pgZV86wz7mXrTIg01PvH6O8Zs0q932WJvp73IxWCBLevXqck24dnlnp24ZPSd4xORfnQFXq8UbZszobh/mFg36HjM/ha98fbxy+Uxo8oD0S5bqvThHvXYM3VM0u9HxvMy3lKG+Ze9AShr1AubL+BeXGFGaz5ha7b9Aa1n+hCEqmyt1dUy2NO95sXdxc7Vb+2rJOsP4gQBPar0wd6kovezILixM1DEz4knXevmZVenO8R2D5Aadx2Rqk3iqSdX+w8X+k7Qq99jUzynbX2uXodFDsfUHltbp3rjwrB2na9/88zy71shGQWtHV9PHTubN0ozTnKqf7ZF4LywI7TYnupzi90Xq97d6vm7bYbLltcXwMxv63zE9bZcwbzLVh+WuaefFD91AYDS1mX4tPIdLz+DL4Dbd6Z82aUe6M1IboGrf5IfVcLVuTnapx7qaNsd8BRv6DQeazu+qaub9jGTvddrCcelLfVN/v3AdqOXvU91R8b0FIfvHphmlOSRVhUCLHnhKjS2TA5fLgpQDOgkig0BMYXRBLCWnnWOhALQTiwX7uBqYIGgibC7AeNlS3CRrqceHsRk4Z04Ash/NqDrPGLYyslIMDm5pmBUimcAvWUOSqkX/vPF/T3LVTB5dUYetRYpkQs9U7Wy55UcdZ/p6O7OAH/DJ3ovmK5Gl/Uo/Lm50PM1tVqcD97c/Cdaj1QpeBjiaB3Km/uJjXoUZ4mcGd3vBZEodR6Bi2mzeXsW1T56ysd66vBXaQHD2pY7m1K+2+nl3o71OjbYUWdRjdRZhGU5pz+W0IeomcoH3xW95ut469zT75UwJ/Tphe+Z/A50XZsek7BheYLundynRIQtee6XZX+S3ie0h2dkxD4YagMcD4Ja11BN80plZzK2K1ftZZOfEC+8CrtaRC2TRk/lrR6pgmIGqKyCfXgq54360dLoRcWJn5IAvskPaeu7UoIqh7NHYPlFz1T+avH00hAtKmcz/TndcPB6uxEBUeQZN5DT/ml8hiVr8p6rN48xhXb/UY/EgQT3qVrtlOF/iFXcQ07k6mM8iqrB2p11NvuSuWNdW2l4vLxhX87Kx1f8u/gr+r5XOWNkGqPXFTeg3nIlYqbKZ2bLWsLollFlVteea+3LTA330n7wCVwYznMx9eXUH+zNfD8WPU3ifoj/6q4i9OEhF9Ywtp11j8QE8HKSBZm4R/AqRkPdUYoYAbwYkK52a+SOIMrRM6RcTQOHxhFg2G1HIKHqVLWM7DwaCRx/cRt0HjxmbBHIs8kDSPGJqZvKVQyB9SIS/7zmSpzFZZGXTceKT/IOWqVm6ufR5JcqoRGluAcZ4Iz1OBPVUaP0ejasAWdKv46dZwfuMDeEUnu2qK1OgIXEv5+iEahRRoBn+evDe3Jtm25j1GJEFTdOd3Frp/MK3ZdqPe8V4zWrmgIYPiPRoLoQiFnLOZhHVS6GvXY1/PRNUcdYaLyVl//UjaFKF3SCRrSJ6FmbWC+JS2gW/c17LilBvqcse6jpNtTXx9TMssaGmoj3IAv441KPU8q8Xco4bSJOpSXi1RWC8jv4FdDwD6ifNxQKZQvXVSYsG3VuLNUa9+pF4aHNe+M8gD5KYkEVE8f8OWjTJ+mDlQbrQdBGQ8eclySljZxemlxtFNU1A5cwCD4RZUJo/5RElZ1IT3Y6dC2omTNNC/3Ks2XJPOtdzbUm164QjYCa9/u01kpDkT7UARBNXWFrG62ucCeK03sa2oUp6p97yfNNdKOIlj7Qmjdp33bkyayIDD2kDQh4YF6FNcWWJrM0lK2DqMwCNlm0U18bTyrIwmQYjSls7JPAcIBzQKBgTnC9m/eCUmm44KHDsXoO9wOQmlgRyeW8zLrsQZhy0FoH1LCt1R//IBU0lOXl8dHm79slu86Ud8rHeGJEvEqO1dyNoxvngKsRpCvS109Tg3jfHWqU+Id2pQrt6s17FkNqruIu/FXjSVfdQUWY9Wh+69ThV6u+y8LAnNCR+n15e00hsHDVNC6lfYK+dARW9Ufoud9cmm5Nz6SpUKNhrR+o1qubFcuhfE9E0w5by/Nh3bPSlDddTDtMjHthvl8bqiNXuppvTLyY7kt9Iq6ys7L9M7jSuVwm3wpn7bStwGU5YxS52fVATor1t6rTvTxwNl5yQLRCHwXeYAzzMsrLh93Rpqjnav7jlb5lwOXa7oJsNIYSNhchSnIZ7UDli1srXycoHL9IfeHNvyur5cGqTsIdcStlah6vluFhPZiPfVRJeIgPXwrSZ59ZizvTV002C1zUNe8TXbE4a7ChkW2X8fn1spdnTrnLg1csHvh9ba3m8qpK1l2rQAtAVMC7YAIS0YJRnF8GAiB5K7HgMAr4iq4h4bFe1lpx/VEbaaZBqx05HrY0EGGAHYovo+hVkW2BEnSc5TKqVK76kJHEvZ4Jf2rM8s9M1gG7wrLr9HIN1W25IEd/W7pK8XcXzXqn6Eij0YUqaVjpLZersweq+dcLVv3CxqhpNL3RBvAzM937aTGfJ3es0jnI3NBo/PPdf32qvRo52NGbjX0vXXd+6cP9N4+v6Nzmq3aBcV8sMGUZS8x29QApVFanz1CmsLtek6tfp2RyuzeLjUy0ubYJEgN9SypqOsvNBuMqhZKv9Lz+XnWg6f390U7VJE/W1y+TNbpnjMri1nJGkGj+f1q7LTzuiCX+fHArErPj9RRbtJzZLrY1/1Y1uyXd+EOG5T7VmgXKo+7dQFmzQNqjLQROtZEqWN7Ti31Rutu9L4Bpf9O5aeeV+fCW2aV+9BUG0BMjBJ2ocp39+7ixM3UOQ/W8bfn5btOlvZztI4jVV3PvFDP3EdpQ2uJykidZdmscm+0f4Tq4GHV289VJ+cO+oh+pc/fnlbq/aYuZs3QBbphL4289wzej/Z2rDr6VqVS5YVcMfe0yu1LwWA7qDo7KjDhlYFzx08v912pfH9a+f287rpZz6X/zFC7kBngTlKdsGSBNNynczl9Hy+32UrLS2p/aO34FuZIZJdkFpwyt9i1pS6+Vam5a3ppwkmBTGe1hY/oHWcp35Oi+iwsf8La8PtqUdG+JGpbBQnQy1Vfp88oL77gmULXbUVMYmsjf5nypCSYfaICGgKEV9NZcWCmje7MKLDqko5MwnG2cR2dH7LEOAnUZkK8uQeNA1uREG+uZ4FZfDceD0YiLzBYKp0GnIksRItvKz5iqGHtqJFklDoIS24jzGubvHHgKtvT4fn8mATk2ELXoSYIe50pPB0/54E6R2VHHbwwXjZqvjPe6WqOSDt7Rmkx4e3RZ1sxewaqwOgCQabLElU0q13DuWZWe6HwymFhedSVM82CBhUYPGLGTVhesPstcmH9x25UKZWNTf6hLQadefPbJm3qXHX2rFIvgj1y2LYX+g4NAvci6eQ7Nf7cwkLnUf3Fws2bvvZifU8Ndb73Whc2xJ+ELniOPD3XNm6jXJjfTY0vGpmABM+LMwf6UndSwm6XNhOZh6E0m9DapdVS9XeYHtEFAvuSqPwaNF417MenVnrQQBvwNzNhXLFodp1R6mOrgjoQHs6EG/vvn8tP2C1vrd+AOEJogtLM8mIif41MuAMlqJ6eJvOHz9RbYILtw7y5feby3oUvjJ48pVoK99Yzi94eUD5fUD5vnds2bsO8y+2YbAdz1Q7UAfPTBnpve67YuXXemZ24VxmTHDCvVm31oQ1jTkScsBJcO9N2uI4CjcooqNzrr6vtOWvDGZXeaKs52k6u4vZV+f9pZrnvsRfbJm0yELqdlK/LGVxylWCPqTpWO1L11kBZmCA3fkO1v6h9h9VdeSfndBG/r9ew6UwaqBz2H0Ad81t0NxvVGQkYdfk9B/wYJARJ+k2RUTJNILF4jBkPKodOjsccpyPr8pvtwZcM9uI9aA5oKhkyZFhLQCVioRcbh+B3YC9AfBPNNnRBW2BhFYu3cIbRkdlliZ1y2MAjTWCwEQeayyUiQomOz2pTfuqtmTOSGRd2ZmLKNkOGDOsIUIOJQyBOATUU9Z6dm+J7TsThtQViGtAWEBhMfeL9xUZsUCUHgVDCJMEJilBC40JFZ7Ylcg5lyJDhjQHCsFmhSVAHjiVsJeZeiVtopi3gwyAWA20BgYGjjJ+pIygrTWAglHB2Et1JlCcmBvs9xn9GLUOGDG8A4OzwYdgsCsPDjT+DIJRm2oI3YZj6QWDEg7Tqzq8YEEqsG0ELWdm9KzNkyPAPBAu/2IWYfSXS1nogFNAMEBJoCnR+hMdQK0kxYdAWMGEQGIQz+yCtuqc/Q4YMbxwwqhMshU8CZyNOx7RdgDE74mHYdH5CqodaSYoJw07STL8gMNA01nCQVIYMGdYUmMlg9SVh2zgmcVCyiMvvjRkHAoPZDR+GTecn1JaVpJgfSaB1ECK+pn7yPkOGDP9AYGJgGjAjgROTqU+i4viNUL9uIwkfhk08BUFU7HRFgFWGDBnWEaTNQKwO8FxWRjLFyV+26yfMlig5lpSvECko4IgkUOp3Yn15dYYMGdYmjPk/JIjDot6atdMAAAAASUVORK5CYII='
codigo_base64_luzitic = 'iVBORw0KGgoAAAANSUhEUgAAAGQAAAApCAYAAADDJIzmAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAALEoAACxKAXd6dE0AAAZxSURBVHhe7Zh7bBRFHMdndu8BbS3vSilRkJcWaTGSYAhR4S9FCtGAEQ2pBh9FRGzLUaTRUlMI9O5aoIRoSawSUxJAidgCvoKa1EcQiAgV2qCGQLCIWFoKvbvdGb+zN3d9knDXrhgzn2SyM9+ZnZ2Z38xvZpYoFAqFQqFQKBQKxb8MlU/bOD/3/gTd0TJf07QUztlXt+9t/ElmKXrBVoNcWjAhjTHyJT4zSaQ5JyanvGDkR41+q4CiB5p82oJpEk/EGAJKia5xsr5pYfpIKSm6YatBNELvk9EOKHURw5gqU4pu2GoQRshpGY3CCTeY3lNXhLHVIEGqe2GBJpkU1jARtqTuafhNKopu2H7K+nPepFHMwbIp4an43MEReuOndDcMo+gV2w1iFx6P5zaTcx+W+A6fz1cn5T6Tl5c3A48nQqFQYUVFRSCsdrB8+fIRLperyOFw+DZu3Pi7lPsNW12WnbS3tw/AbHoa0XFhpX/glD6Eg8cziYmJCSKdn5//St7KlaVWJoAxxiB/MQyWIaV+pU8Gufv9+mFTqk69llnV4M94tyGL7Nqly6yYWPNJZlphTfrqwtr00tdrJs+S8i3h3NmzpZyxyRs2bPjbEiidC8PfYcWB3+8/zExzbFlZ2T4p9Stxu6zM906nMU7rUMGdlsBxDyd8+4nnJuWgE9zSboJVe6dMcbnMr9GUISKNyyMmKS8qeay+RCSF1huW63C7f8X3l8Fl7ZCyRUFBwRjDNNdj4Dzl5eXnhZabm5um6boXg/0WY2yY7nAsswp3glH6gcZYO6JPBoPBVVgNJejLAnQtRCmtQ2PqKOe7oW1CPWthlAaUpaj7QU3Tnoc+GWVaMcvfQZuqrUpjJO4VYjKyMmoMAVqMe8cLU6tOT5OKBTpD4ZeXYdnvQ8PHSzmK02GujRhDgMujsObqgs/vSpZSzAQCgaF4LHI6ndE6MGCD8FiEZqZwl+sy4/xYJKCRg9DO+TBGE/LHYUY8jvIatO/wzmWEi+jrIej1iCeKepA9Ak+CfmXD0PvR5uuopxAdeNsg5JLIi4f4DIJBxsDdI1MdQDWoNlGmLGCMIVTT1qNDWehEoZSjoJ57ZTQKbJLgNAeOkcl+Z3Np6S/lfr9XBAzAATRiBtr3JtzREVlE7BUmVsBORM+hW2cw47eXe72HwrlhioqKEjAWZQhby/z+l/D+gTKvd+cmn+8zWSRm4jOI5ZLoUZnqBLwW1X+WCYu2trZWNLgGQZxI9oTVztDDMtIB51d0Nz8jU7aBFTsUffkQnTmIwSyT8k3T0tIyHcYajIm2GUlU03fidlkhLcEPdy+WsNUQ+H4DV3PviWfHHRfpCJWVlaGkpKTF8Nvp6HStlKNwjRThZcvPC4S/hvZG8az6q1KKG9M0MfF7B7NbeKWtaLzh1PUcSDEPKN5PQns5nm1S6jNxG+RU9ui/giZ7ABvnYqyX1YxqDx8/W71GZnehuLiYYXO9LpNdWDfn5BnDzTIxU5fAqB6q8enr5tRXyOy4wABZxjQpHW0JYbr80Lx67doKfPNRbNILcZ+4IuWeiFMGIb2eHk2n8yRWCDUYmyelPnPDGfRfJ3LKwmD5cDqKrjwH50GsjEbsWw0Y8Aad0hyszmSUq0RnM1FkNgyGQxITfn4TxlPsExEuImQxQkrw3niv19ua7/FUwSgz8e4juAxexv1jMOoVp7uZ4kKKTX038mdpmFD47o/E7R4oVuYWr7cxXGVsxL1CbjU4QQlv0YLByMXp6ItIwKmpOjk5mWPAchAyTMaO4AgiBq0Uz3NM03QM2MswUABhKcp/Ewmo70WEIMq2ys/gLMy2oVwy8o8GDeNVoaDMFdRh/f4JBQJLYKAalNmB09ZxzTB+cJjmUyIvHtBme2mvHT5RJ/pSdCIV3vaAK/WPajqNhGR2X7lR+zE+1qY9EKshBXvYBbjNICRR3soDvb3bOS8St+rBBBjW3NzcJPZESF3yBbjRJ6KPQ3VdbxYrS8oxc6MO9QuBj1MyiI7LIyVJIi16gEZXuedeXIIPd+mQIoytLovqJD9iDIGwPnx2Nq8dNSGsKLpjq0HgyHtc7mAUzWShsTKp6Ibdm/q38hkFLuuagzuiN2JFV2w1SMDU/dgpjoWP8mLT4O1YIivovAtx/+v5v2Prpi5g+8e7Td48m3E6HKvj+wFZl+I6nysUCoVCoVAoFAqFwg4I+QfWBJ64HHCcOQAAAABJRU5ErkJggg=='

# Decodifica o código base64 para dados binários
dados_imagem = base64.b64decode(codigo_base64_isaac)

# Cria um objeto PhotoImage com os dados da imagem
imagem = ImageTk.PhotoImage(data=dados_imagem)

# Decodifica o código base64 para dados binários
dados_imagem2 = base64.b64decode(codigo_base64_luzitic)

# Cria um objeto PhotoImage com os dados da imagem
imagem2 = ImageTk.PhotoImage(data=dados_imagem2)

# Exibe a imagem no canvas
entry_bg_3 = canvas.create_image(
    700, 50, image=imagem)

entry_bg_4 = canvas.create_image(
    800, 500, image=imagem2)

window.resizable(False, False)
window.mainloop()